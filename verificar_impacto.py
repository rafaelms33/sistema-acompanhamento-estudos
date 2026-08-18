"""
Confere, ANTES de importar, se a nova planilha desloca alguma tarefa que já
tem histórico de estudo registrado.

Uso:
    python verificar_impacto.py planilha_com_simulados.xlsx [banco.db]

O importador casa as linhas pelo número da tarefa (ON CONFLICT(numero)).
Se um número passar a descrever outro conteúdo, o histórico preso a ele fica
atribuído à tarefa errada — sem erro nenhum, silenciosamente.

Este script compara o conteúdo que cada número tem HOJE no banco com o que
terá DEPOIS da importação, e só reclama quando a tarefa afetada tem estudo
registrado. Ele apenas lê os dados; nada é alterado.
"""

import sys
import sqlite3
import unicodedata
import re
from pathlib import Path

import openpyxl

COL_TAREFA, COL_DISCIPLINA, COL_CONTEUDO = 5, 7, 14
PRIMEIRA_LINHA = 4


def normalizar(texto):
    """Compara textos ignorando acentos, caixa e espaços repetidos."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = t.encode("ascii", "ignore").decode()
    t = re.sub(r"\s+", " ", t).strip().lower()
    # aspas tipográficas viram aspas simples
    return t.replace("\u201c", '"').replace("\u201d", '"')


def ler_planilha(caminho):
    ws = openpyxl.load_workbook(caminho, data_only=True)["Tarefas"]
    novo = {}
    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        numero = ws.cell(r, COL_TAREFA).value
        if numero is None:
            continue
        novo[int(numero)] = (
            ws.cell(r, COL_DISCIPLINA).value,
            ws.cell(r, COL_CONTEUDO).value,
        )
    return novo


def ler_banco(caminho_db):
    conn = sqlite3.connect(caminho_db)
    atual, historico = {}, {}
    for numero, disciplina, conteudo in conn.execute("""
        SELECT t.numero, d.nome, t.conteudo
        FROM tarefas t JOIN disciplinas d ON d.id = t.disciplina_id
    """):
        atual[int(numero)] = (disciplina, conteudo)

    for numero, n_exec, horas, questoes in conn.execute("""
        SELECT t.numero, COUNT(*), COALESCE(SUM(e.ch_efetiva),0),
               COALESCE(SUM(e.qtd_questoes_feitas),0)
        FROM execucoes e JOIN tarefas t ON t.id = e.tarefa_id
        WHERE e.ch_efetiva > 0 OR e.qtd_questoes_feitas > 0
           OR e.status <> 'NAO_INICIADA'
        GROUP BY t.numero
    """):
        historico[int(numero)] = (n_exec, horas, questoes)
    conn.close()
    return atual, historico


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    planilha = Path(sys.argv[1])
    banco = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("trilha_tjs_ajaa.db")

    for arq in (planilha, banco):
        if not arq.exists():
            print(f"Arquivo não encontrado: {arq}")
            sys.exit(1)

    novo = ler_planilha(planilha)
    atual, historico = ler_banco(banco)

    print(f"Banco:    {len(atual)} tarefas, {len(historico)} com histórico")
    print(f"Planilha: {len(novo)} tarefas\n")

    deslocadas, criticas = [], []
    for numero, (disc_atual, cont_atual) in sorted(atual.items()):
        if numero not in novo:
            continue
        disc_novo, cont_novo = novo[numero]
        mudou = (normalizar(disc_atual) != normalizar(disc_novo)
                 or normalizar(cont_atual) != normalizar(cont_novo))
        if not mudou:
            continue
        deslocadas.append(numero)
        if numero in historico:
            criticas.append((numero, disc_atual, disc_novo, historico[numero]))

    novas = sorted(set(novo) - set(atual))
    print(f"Tarefas que mudam de conteúdo: {len(deslocadas)}")
    print(f"Tarefas novas na planilha:     {len(novas)}")
    if novas:
        print(f"  (números {novas[0]} a {novas[-1]})")

    print()
    if not criticas:
        print("=" * 62)
        print(" SEGURO — nenhuma tarefa com histórico muda de conteúdo.")
        print(" Pode importar.")
        print("=" * 62)
        return 0

    print("=" * 62)
    print(f" ATENÇÃO — {len(criticas)} tarefa(s) COM HISTÓRICO mudam de conteúdo.")
    print(" Importar assim vai atribuir estudo já feito à tarefa errada.")
    print("=" * 62)
    for numero, da, dn, (n_exec, horas, questoes) in criticas:
        print(f"\n  Tarefa {numero}:  {da}  ->  {dn}")
        print(f"    {n_exec} execução(ões), {horas:.2f}h, {questoes} questões em risco")
    print("\n  O que fazer: gere os simulados a partir da MESMA planilha original")
    print("  usada da última vez, ou coloque os novos simulados apenas em")
    print("  posições acima da última tarefa já estudada.")
    return 1


if __name__ == "__main__":
    sys.exit(main())