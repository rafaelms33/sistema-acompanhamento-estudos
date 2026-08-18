"""
Insere simulados na planilha de referência, de forma determinística e repetível.

Uso:
    python inserir_simulados.py planilha_original.xlsx [saida.xlsx]

O que faz:
  1. Lê a aba "Tarefas" da planilha original (numeração 1..N do fornecedor).
  2. Insere um simulado nas posições 50, 60, 70, ... e renumera tudo.
  3. Cria/atualiza a aba "Simulados" e a registra na aba "Estatísticas".
  4. Corrige defeitos de exportação do Google Sheets que quebram o Excel.

Por que é seguro rodar de novo a cada atualização da planilha:
  a posição de cada simulado é contada a partir do início da lista, então
  enquanto o fornecedor apenas ACRESCENTAR tarefas no fim, o número final de
  toda tarefa já existente permanece o mesmo. Rode `verificar_impacto.py`
  antes de importar para confirmar isso com os dados reais do seu banco.
"""

import sys
from copy import copy
from pathlib import Path

import openpyxl

# ── Parâmetros da regra de simulados ──
PRIMEIRO_SIMULADO = 50   # posição do primeiro simulado
INTERVALO         = 10   # um simulado a cada N posições

LINHA_CABECALHO = 3      # aba Tarefas: cabeçalho na linha 3, dados a partir da 4
PRIMEIRA_LINHA  = 4
COL_TAREFA, COL_TRILHA, COL_DISCIPLINA = 5, 6, 7
COL_SEQ, COL_AULA, COL_QTD             = 8, 9, 10
COL_DESEMPENHO, COL_TIPO, COL_CONTEUDO = 12, 13, 14
N_COLS = 14

TEXTO_SIMULADO = ("Simulado {n} — resolução de bateria de questões "
                  "abrangendo as disciplinas estudadas até esta etapa da trilha.")


def ler_tarefas(ws):
    """Extrai as linhas de dados da aba Tarefas, ignorando linhas em branco."""
    linhas = []
    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        if ws.cell(r, COL_TAREFA).value is None:
            continue
        linhas.append([ws.cell(r, c).value for c in range(1, N_COLS + 1)])
    return linhas


def montar_sequencia(tarefas):
    """Intercala simulados. Devolve a lista final e quantos simulados entraram."""
    final, i, num, n_sim, trilha = [], 0, 0, 0, None
    while i < len(tarefas):
        num += 1
        if num >= PRIMEIRO_SIMULADO and num % INTERVALO == 0:
            n_sim += 1
            final.append({"simulado": True, "numero": num,
                          "trilha": trilha, "seq": n_sim})
        else:
            linha = list(tarefas[i])
            if linha[COL_TRILHA - 1] is not None:
                trilha = linha[COL_TRILHA - 1]
            linha[COL_TAREFA - 1] = num
            final.append({"simulado": False, "numero": num, "linha": linha})
            i += 1
    return final, n_sim


def escrever_tarefas(ws, sequencia):
    """Reescreve a aba Tarefas preservando o estilo das linhas de dados."""
    estilo = [copy(ws.cell(PRIMEIRA_LINHA, c)._style) for c in range(1, N_COLS + 1)]

    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        for c in range(1, N_COLS + 1):
            ws.cell(r, c).value = None

    for idx, item in enumerate(sequencia):
        r = PRIMEIRA_LINHA + idx
        if item["simulado"]:
            valores = {
                COL_TAREFA:     item["numero"],
                COL_TRILHA:     item["trilha"],
                COL_DISCIPLINA: "Simulados",
                COL_SEQ:        item["seq"],
                COL_AULA:       item["seq"],
                COL_TIPO:       "Exercícios",
                COL_CONTEUDO:   TEXTO_SIMULADO.format(n=item["seq"]),
            }
            for c in range(1, N_COLS + 1):
                ws.cell(r, c).value = valores.get(c)
            ws.cell(r, COL_DESEMPENHO).value = (
                f'=IF(OR(J{r}="",K{r}=""),"",K{r}/J{r})'
            )
        else:
            for c in range(1, N_COLS + 1):
                ws.cell(r, c).value = item["linha"][c - 1]
        for c in range(1, N_COLS + 1):
            ws.cell(r, c)._style = copy(estilo[c - 1])

    fim = PRIMEIRA_LINHA + len(sequencia) - 1
    for r in range(fim + 1, ws.max_row + 1):
        ws.cell(r, COL_DESEMPENHO).value = f'=IF(OR(J{r}="",K{r}=""),"",K{r}/J{r})'
    return fim


def criar_aba_simulados(wb, sequencia):
    """Cria a aba Simulados espelhando a estrutura das abas de disciplina."""
    if "Simulados" in wb.sheetnames:
        del wb["Simulados"]
    modelo = wb["Português"]
    ws = wb.create_sheet("Simulados")

    for letra, dim in modelo.column_dimensions.items():
        ws.column_dimensions[letra].width = dim.width
    for r in (1, 2):
        for c in range(1, 12):
            ws.cell(r, c).value = modelo.cell(r, c).value
            ws.cell(r, c)._style = copy(modelo.cell(r, c)._style)
    ws["A1"] = "Simulados"

    sims = [i for i in sequencia if i["simulado"]]
    for i, item in enumerate(sims):
        r = 3 + i
        for c in range(1, 12):
            ws.cell(r, c)._style = copy(modelo.cell(3, c)._style)
        ws.cell(r, 1).value = item["seq"]
        ws.cell(r, 2).value = TEXTO_SIMULADO.format(n=item["seq"])
        ws.cell(r, 9).value = f"=IF(H{r}>0,H{r}/G{r},0)"
    return len(sims)


def converter_formulas_google(wb, caminho_original):
    """
    Troca as células __xludf.DUMMYFUNCTION pelos valores em cache.

    Essas funções (QUERY, SPARKLINE) só existem no Google Sheets. No Excel elas
    dependem do valor em cache, que o openpyxl descarta ao salvar — sem essa
    conversão as células virariam #NAME?.
    """
    cache = openpyxl.load_workbook(caminho_original, data_only=True)
    n = 0
    for ws in wb.worksheets:
        if ws.title in ("Tarefas", "Simulados") or ws.title not in cache.sheetnames:
            continue
        wc = cache[ws.title]
        for row in ws.iter_rows():
            for cel in row:
                v = cel.value
                if isinstance(v, str) and "__xludf.DUMMYFUNCTION" in v:
                    cel.value = wc.cell(cel.row, cel.column).value
                    n += 1
    return n


def registrar_em_estatisticas(wb, total_simulados):
    """
    Acrescenta Simulados aos dois blocos da aba Estatísticas e corrige três
    defeitos da exportação do Google que fazem as fórmulas falharem no Excel:

      1. INDIRECT com intervalo aberto ("B$3:B") — retorna 1 em vez da contagem.
      2. Nome de aba com espaço sem aspas — referência não resolve.
      3. Coluna AL com o nome completo da disciplina, mas o Excel trunca o nome
         da aba em 31 caracteres — o INDIRECT não encontra a aba.
    """
    est = wb["Estatísticas"]
    abas = wb.sheetnames

    def achar_aba(nome):
        if nome in abas:
            return nome
        return next((a for a in abas if nome and nome.startswith(a)), None)

    # Insere a linha de Simulados: bloco 2 primeiro, para não deslocar o bloco 1
    for linha_total in (34, 20):
        est.insert_rows(linha_total)
        for c in range(1, 5):
            est.cell(linha_total, c)._style = copy(est.cell(linha_total - 1, c)._style)
        est.cell(linha_total, 1).value = "Simulados"

    # Blocos: dados 10..20 (total 21) e 25..35 (total 36)
    for inicio, fim, bloco2 in ((10, 20, False), (25, 35, True)):
        for r in range(inicio, fim + 1):
            disciplina = est.cell(r, 1).value
            aba = achar_aba(disciplina)
            if aba:
                est.cell(r, 38).value = aba
            est.cell(r, 2).value = f"=COUNTA(INDIRECT(\"'\"&$AL{r}&\"'!B$3:B$500\"))"
            est.cell(r, 5).value = f"=IF(C{r}>0,C{r}/B{r},0)"
            if bloco2:
                est.cell(r, 7).value = (
                    f'=SUMIFS(Tarefas!J$4:J1000,Tarefas!G$4:G1000,A{r},'
                    f'Tarefas!A$4:A1000,">=0")'
                )
                est.cell(r, 8).value = f'=SUMIFS(Tarefas!K$4:K1000,Tarefas!G$4:G1000,A{r})'
                est.cell(r, 9).value = f'=IF(H{r}>0,H{r}/G{r},0)'
            else:
                est.cell(r, 7).value = f'=SUMIFS(Tarefas!J$4:J1000,Tarefas!G$4:G1000,A{r})'

        total = fim + 1 if not bloco2 else 36
        colunas = ((2, "B"), (3, "C"), (4, "D"), (7, "G"))
        if bloco2:
            colunas += ((8, "H"),)
        for c, letra in colunas:
            est.cell(total, c).value = f"=SUM({letra}{inicio}:{letra}{fim})"
        est.cell(total, 5).value = f"=IF(C{total}>0,C{total}/B{total},0)"
        if bloco2:
            est.cell(total, 9).value = f"=IF(H{total}>0,H{total}/G{total},0)"

        # Linha de Simulados: sem cache do Google, então usa fórmula de verdade
        est.cell(fim, 3).value = f'=COUNTIF(Tarefas!$G:$G,$A{fim})'
        est.cell(fim, 4).value = f'=COUNTIF(Tarefas!$G:$G,$A{fim})'


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    origem = Path(sys.argv[1])
    destino = Path(sys.argv[2]) if len(sys.argv) > 2 else \
        origem.with_name(origem.stem + "_com_simulados.xlsx")

    if not origem.exists():
        print(f"Arquivo não encontrado: {origem}")
        sys.exit(1)

    wb = openpyxl.load_workbook(origem)
    if "Tarefas" not in wb.sheetnames:
        print("A planilha precisa ter a aba 'Tarefas'.")
        sys.exit(1)

    ws = wb["Tarefas"]
    tarefas = ler_tarefas(ws)
    if any(str(ws.cell(r, COL_DISCIPLINA).value) == "Simulados"
           for r in range(PRIMEIRA_LINHA, ws.max_row + 1)):
        print("ERRO: esta planilha JÁ contém simulados.")
        print("      Rode o script sempre sobre a planilha original do fornecedor.")
        sys.exit(1)

    print(f"Tarefas originais lidas: {len(tarefas)}")

    sequencia, n_sim = montar_sequencia(tarefas)
    escrever_tarefas(ws, sequencia)
    criar_aba_simulados(wb, sequencia)
    convertidas = converter_formulas_google(wb, origem)
    registrar_em_estatisticas(wb, n_sim)

    wb.save(destino)

    posicoes = [i["numero"] for i in sequencia if i["simulado"]]
    print(f"Simulados inseridos:     {n_sim}")
    print(f"Posições:                {posicoes[:5]} ... {posicoes[-2:]}")
    print(f"Total de tarefas:        {len(sequencia)}")
    print(f"Fórmulas Google fixadas: {convertidas}")
    print(f"\nSalvo em: {destino}")
    print("\nAntes de importar, rode:  python verificar_impacto.py "
          f'"{destino}"')


if __name__ == "__main__":
    main()