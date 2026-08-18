"""
Sistema de Acompanhamento de Estudos — versão evoluída.
Dashboard analítico avançado, menu moderno, KPIs inteligentes, análise por IA e tooltips completos.
"""

import hashlib
import hmac
import html as html_lib
import os
import re
import sqlite3
from sqlite3 import IntegrityError
from contextlib import contextmanager
from datetime import date, datetime, time, timedelta
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    import psycopg
    from psycopg.errors import UniqueViolation
except ImportError:
    psycopg = None
    UniqueViolation = IntegrityError


# ─────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────

def config_valor(chave, padrao=None):
    valor = os.getenv(chave)
    if valor not in (None, ""):
        return valor
    try:
        if chave in st.secrets and st.secrets[chave] not in (None, ""):
            return str(st.secrets[chave])
    except Exception:
        pass
    return padrao


APP_NAME = "Sistema de Acompanhamento de Estudos"
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / config_valor("ESTUDOS_DB", "trilha_tjs_ajaa.db")
APP_ENV = config_valor("APP_ENV", "development").lower()
DATABASE_URL = config_valor("DATABASE_URL")
DB_BACKEND = "postgresql" if DATABASE_URL else "sqlite"
IS_PRODUCTION = APP_ENV == "production"
DEBUG = config_valor("DEBUG", "0") == "1" and not IS_PRODUCTION
def _resolver_referencia(valor):
    """
    Caminho da planilha de referência. Um valor absoluto é usado como veio;
    um valor relativo é resolvido a partir da pasta do app, para o código não
    depender da máquina de ninguém.
    """
    caminho = Path(str(valor).strip()).expanduser()
    return caminho if caminho.is_absolute() else (BASE_DIR / caminho)


PLANILHA_REFERENCIA = _resolver_referencia(
    config_valor("ESTUDOS_REFERENCIA", "Acompanhamento_da_Trilha.xlsx")
)
ADMIN_EMAIL = config_valor("ESTUDOS_ADMIN_EMAIL", "admin@admin.com")
ADMIN_PASSWORD = config_valor("ESTUDOS_ADMIN_PASSWORD", "123")

STATUS_NAO_INICIADA = "NAO_INICIADA"
STATUS_EM_ANDAMENTO = "EM_ANDAMENTO"
STATUS_CONCLUIDA = "CONCLUIDA"
STATUS_VALIDOS = [STATUS_NAO_INICIADA, STATUS_EM_ANDAMENTO, STATUS_CONCLUIDA]
STATUS_ANALISE = [STATUS_EM_ANDAMENTO, STATUS_CONCLUIDA]
STATUS_LABELS = {
    STATUS_NAO_INICIADA: "Não iniciada",
    STATUS_EM_ANDAMENTO: "Em andamento",
    STATUS_CONCLUIDA: "Concluída",
}
STATUS_CORES = {
    STATUS_NAO_INICIADA: "#94a3b8",
    STATUS_EM_ANDAMENTO: "#f59e0b",
    STATUS_CONCLUIDA: "#22c55e",
}
TIPOS_ESTUDO = [
    "Leitura", "Exercícios", "Revisão", "Simulado",
    "Videoaula", "Resumo", "Flashcards", "Projeto", "Pesquisa", "Outro",
]

# Paleta de cores
COR_PRIMARIA = "#1e40af"
COR_SUCESSO  = "#16a34a"
COR_ALERTA   = "#d97706"
COR_PERIGO   = "#dc2626"

# Itens de menu com ícones (usados no menu lateral customizado)
MENU_ITENS_GESTOR = [
    ("📊", "Dashboard"), ("⚡", "Registro rápido"), ("📋", "Tarefas"),
    ("📖", "Aulas e assuntos"), ("🎓", "Disciplinas"), ("👥", "Alunos"),
    ("📥", "Importações"), ("⚙️", "Configurações"),
]
MENU_ITENS_ALUNO = [
    ("📊", "Dashboard"), ("⚡", "Registro rápido"), ("📋", "Tarefas"),
    ("📖", "Aulas e assuntos"), ("⚙️", "Configurações"),
]

st.set_page_config(
    page_title=APP_NAME,
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ═══════════════════════════════════════════════════════════
# CSS GLOBAL — menu moderno, dashboard analítico, tooltips
# ═══════════════════════════════════════════════════════════

CSS = """
<style>
:root {
  --c-bg:      #f1f5f9;
  --c-sidebar: #0f172a;
  --c-accent:  #3b82f6;
  --c-text:    #0f172a;
  --c-muted:   #64748b;
  --c-border:  #e2e8f0;
  --c-white:   #ffffff;
  --c-ok:      #16a34a;
  --c-warn:    #d97706;
  --c-danger:  #dc2626;
  color-scheme: light;
}

/* ══ LAYOUT ══ */
.block-container { padding: 1rem 1.5rem 2rem; max-width: 1560px; }
[data-testid="stAppViewContainer"] > .main { background: var(--c-bg); }

/* ══ SIDEBAR — textos sempre visíveis, menu moderno ══ */
[data-testid="stSidebar"] {
  background: var(--c-sidebar) !important;
  min-width: 240px !important;
  max-width: 260px !important;
}
/* Labels de widgets na sidebar sempre visíveis */
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stMultiSelect label,
[data-testid="stSidebar"] .stDateInput label,
[data-testid="stSidebar"] .stNumberInput label,
[data-testid="stSidebar"] .stTextInput label,
[data-testid="stSidebar"] .stToggle label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span {
  color: #cbd5e1 !important;
  opacity: 1 !important;
  visibility: visible !important;
}
/* Título da aplicação (st.sidebar.title → h1).
   Sem esta regra o h1 ficava com a cor padrão do Streamlit (#31333F) sobre o
   fundo #0f172a: contraste 1.42, praticamente invisível. */
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] .stMarkdown h1,
[data-testid="stSidebar"] .stMarkdown h2 {
  color: #f1f5f9 !important;
  font-size: 1.02rem !important;
  font-weight: 800 !important;
  line-height: 1.3 !important;
  padding: 0 !important;
  margin: .2rem 0 .8rem !important;
}
[data-testid="stSidebar"] h4,
[data-testid="stSidebar"] h5,
[data-testid="stSidebar"] h6 { color: #e2e8f0 !important; }

/* Títulos de seção na sidebar */
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] .stMarkdown h3 {
  color: #94a3b8 !important;
  font-size: .68rem !important;
  font-weight: 800 !important;
  letter-spacing: .12em !important;
  text-transform: uppercase !important;
  margin: 1.2rem 0 .4rem !important;
  padding: 0 !important;
}
/* Radio buttons do menu — sempre visíveis */
[data-testid="stSidebar"] .stRadio > div { gap: 2px !important; }
[data-testid="stSidebar"] .stRadio label {
  display: flex !important;
  align-items: center !important;
  gap: 8px !important;
  padding: 9px 14px !important;
  border-radius: 8px !important;
  margin: 1px 0 !important;
  cursor: pointer !important;
  color: #cbd5e1 !important;
  font-size: .85rem !important;
  font-weight: 500 !important;
  opacity: 1 !important;
  visibility: visible !important;
  transition: background .15s, color .15s !important;
}
[data-testid="stSidebar"] .stRadio label:hover {
  background: rgba(255,255,255,.08) !important;
  color: #f1f5f9 !important;
}
[data-testid="stSidebar"] .stRadio label[data-testid="stWidgetLabel"] { display: none !important; }

/* Item ativo do menu.
   O Streamlit marca a opção selecionada com data-selected="true" em
   stRadioOption. Os seletores antigos (input:checked ~ label) não existem
   nessa árvore, então o item ativo não recebia destaque nenhum. */
[data-testid="stSidebar"] [data-testid="stRadioOption"] {
  border-radius: 8px !important;
  padding: 2px 6px !important;
  margin: 1px 0 !important;
}
[data-testid="stSidebar"] [data-testid="stRadioOption"] p {
  color: #cbd5e1 !important;
  font-size: .85rem !important;
  font-weight: 500 !important;
}
[data-testid="stSidebar"] [data-testid="stRadioOption"]:hover {
  background: rgba(255,255,255,.08) !important;
}
[data-testid="stSidebar"] [data-testid="stRadioOption"][data-selected="true"] {
  background: rgba(59,130,246,.25) !important;
}
[data-testid="stSidebar"] [data-testid="stRadioOption"][data-selected="true"] p {
  color: #bfdbfe !important;
  font-weight: 700 !important;
}

/* Chips do multiselect: o tema do Streamlit pinta o fundo de vermelho (#ff4b4b)
   e a regra genérica de span acima deixava o texto cinza-claro em cima —
   contraste 2.22. Fundo azul do tema e texto branco. */
[data-testid="stSidebar"] [data-testid="stMultiSelectTagsContainer"] span[data-tag] {
  background: #2563eb !important;
  border-radius: 6px !important;
}
[data-testid="stSidebar"] [data-testid="stMultiSelectTagsContainer"] span[data-tag] span,
[data-testid="stSidebar"] [data-testid="stMultiSelectTagsContainer"] span[data-tag] {
  color: #ffffff !important;
}
[data-testid="stSidebar"] [data-testid="stMultiSelectTagsContainer"] span[data-tag] svg {
  fill: #ffffff !important;
}
/* Botões da sidebar */
[data-testid="stSidebar"] .stButton > button {
  background: rgba(255,255,255,.07) !important;
  color: #e2e8f0 !important;
  border: 1px solid rgba(255,255,255,.12) !important;
  border-radius: 8px !important;
  font-size: .82rem !important;
  font-weight: 600 !important;
  width: 100% !important;
  text-align: left !important;
  padding: 9px 14px !important;
  margin-bottom: 3px !important;
  display: flex !important;
  align-items: center !important;
  gap: 8px !important;
  transition: background .15s !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
  background: rgba(255,255,255,.14) !important;
  color: #f8fafc !important;
}
[data-testid="stSidebar"] .stButton > button p { color: inherit !important; font-size: .82rem !important; }
/* Selectbox/inputs da sidebar */
[data-testid="stSidebar"] .stSelectbox > div > div,
[data-testid="stSidebar"] .stMultiSelect > div > div {
  background: rgba(255,255,255,.07) !important;
  border-color: rgba(255,255,255,.15) !important;
  color: #e2e8f0 !important;
}
[data-testid="stSidebar"] .stTextInput input,
[data-testid="stSidebar"] .stNumberInput input {
  background: rgba(255,255,255,.07) !important;
  border-color: rgba(255,255,255,.15) !important;
  color: #e2e8f0 !important;
}
/* Valor selecionado do selectbox/multiselect.
   No Streamlit atual isso é o value de um <input> dentro de react-aria-ComboBox,
   não um nó de texto — as regras de `span`/`p`/`> div > div` não o alcançavam e
   ele ficava com a cor escura padrão sobre o fundo escuro da sidebar. */
[data-testid="stSidebar"] input {
  color: #e2e8f0 !important;
  -webkit-text-fill-color: #e2e8f0 !important;
  opacity: 1 !important;
}
[data-testid="stSidebar"] input::placeholder {
  color: #94a3b8 !important;
  -webkit-text-fill-color: #94a3b8 !important;
  opacity: 1 !important;
}
/* Ícones (seta do selectbox, calendário, limpar) */
[data-testid="stSidebar"] [data-testid="stSelectbox"] svg,
[data-testid="stSidebar"] [data-testid="stMultiSelect"] svg,
[data-testid="stSidebar"] [data-testid="stDateInput"] svg {
  fill: #cbd5e1 !important;
}
/* Cabeçalho do usuário na sidebar */
.sidebar-user {
  padding: 16px 14px 12px;
  border-bottom: 1px solid rgba(255,255,255,.1);
  margin-bottom: 8px;
}
.sidebar-user-name { color: #f1f5f9; font-size: .92rem; font-weight: 700; margin: 0; }
.sidebar-user-role { color: #64748b; font-size: .72rem; margin: 2px 0 0; }
.sidebar-nav-section {
  padding: 6px 14px 2px;
  color: #475569;
  font-size: .66rem;
  font-weight: 800;
  letter-spacing: .1em;
  text-transform: uppercase;
}

/* ══ MÉTRICAS NATIVAS ══ */
[data-testid="stMetric"] {
  background: var(--c-white);
  border: 1px solid var(--c-border);
  border-radius: 10px;
  padding: 14px 16px;
  box-shadow: 0 1px 3px rgba(15,23,42,.05);
}
[data-testid="stMetricValue"] { font-size: 1.4rem; font-weight: 800; color: var(--c-text); }
[data-testid="stMetricLabel"] { color: var(--c-muted); font-size: .78rem; font-weight: 600; }
[data-testid="stMetricDelta"] { font-size: .76rem !important; }

/* ══ DATAFRAMES ══ */
div[data-testid="stDataFrame"] { border: 1px solid var(--c-border); border-radius: 10px; overflow: hidden; }

/* ══ HERO ══ */
.hero {
  background: linear-gradient(135deg,#eff6ff 0%,#f8fafc 50%,#ecfdf5 100%);
  border: 1px solid #dbeafe; border-radius: 12px; padding: 18px 24px; margin-bottom: 16px;
}
.hero h1 { margin: 0 0 4px; font-size: 1.5rem; color: var(--c-text); font-weight: 800; }
.hero p  { margin: 0; color: var(--c-muted); font-size: .88rem; }

/* ══ CARDS DE ATIVIDADE ══ */
.quick-card {
  background: var(--c-white); border: 1px solid var(--c-border);
  border-left: 5px solid #94a3b8; border-radius: 10px;
  padding: 14px 18px; margin-bottom: 14px;
  box-shadow: 0 2px 8px rgba(15,23,42,.05);
}
.quick-card.ok   { border-left-color: var(--c-ok); }
.quick-card.warn { border-left-color: var(--c-warn); }
.quick-card.off  { border-left-color: #94a3b8; }
.quick-card h3   { margin: 4px 0 0; font-size: 1rem; }
.quick-grid { display: grid; grid-template-columns: repeat(4,minmax(0,1fr)); gap: 8px; margin-top: 12px; }
@media(max-width:900px){ .quick-grid{ grid-template-columns: repeat(2,1fr); } }
@media(max-width:560px){ .quick-grid{ grid-template-columns: 1fr; } }
.quick-item { background: #f8fafc; border: 1px solid var(--c-border); border-radius: 7px; padding: 8px 10px; }
.quick-label { color: var(--c-muted); font-size: .66rem; font-weight: 800; text-transform: uppercase; letter-spacing: .06em; }
.quick-value { color: var(--c-text); font-weight: 700; margin-top: 2px; font-size: .87rem; overflow-wrap: anywhere; }

/* ══ STATUS BADGES ══ */
.status-ok, .status-warn, .status-off {
  display: inline-flex; align-items: center; gap: 4px;
  padding: 3px 10px; border-radius: 999px; font-weight: 700; font-size: .77rem; margin-top: 4px;
}
.status-ok   { background: #dcfce7; color: #166534; }
.status-warn { background: #fef3c7; color: #92400e; }
.status-off  { background: #f1f5f9; color: #475569; }

/* ══ KPI CARD CUSTOMIZADO com tooltip ══ */
.kpi-card {
  background: var(--c-white); border: 1px solid var(--c-border); border-radius: 10px;
  padding: 14px 16px; position: relative; box-shadow: 0 1px 4px rgba(15,23,42,.05);
  height: 100%; min-height: 100px;
}
.kpi-label  { font-size: .67rem; font-weight: 800; color: var(--c-muted); text-transform: uppercase; letter-spacing: .07em; margin-bottom: 2px; }
.kpi-value  { font-size: 1.5rem; font-weight: 900; color: var(--c-text); margin: 4px 0 2px; line-height: 1.1; }
.kpi-sub    { font-size: .7rem; color: var(--c-muted); line-height: 1.4; }
.kpi-delta-pos { font-size: .7rem; color: var(--c-ok); font-weight: 700; margin-top: 3px; }
.kpi-delta-neg { font-size: .7rem; color: var(--c-danger); font-weight: 700; margin-top: 3px; }
/* Tooltip */
.kpi-tooltip {
  position: absolute; top: 8px; right: 8px;
  width: 17px; height: 17px; border-radius: 50%;
  background: #e2e8f0; color: #64748b;
  font-size: .64rem; font-weight: 900;
  display: inline-flex; align-items: center; justify-content: center;
  cursor: help; user-select: none;
}
.kpi-tooltip[title] { text-decoration: none; }
/* Tooltip via CSS puro — visível em hover */
.kpi-ttip-wrap { display: inline-block; position: absolute; top: 8px; right: 8px; }
.kpi-ttip-icon {
  width: 17px; height: 17px; border-radius: 50%; background: #e2e8f0;
  color: #64748b; font-size: .63rem; font-weight: 900;
  display: inline-flex; align-items: center; justify-content: center; cursor: help;
}
.kpi-ttip-box {
  display: none; position: absolute; z-index: 9999;
  bottom: 130%; right: 0;
  background: #1e293b; color: #f1f5f9;
  font-size: .72rem; line-height: 1.55; padding: 10px 13px;
  border-radius: 9px; width: 270px;
  box-shadow: 0 6px 20px rgba(0,0,0,.35); white-space: normal;
}
.kpi-ttip-box::after {
  content: ""; position: absolute; top: 100%; right: 4px;
  border: 6px solid transparent; border-top-color: #1e293b;
}
.kpi-ttip-wrap:hover .kpi-ttip-box { display: block; }

/* ══ INSIGHT CARDS ══ */
.insight-card {
  border-radius: 10px; padding: 12px 16px; margin-bottom: 8px;
  display: flex; gap: 12px; align-items: flex-start;
}
.insight-card.info    { background: #eff6ff; border-left: 4px solid #3b82f6; }
.insight-card.success { background: #f0fdf4; border-left: 4px solid #22c55e; }
.insight-card.warning { background: #fffbeb; border-left: 4px solid #f59e0b; }
.insight-card.danger  { background: #fef2f2; border-left: 4px solid #ef4444; }
.insight-icon  { font-size: 1.1rem; min-width: 20px; padding-top: 1px; }
.insight-body  { flex: 1; }
.insight-title { font-size: .83rem; font-weight: 800; color: var(--c-text); margin: 0 0 2px; }
.insight-text  { font-size: .77rem; color: var(--c-muted); margin: 0; line-height: 1.45; }

/* ══ REGRAS DE NEGÓCIO ══ */
.rule-warning { background:#fffbeb; border:1px solid #fde68a; border-left:5px solid #f59e0b; border-radius:8px; padding:12px 14px; margin:10px 0; color:#78350f; font-size:.88rem; }
.rule-error   { background:#fef2f2; border:1px solid #fecaca; border-left:5px solid #ef4444; border-radius:8px; padding:12px 14px; margin:10px 0; color:#7f1d1d; font-size:.88rem; }

/* ══ UTILITÁRIOS ══ */
.section-title { font-size:.93rem; font-weight:800; color:var(--c-text); margin:14px 0 6px; }
.muted { color:var(--c-muted); font-size:.84rem; }

/* ══ BOTÕES ══ */
.stButton > button[kind="primary"] { background: var(--c-accent); color:#fff; border-radius:8px; font-weight:700; }
.stButton > button[kind="primary"]:hover { background:#1d4ed8; }

/* ══ FORMULÁRIOS ══ */
[data-testid="stForm"] { border:1px solid var(--c-border) !important; border-radius:10px !important; padding:16px !important; }

/* ══ GRÁFICO RANK ══ */
.rank-row {
  display: flex; align-items: center; gap: 10px;
  padding: 9px 12px; border-radius: 8px; background: #f8fafc;
  border: 1px solid var(--c-border); margin-bottom: 6px;
}
.rank-pos { font-size: 1rem; font-weight: 900; color: var(--c-muted); min-width: 28px; text-align: center; }
.rank-name { flex: 1; font-size: .85rem; font-weight: 700; color: var(--c-text); }
.rank-bar-wrap { width: 100px; height: 8px; background: #e2e8f0; border-radius: 999px; overflow: hidden; }
.rank-bar { height: 100%; border-radius: 999px; background: var(--c-accent); }
.rank-val { font-size: .8rem; font-weight: 700; color: var(--c-text); min-width: 48px; text-align: right; }

/* ══ RESPONSIVIDADE ══ */
@media(max-width:768px){
  .block-container { padding: .75rem 1rem 1.5rem; }
  .kpi-value { font-size: 1.2rem !important; }
}
</style>
"""

st.markdown(CSS, unsafe_allow_html=True)



# ─────────────────────────────────────────────
# UTILITÁRIOS DE RENDERIZAÇÃO
# ─────────────────────────────────────────────

def render_html(conteudo: str):
    if hasattr(st, "html"):
        st.html(conteudo)
    else:
        st.markdown(conteudo, unsafe_allow_html=True)


def escape_html(valor):
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    return html_lib.escape(str(valor))


def status_badge(status: str) -> str:
    cls = {STATUS_CONCLUIDA:"status-ok", STATUS_EM_ANDAMENTO:"status-warn", STATUS_NAO_INICIADA:"status-off"}.get(status,"status-off")
    icone = {"CONCLUIDA":"✓","EM_ANDAMENTO":"●","NAO_INICIADA":"○"}.get(status,"○")
    return f'<span class="{cls}">{icone} {escape_html(STATUS_LABELS.get(status,status))}</span>'


def status_card_class(status: str) -> str:
    return {STATUS_CONCLUIDA:"ok",STATUS_EM_ANDAMENTO:"warn",STATUS_NAO_INICIADA:"off"}.get(status,"off")


def kpi_card(label: str, valor: str, subtexto: str = "", delta: str = "", delta_pos: bool = True, tooltip: str = "") -> str:
    """KPI card com tooltip sempre visível ao hover (CSS puro, sem JS)."""
    delta_html = ""
    if delta:
        cls = "kpi-delta-pos" if delta_pos else "kpi-delta-neg"
        seta = "▲" if delta_pos else "▼"
        delta_html = f'<div class="{cls}">{seta} {escape_html(delta)}</div>'
    tooltip_html = ""
    if tooltip:
        tooltip_html = (
            '<div class="kpi-ttip-wrap">'
            '<div class="kpi-ttip-icon">?</div>'
            f'<div class="kpi-ttip-box">{escape_html(tooltip)}</div>'
            '</div>'
        )
    return (
        '<div class="kpi-card">'
        + tooltip_html
        + f'<div class="kpi-label">{escape_html(label)}</div>'
        + f'<div class="kpi-value">{escape_html(str(valor))}</div>'
        + f'<div class="kpi-sub">{escape_html(subtexto)}</div>'
        + delta_html
        + '</div>'
    )


def insight_card(tipo: str, icone: str, titulo: str, texto: str) -> str:
    return (
        f'<div class="insight-card {tipo}">'
        f'<div class="insight-icon">{icone}</div>'
        '<div class="insight-body">'
        f'<div class="insight-title">{escape_html(titulo)}</div>'
        f'<p class="insight-text">{escape_html(texto)}</p>'
        '</div></div>'
    )


# ─────────────────────────────────────────────
# UTILITÁRIOS GERAIS
# ─────────────────────────────────────────────

def limpar_texto(valor):
    if pd.isna(valor):
        return None
    texto = str(valor).replace("\n", " ").strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto if texto else None


def quebrar_texto(texto, limite=26):
    texto = limpar_texto(texto) or ""
    partes, linha = [], []
    for palavra in texto.split():
        candidato = " ".join(linha + [palavra])
        if linha and len(candidato) > limite:
            partes.append(" ".join(linha))
            linha = [palavra]
        else:
            linha.append(palavra)
    if linha:
        partes.append(" ".join(linha))
    return "<br>".join(partes) if partes else texto


def normalizar_email(valor):
    email = limpar_texto(valor)
    return email.lower() if email else None


def email_local(nome):
    nome = str(nome or "aluno").lower().strip()
    nome = re.sub(r"[^a-z0-9\s.]", "", nome)
    nome = re.sub(r"\s+", ".", nome)
    nome = re.sub(r"\.+", ".", nome).strip(".")
    return f"{nome or 'aluno'}@local.com"


def chave_texto(texto):
    texto = limpar_texto(texto) or ""
    mapa = str.maketrans("áàâãéêíóôõúüçÁÀÂÃÉÊÍÓÔÕÚÜÇ", "aaaaeeiooouucAAAAEEIOOOUUC")
    texto = texto.translate(mapa).lower()
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def converter_horas(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, time):
        return valor.hour + valor.minute / 60 + valor.second / 3600
    if isinstance(valor, datetime):
        return valor.hour + valor.minute / 60 + valor.second / 3600
    if isinstance(valor, timedelta):
        return valor.total_seconds() / 3600
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        valor = valor.strip()
        if not valor:
            return 0.0
        if ":" in valor:
            partes = valor.split(":")
            try:
                h = int(partes[0])
                m = int(partes[1]) if len(partes) > 1 else 0
                s = int(partes[2]) if len(partes) > 2 else 0
                return h + m / 60 + s / 3600
            except ValueError:
                return 0.0
        try:
            return float(valor.replace(",", "."))
        except ValueError:
            return 0.0
    return 0.0


def horas_para_hm(horas_float: float) -> str:
    """
    Converte um valor decimal de horas para string legível 'Xh Ymin'.
    Exemplos: 1.5 → '1h 30min' | 0.25 → '15min' | 2.0 → '2h' | 0.0 → '—'
    """
    try:
        total_min = int(round(float(horas_float or 0) * 60))
    except (TypeError, ValueError):
        return "—"
    if total_min <= 0:
        return "—"
    h = total_min // 60
    m = total_min % 60
    if h > 0 and m > 0:
        return f"{h}h {m}min"
    if h > 0:
        return f"{h}h"
    return f"{m}min"


def hm_para_horas(horas: int, minutos: int) -> float:
    """Converte horas inteiras + minutos inteiros para float decimal."""
    return max(0.0, float(horas or 0) + float(minutos or 0) / 60)


def converter_numero(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, (time, datetime, timedelta)):
        return converter_horas(valor)
    if isinstance(valor, str):
        valor = valor.strip().replace("%", "").replace(",", ".")
        if not valor:
            return 0.0
        try:
            return float(valor)
        except ValueError:
            return 0.0
    return 0.0


def converter_inteiro(valor):
    return int(round(converter_numero(valor)))


def converter_data(valor):
    if pd.isna(valor) or valor in ("", None):
        return None
    try:
        return str(pd.to_datetime(valor).date())
    except Exception:
        return None


def formatar_data_br(valor):
    data = converter_data(valor)
    if not data:
        return ""
    return pd.to_datetime(data).strftime("%d/%m/%Y")


def nome_aluno_data_arquivo(caminho):
    caminho = Path(caminho)
    nome = caminho.stem.split("-", 1)[0].strip()
    data_execucao = date.today()
    match = re.search(r"(\d{2})_(\d{2})_(\d{4})", caminho.stem)
    if match:
        dia, mes, ano = map(int, match.groups())
        data_execucao = date(ano, mes, dia)
    return nome, data_execucao


def distribuir_inteiro(total, pesos):
    total = int(round(total or 0))
    if not pesos:
        return []
    if total == 0:
        return [0 for _ in pesos]
    pesos = [float(p or 0) for p in pesos]
    if sum(pesos) <= 0:
        pesos = [1 for _ in pesos]
    bruto = [total * p / sum(pesos) for p in pesos]
    base = [int(v) for v in bruto]
    sobra = total - sum(base)
    ordem = sorted(range(len(bruto)), key=lambda i: bruto[i] - base[i], reverse=True)
    for i in ordem[:sobra]:
        base[i] += 1
    return base


def normalizar_tipo_estudo(valor):
    tipo = limpar_texto(valor) or "Outro"
    aliases = {
        "teoria": "Leitura",
        "questoes": "Exercícios",
        "questões": "Exercícios",
        "exercicios": "Exercícios",
        "exercícios": "Exercícios",
        "video aula": "Videoaula",
        "vídeo aula": "Videoaula",
    }
    tipo = aliases.get(chave_texto(tipo), tipo)
    return tipo if tipo in TIPOS_ESTUDO else "Outro"


def erro_usuario(mensagem, exc=None):
    if DEBUG and exc is not None:
        st.error(f"{mensagem}: {exc}")
    else:
        st.error(mensagem)


# ─────────────────────────────────────────────
# SEGURANÇA — SENHAS
# ─────────────────────────────────────────────

def hash_senha(senha):
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", str(senha).encode(), salt, 120_000)
    return f"pbkdf2_sha256${salt.hex()}${digest.hex()}"


def verificar_senha(senha, senha_armazenada):
    if not senha_armazenada:
        return False
    if not str(senha_armazenada).startswith("pbkdf2_sha256$"):
        return hmac.compare_digest(str(senha), str(senha_armazenada))
    try:
        _, salt_hex, digest_hex = str(senha_armazenada).split("$", 2)
        digest = hashlib.pbkdf2_hmac("sha256", str(senha).encode(), bytes.fromhex(salt_hex), 120_000)
        return hmac.compare_digest(digest.hex(), digest_hex)
    except ValueError:
        return False


def senha_valida(nova):
    return (
        len(str(nova or "")) >= 8
        and bool(re.search(r"[A-Za-z]", str(nova)))
        and bool(re.search(r"\d", str(nova)))
    )


def atualizar_senha_usuario(usuario_id, nova_senha, forcar_troca=0):
    executar(
        "UPDATE alunos SET senha = ?, force_troca_senha = ? WHERE id = ?",
        (hash_senha(nova_senha), int(forcar_troca), int(usuario_id)),
    )


# ─────────────────────────────────────────────
# BANCO DE DADOS
# ─────────────────────────────────────────────

def adaptar_sql(sql):
    if DB_BACKEND != "postgresql":
        return sql
    return sql.replace("?", "%s")


class ConexaoDB:
    """Wrapper para compatibilidade entre SQLite e PostgreSQL."""
    def __init__(self, conn):
        self.conn = conn
        self._cursor = None

    def execute(self, sql, params=()):
        cur = self.conn.cursor()
        cur.execute(adaptar_sql(sql), params)
        self._cursor = cur
        return cur

    def cursor(self):
        return self

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()


def abrir_conexao_raw():
    if DB_BACKEND == "postgresql":
        if psycopg is None:
            raise RuntimeError("Instale psycopg[binary] para usar PostgreSQL em produção.")
        return psycopg.connect(DATABASE_URL)
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


@contextmanager
def conectar():
    raw = abrir_conexao_raw()
    conn = ConexaoDB(raw) if DB_BACKEND == "postgresql" else raw
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def executar(sql, params=()):
    with conectar() as conn:
        conn.execute(sql, params)


def consultar(sql, params=()):
    conn = abrir_conexao_raw()
    try:
        return pd.read_sql_query(adaptar_sql(sql), conn, params=params)
    finally:
        conn.close()


def ultimo_id(conn):
    if DB_BACKEND == "postgresql":
        return int(conn.execute("SELECT LASTVAL()").fetchone()[0])
    return int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])


def erro_integridade(exc):
    return isinstance(exc, (IntegrityError, UniqueViolation))


# ─────────────────────────────────────────────
# CRIAÇÃO DAS TABELAS
# ─────────────────────────────────────────────

def criar_tabelas():
    if DB_BACKEND == "postgresql":
        _criar_tabelas_postgresql()
        return
    _criar_tabelas_sqlite()


def _criar_tabelas_postgresql():
    with conectar() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS alunos (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL UNIQUE,
                email TEXT UNIQUE,
                senha TEXT NOT NULL,
                perfil TEXT NOT NULL DEFAULT 'Aluno' CHECK(perfil IN ('Gestor','Aluno')),
                ativo INTEGER DEFAULT 1,
                force_troca_senha INTEGER DEFAULT 1
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS disciplinas (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL UNIQUE,
                ativo INTEGER DEFAULT 1
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS aulas (
                id SERIAL PRIMARY KEY,
                disciplina_id INTEGER NOT NULL REFERENCES disciplinas(id),
                aula TEXT NOT NULL,
                assunto TEXT,
                estudada_padrao TEXT DEFAULT 'Não',
                revisao_24h_padrao TEXT DEFAULT 'Não',
                tipo_estudo TEXT DEFAULT 'Outro',
                ativo INTEGER DEFAULT 1
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS assuntos (
                id SERIAL PRIMARY KEY,
                aula_id INTEGER NOT NULL REFERENCES aulas(id),
                titulo TEXT NOT NULL,
                ativo INTEGER DEFAULT 1,
                UNIQUE(aula_id, titulo)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tarefas (
                id SERIAL PRIMARY KEY,
                numero INTEGER NOT NULL UNIQUE,
                trilha INTEGER,
                disciplina_id INTEGER NOT NULL REFERENCES disciplinas(id),
                seq_disciplina INTEGER,
                aula TEXT,
                qtd_exercicios_previstos INTEGER DEFAULT 0,
                tipo TEXT,
                conteudo TEXT,
                ativo INTEGER DEFAULT 1,
                aula_id INTEGER REFERENCES aulas(id),
                assunto_id INTEGER REFERENCES assuntos(id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS execucoes (
                id SERIAL PRIMARY KEY,
                aluno_id INTEGER NOT NULL REFERENCES alunos(id),
                tarefa_id INTEGER NOT NULL REFERENCES tarefas(id),
                data_execucao TEXT,
                ch_efetiva REAL DEFAULT 0,
                data_revisao_24h TEXT,
                ch_revisao REAL DEFAULT 0,
                qtd_acertos INTEGER DEFAULT 0,
                desempenho REAL DEFAULT 0,
                comentario TEXT,
                concluida INTEGER DEFAULT 0,
                atualizado_em TEXT DEFAULT CURRENT_TIMESTAMP,
                qtd_questoes_feitas INTEGER DEFAULT 0,
                status TEXT DEFAULT 'NAO_INICIADA',
                tipo_estudo TEXT DEFAULT 'Outro',
                UNIQUE(aluno_id, tarefa_id)
            )
        """)
        # Migrações seguras
        conn.execute("ALTER TABLE alunos ADD COLUMN IF NOT EXISTS force_troca_senha INTEGER DEFAULT 1")
        conn.execute("ALTER TABLE execucoes ADD COLUMN IF NOT EXISTS qtd_questoes_feitas INTEGER DEFAULT 0")
        conn.execute("ALTER TABLE execucoes ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'NAO_INICIADA'")
        conn.execute("ALTER TABLE execucoes ADD COLUMN IF NOT EXISTS tipo_estudo TEXT DEFAULT 'Outro'")
        conn.execute("ALTER TABLE aulas ADD COLUMN IF NOT EXISTS tipo_estudo TEXT DEFAULT 'Outro'")
        conn.execute("ALTER TABLE tarefas ADD COLUMN IF NOT EXISTS aula_id INTEGER")
        conn.execute("ALTER TABLE tarefas ADD COLUMN IF NOT EXISTS assunto_id INTEGER")
        # Tabela de sessões individuais de estudo
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sessoes_estudo (
                id SERIAL PRIMARY KEY,
                aluno_id INTEGER NOT NULL REFERENCES alunos(id),
                tarefa_id INTEGER NOT NULL REFERENCES tarefas(id),
                data_sessao TEXT NOT NULL,
                ch_sessao REAL DEFAULT 0,
                qtd_questoes INTEGER DEFAULT 0,
                qtd_acertos INTEGER DEFAULT 0,
                tipo_estudo TEXT DEFAULT 'Outro',
                comentario TEXT,
                criado_em TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sessoes_aluno ON sessoes_estudo(aluno_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sessoes_tarefa ON sessoes_estudo(tarefa_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sessoes_data ON sessoes_estudo(data_sessao)")
        # Migração: popula sessoes_estudo com dados existentes (apenas uma vez)
        conn.execute("""
            INSERT INTO sessoes_estudo
                (aluno_id, tarefa_id, data_sessao, ch_sessao, qtd_questoes, qtd_acertos, tipo_estudo, comentario)
            SELECT
                e.aluno_id, e.tarefa_id,
                COALESCE(e.data_execucao, CURRENT_DATE::text),
                e.ch_efetiva, e.qtd_questoes_feitas, e.qtd_acertos,
                COALESCE(e.tipo_estudo, 'Outro'), e.comentario
            FROM execucoes e
            WHERE e.ch_efetiva > 0
              AND NOT EXISTS (
                  SELECT 1 FROM sessoes_estudo s
                  WHERE s.aluno_id = e.aluno_id AND s.tarefa_id = e.tarefa_id
              )
        """)
        _normalizar_status(conn)
        conn.execute("DELETE FROM alunos WHERE email = 'aluno@local.com' OR nome = 'Aluno Padrão'")
        _criar_indices(conn)


def _criar_tabelas_sqlite():
    with conectar() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS alunos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE,
                email TEXT UNIQUE,
                senha TEXT NOT NULL,
                perfil TEXT NOT NULL DEFAULT 'Aluno' CHECK(perfil IN ('Gestor','Aluno')),
                force_troca_senha INTEGER DEFAULT 1,
                ativo INTEGER DEFAULT 1
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS disciplinas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL UNIQUE,
                ativo INTEGER DEFAULT 1
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS aulas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                disciplina_id INTEGER NOT NULL,
                aula TEXT NOT NULL,
                assunto TEXT,
                estudada_padrao TEXT DEFAULT 'Não',
                revisao_24h_padrao TEXT DEFAULT 'Não',
                tipo_estudo TEXT DEFAULT 'Outro',
                ativo INTEGER DEFAULT 1,
                FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS assuntos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aula_id INTEGER NOT NULL,
                titulo TEXT NOT NULL,
                ativo INTEGER DEFAULT 1,
                UNIQUE(aula_id, titulo),
                FOREIGN KEY (aula_id) REFERENCES aulas(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tarefas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero INTEGER NOT NULL UNIQUE,
                trilha INTEGER,
                disciplina_id INTEGER NOT NULL,
                seq_disciplina INTEGER,
                aula TEXT,
                qtd_exercicios_previstos INTEGER DEFAULT 0,
                tipo TEXT,
                conteudo TEXT,
                ativo INTEGER DEFAULT 1,
                aula_id INTEGER,
                assunto_id INTEGER,
                FOREIGN KEY (disciplina_id) REFERENCES disciplinas(id),
                FOREIGN KEY (aula_id) REFERENCES aulas(id),
                FOREIGN KEY (assunto_id) REFERENCES assuntos(id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS execucoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aluno_id INTEGER NOT NULL,
                tarefa_id INTEGER NOT NULL,
                data_execucao TEXT,
                ch_efetiva REAL DEFAULT 0,
                data_revisao_24h TEXT,
                ch_revisao REAL DEFAULT 0,
                qtd_acertos INTEGER DEFAULT 0,
                desempenho REAL DEFAULT 0,
                comentario TEXT,
                concluida INTEGER DEFAULT 0,
                atualizado_em TEXT DEFAULT CURRENT_TIMESTAMP,
                qtd_questoes_feitas INTEGER DEFAULT 0,
                status TEXT DEFAULT 'NAO_INICIADA',
                tipo_estudo TEXT DEFAULT 'Outro',
                UNIQUE(aluno_id, tarefa_id),
                FOREIGN KEY (aluno_id) REFERENCES alunos(id),
                FOREIGN KEY (tarefa_id) REFERENCES tarefas(id)
            )
        """)
        # Migrações: adiciona colunas ausentes com segurança
        _migrar_coluna_sqlite(cur, "execucoes", "qtd_questoes_feitas", "INTEGER DEFAULT 0")
        _migrar_coluna_sqlite(cur, "execucoes", "status", "TEXT DEFAULT 'NAO_INICIADA'")
        _migrar_coluna_sqlite(cur, "execucoes", "tipo_estudo", "TEXT DEFAULT 'Outro'")
        _migrar_coluna_sqlite(cur, "aulas", "tipo_estudo", "TEXT DEFAULT 'Outro'")
        _migrar_coluna_sqlite(cur, "tarefas", "aula_id", "INTEGER")
        _migrar_coluna_sqlite(cur, "tarefas", "assunto_id", "INTEGER")
        _migrar_coluna_sqlite(cur, "alunos", "force_troca_senha", "INTEGER DEFAULT 0")
        # Tabela de sessões individuais de estudo
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sessoes_estudo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aluno_id INTEGER NOT NULL,
                tarefa_id INTEGER NOT NULL,
                data_sessao TEXT NOT NULL,
                ch_sessao REAL DEFAULT 0,
                qtd_questoes INTEGER DEFAULT 0,
                qtd_acertos INTEGER DEFAULT 0,
                tipo_estudo TEXT DEFAULT 'Outro',
                comentario TEXT,
                criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (aluno_id) REFERENCES alunos(id),
                FOREIGN KEY (tarefa_id) REFERENCES tarefas(id)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sessoes_aluno ON sessoes_estudo(aluno_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sessoes_tarefa ON sessoes_estudo(tarefa_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sessoes_data ON sessoes_estudo(data_sessao)")
        # Migração: popula sessoes_estudo com dados existentes (apenas uma vez por aluno+tarefa)
        cur.execute("""
            INSERT INTO sessoes_estudo
                (aluno_id, tarefa_id, data_sessao, ch_sessao, qtd_questoes, qtd_acertos, tipo_estudo, comentario)
            SELECT
                e.aluno_id, e.tarefa_id,
                COALESCE(e.data_execucao, date('now')),
                e.ch_efetiva, e.qtd_questoes_feitas, e.qtd_acertos,
                COALESCE(e.tipo_estudo, 'Outro'), e.comentario
            FROM execucoes e
            WHERE e.ch_efetiva > 0
              AND NOT EXISTS (
                  SELECT 1 FROM sessoes_estudo s
                  WHERE s.aluno_id = e.aluno_id AND s.tarefa_id = e.tarefa_id
              )
        """)
        _normalizar_status(conn)
        cur.execute("DELETE FROM alunos WHERE email = 'aluno@local.com' OR nome = 'Aluno Padrão'")
        _criar_indices(conn)


def _migrar_coluna_sqlite(cur, tabela, coluna, definicao):
    colunas = {row[1] for row in cur.execute(f"PRAGMA table_info({tabela})").fetchall()}
    if coluna not in colunas:
        cur.execute(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {definicao}")


def _normalizar_status(conn):
    conn.execute("""
        UPDATE execucoes
        SET status = CASE
            WHEN status IN ('NAO_INICIADA','EM_ANDAMENTO','CONCLUIDA') THEN status
            WHEN concluida = 1 THEN 'CONCLUIDA'
            ELSE 'NAO_INICIADA'
        END
    """)
    conn.execute("UPDATE execucoes SET concluida = CASE WHEN status = 'CONCLUIDA' THEN 1 ELSE 0 END")


def _criar_indices(conn):
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tarefas_disciplina ON tarefas(disciplina_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tarefas_trilha ON tarefas(trilha)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_exec_aluno ON execucoes(aluno_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_exec_tarefa ON execucoes(tarefa_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_exec_status ON execucoes(status)")


def inserir_admin():
    existe = consultar("SELECT id FROM alunos WHERE email = ?", (ADMIN_EMAIL,))
    if existe.empty:
        executar(
            "INSERT INTO alunos (nome, email, senha, perfil, ativo, force_troca_senha) VALUES (?, ?, ?, 'Gestor', 1, 1)",
            ("Administrador", ADMIN_EMAIL, hash_senha(ADMIN_PASSWORD)),
        )


# ─────────────────────────────────────────────
# CACHE
# ─────────────────────────────────────────────

def limpar_cache():
    carregar_visao_tarefas.clear()
    carregar_execucoes.clear()
    carregar_aulas.clear()
    carregar_assuntos.clear()
    carregar_tarefas_base.clear()
    carregar_sessoes_dashboard.clear()
    carregar_sessoes.clear()
    carregar_sessoes_aluno.clear()


# ─────────────────────────────────────────────
# UPSERTS
# ─────────────────────────────────────────────

def upsert_disciplina(conn, nome):
    nome = limpar_texto(nome)
    if not nome:
        return None
    conn.execute(
        "INSERT INTO disciplinas (nome, ativo) VALUES (?, 1) ON CONFLICT(nome) DO UPDATE SET ativo = 1",
        (nome,),
    )
    return conn.execute("SELECT id FROM disciplinas WHERE nome = ?", (nome,)).fetchone()[0]


def upsert_aula(conn, disciplina_id, aula, estudada="Não", revisao="Não", tipo_estudo="Outro"):
    aula = str(limpar_texto(aula) or "Aula não informada")
    existente = conn.execute(
        "SELECT id FROM aulas WHERE disciplina_id = ? AND aula = ? AND ativo = 1 ORDER BY id LIMIT 1",
        (int(disciplina_id), aula),
    ).fetchone()
    if existente:
        conn.execute(
            "UPDATE aulas SET estudada_padrao = ?, revisao_24h_padrao = ?, tipo_estudo = ?, ativo = 1 WHERE id = ?",
            (estudada or "Não", revisao or "Não", normalizar_tipo_estudo(tipo_estudo), int(existente[0])),
        )
        return int(existente[0])
    conn.execute(
        "INSERT INTO aulas (disciplina_id, aula, estudada_padrao, revisao_24h_padrao, tipo_estudo, ativo) VALUES (?, ?, ?, ?, ?, 1)",
        (int(disciplina_id), aula, estudada or "Não", revisao or "Não", normalizar_tipo_estudo(tipo_estudo)),
    )
    return ultimo_id(conn)


def upsert_assunto(conn, aula_id, titulo):
    titulo = limpar_texto(titulo) or "Conteúdo não informado"
    conn.execute(
        "INSERT INTO assuntos (aula_id, titulo, ativo) VALUES (?, ?, 1) ON CONFLICT(aula_id, titulo) DO UPDATE SET ativo = 1",
        (int(aula_id), titulo),
    )
    return conn.execute("SELECT id FROM assuntos WHERE aula_id = ? AND titulo = ?", (int(aula_id), titulo)).fetchone()[0]


def upsert_execucao(
    conn,
    aluno_id,
    tarefa_id,
    data_execucao=None,
    ch_efetiva=0,
    data_revisao=None,
    ch_revisao=0,
    acertos=0,
    comentario=None,
    questoes_feitas=0,
    status=STATUS_NAO_INICIADA,
    tipo_estudo=None,
):
    if status not in STATUS_VALIDOS:
        status = STATUS_NAO_INICIADA
    tipo_estudo = normalizar_tipo_estudo(tipo_estudo)
    if tipo_estudo == "Outro":
        tarefa_tipo = conn.execute(
            "SELECT COALESCE(tipo, 'Outro') FROM tarefas WHERE id = ?", (int(tarefa_id),)
        ).fetchone()
        tipo_estudo = normalizar_tipo_estudo(tarefa_tipo[0] if tarefa_tipo else None)
    questoes = int(round(questoes_feitas or 0))
    acertos = int(acertos or 0)
    desempenho = round((acertos / questoes) * 100, 2) if questoes > 0 else 0
    concluida = 1 if status == STATUS_CONCLUIDA else 0
    conn.execute(
        """
        INSERT INTO execucoes
            (aluno_id, tarefa_id, data_execucao, ch_efetiva, data_revisao_24h, ch_revisao,
             qtd_questoes_feitas, qtd_acertos, desempenho, status, comentario, concluida, tipo_estudo)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(aluno_id, tarefa_id) DO UPDATE SET
            data_execucao      = excluded.data_execucao,
            ch_efetiva         = excluded.ch_efetiva,
            data_revisao_24h   = excluded.data_revisao_24h,
            ch_revisao         = excluded.ch_revisao,
            qtd_questoes_feitas= excluded.qtd_questoes_feitas,
            qtd_acertos        = excluded.qtd_acertos,
            desempenho         = excluded.desempenho,
            status             = excluded.status,
            comentario         = excluded.comentario,
            concluida          = excluded.concluida,
            tipo_estudo        = excluded.tipo_estudo,
            atualizado_em      = CURRENT_TIMESTAMP
        """,
        (
            int(aluno_id), int(tarefa_id), data_execucao,
            float(ch_efetiva or 0), data_revisao, float(ch_revisao or 0),
            questoes, acertos, desempenho, status, comentario, concluida, tipo_estudo,
        ),
    )



# ─────────────────────────────────────────────
# SESSÕES DE ESTUDO
# ─────────────────────────────────────────────

def inserir_sessao(
    aluno_id: int,
    tarefa_id: int,
    data_sessao: str,
    ch_sessao: float,
    qtd_questoes: int = 0,
    qtd_acertos: int = 0,
    tipo_estudo: str = "Outro",
    comentario: str | None = None,
) -> int:
    """
    Insere uma nova sessão de estudo e atualiza o acumulado em execucoes.
    Retorna o id da sessão criada.
    """
    tipo_estudo = normalizar_tipo_estudo(tipo_estudo)
    with conectar() as conn:
        conn.execute(
            """
            INSERT INTO sessoes_estudo
                (aluno_id, tarefa_id, data_sessao, ch_sessao,
                 qtd_questoes, qtd_acertos, tipo_estudo, comentario)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (int(aluno_id), int(tarefa_id), str(data_sessao),
             float(ch_sessao or 0), int(qtd_questoes or 0), int(qtd_acertos or 0),
             tipo_estudo, limpar_texto(comentario)),
        )
        sessao_id = ultimo_id(conn)

        # Recalcula acumulados em execucoes a partir das sessões
        _recalcular_execucao_por_sessoes(conn, aluno_id, tarefa_id)

    limpar_cache()
    return sessao_id


def excluir_sessao(sessao_id: int, aluno_id: int, tarefa_id: int) -> None:
    """Remove uma sessão e recalcula o acumulado em execucoes."""
    with conectar() as conn:
        conn.execute("DELETE FROM sessoes_estudo WHERE id = ?", (int(sessao_id),))
        _recalcular_execucao_por_sessoes(conn, aluno_id, tarefa_id)
    limpar_cache()


def _recalcular_execucao_por_sessoes(conn, aluno_id: int, tarefa_id: int) -> None:
    """
    Soma todas as sessões da tarefa+aluno e atualiza ch_efetiva,
    qtd_questoes_feitas, qtd_acertos e desempenho em execucoes.
    Preserva status, tipo_estudo, data_execucao (usa a data da sessão mais recente).
    """
    # Garante que a linha de execução existe antes de atualizar.
    # Sem isto, uma tarefa importada e ainda não vinculada ao aluno recebia a
    # sessão em sessoes_estudo e o UPDATE abaixo afetava ZERO linhas: as horas
    # ficavam órfãs, sem erro nenhum e com mensagem de sucesso na tela.
    conn.execute(
        """
        INSERT INTO execucoes (aluno_id, tarefa_id, status, concluida)
        VALUES (?, ?, 'NAO_INICIADA', 0)
        ON CONFLICT(aluno_id, tarefa_id) DO NOTHING
        """,
        (int(aluno_id), int(tarefa_id)),
    )

    row = conn.execute(
        """
        SELECT
            COALESCE(SUM(ch_sessao), 0),
            COALESCE(SUM(qtd_questoes), 0),
            COALESCE(SUM(qtd_acertos), 0),
            MAX(data_sessao)
        FROM sessoes_estudo
        WHERE aluno_id = ? AND tarefa_id = ?
        """,
        (int(aluno_id), int(tarefa_id)),
    ).fetchone()

    ch_total    = float(row[0] or 0)
    q_total     = int(row[1] or 0)
    a_total     = int(row[2] or 0)
    data_ultima = row[3]
    desempenho  = round(a_total / q_total * 100, 2) if q_total > 0 else 0

    conn.execute(
        """
        UPDATE execucoes SET
            ch_efetiva          = ?,
            qtd_questoes_feitas = ?,
            qtd_acertos         = ?,
            desempenho          = ?,
            data_execucao       = COALESCE(?, data_execucao),
            atualizado_em       = CURRENT_TIMESTAMP
        WHERE aluno_id = ? AND tarefa_id = ?
        """,
        (ch_total, q_total, a_total, desempenho,
         data_ultima, int(aluno_id), int(tarefa_id)),
    )


@st.cache_data(ttl=20)
def carregar_sessoes(aluno_id: int, tarefa_id: int) -> pd.DataFrame:
    """Retorna todas as sessões de uma tarefa de um aluno, mais recentes primeiro."""
    return consultar(
        """
        SELECT
            s.id, s.data_sessao, s.ch_sessao,
            s.qtd_questoes, s.qtd_acertos, s.tipo_estudo, s.comentario, s.criado_em,
            CASE WHEN s.qtd_questoes > 0
                 THEN ROUND(CAST(s.qtd_acertos AS NUMERIC) / s.qtd_questoes * 100, 1)
                 ELSE 0 END AS desempenho_sessao
        FROM sessoes_estudo s
        WHERE s.aluno_id = ? AND s.tarefa_id = ?
        ORDER BY s.data_sessao DESC, s.criado_em DESC
        """,
        (int(aluno_id), int(tarefa_id)),
    )


@st.cache_data(ttl=20)
def carregar_sessoes_aluno(aluno_id: int) -> pd.DataFrame:
    """Retorna todas as sessões de um aluno com info de tarefa/disciplina."""
    return consultar(
        """
        SELECT
            s.id, s.data_sessao, s.ch_sessao,
            s.qtd_questoes, s.qtd_acertos, s.tipo_estudo, s.comentario,
            t.numero AS tarefa, d.nome AS disciplina,
            COALESCE(ass.titulo, t.conteudo) AS assunto
        FROM sessoes_estudo s
        JOIN tarefas t ON t.id = s.tarefa_id
        JOIN disciplinas d ON d.id = t.disciplina_id
        LEFT JOIN assuntos ass ON ass.id = t.assunto_id
        WHERE s.aluno_id = ?
        ORDER BY s.data_sessao DESC, s.criado_em DESC
        """,
        (int(aluno_id),),
    )


# ─────────────────────────────────────────────
# IMPORTADORES
# ─────────────────────────────────────────────

def importar_planilha_referencia(caminho=PLANILHA_REFERENCIA, substituir=False):
    """
    Importa a planilha de referência (disciplinas, aulas, assuntos, tarefas).

    NUNCA apaga execuções existentes.
    Apenas insere ou atualiza estrutura (disciplinas, aulas, assuntos, tarefas).
    Execuções já existentes (status, datas, horas, acertos) são preservadas.
    O parâmetro `substituir` foi mantido por compatibilidade mas não apaga execuções.
    """
    caminho = Path(caminho)
    if not caminho.exists():
        st.error(f"Planilha não encontrada: {caminho}")
        return False
    try:
        xl = pd.ExcelFile(caminho)
    except Exception as exc:
        erro_usuario("Erro ao abrir a planilha.", exc)
        return False
    if "Tarefas" not in xl.sheet_names:
        st.error("A planilha precisa ter a aba Tarefas.")
        return False

    abas_ignoradas = {"Tutorial", "Estatísticas", "Tarefas"}
    try:
        with conectar() as conn:
            # ── Estrutura: disciplinas, aulas, assuntos ──
            # Nunca apaga. Apenas insere novos ou reactiva inativados.
            for aba in xl.sheet_names:
                if aba in abas_ignoradas:
                    continue
                try:
                    raw = pd.read_excel(caminho, sheet_name=aba, header=None, nrows=1)
                    nome_disciplina = limpar_texto(raw.iloc[0, 0]) or aba
                except Exception:
                    nome_disciplina = aba
                disciplina_id = upsert_disciplina(conn, nome_disciplina)

                try:
                    aulas_df = pd.read_excel(caminho, sheet_name=aba, header=1).dropna(how="all")
                except Exception:
                    continue
                if "Aula" not in aulas_df.columns or "Assunto" not in aulas_df.columns:
                    continue
                for _, row in aulas_df.iterrows():
                    aula    = limpar_texto(row.get("Aula"))
                    assunto = limpar_texto(row.get("Assunto"))
                    if not aula or not assunto:
                        continue
                    aula_id = upsert_aula(
                        conn, disciplina_id, aula,
                        limpar_texto(row.get("Estudada")) or "Não",
                        limpar_texto(row.get("Revisão 24h")) or "Não",
                    )
                    upsert_assunto(conn, aula_id, assunto)

            # ── Tarefas: upsert seguro — não toca em execuções ──
            tarefas = pd.read_excel(caminho, sheet_name="Tarefas", header=2).dropna(how="all")
            for _, row in tarefas.iterrows():
                numero     = converter_inteiro(row.get("Tarefa"))
                disciplina = limpar_texto(row.get("Disciplina"))
                if not numero or not disciplina:
                    continue
                disciplina_id = upsert_disciplina(conn, disciplina)
                aula_valor    = str(limpar_texto(row.get("Aula")) or "Aula não informada")
                aula_id       = upsert_aula(conn, disciplina_id, aula_valor)
                assunto_id    = upsert_assunto(conn, aula_id, limpar_texto(row.get("Conteúdo")))
                tipo          = normalizar_tipo_estudo(row.get("Tipo"))
                conn.execute(
                    """
                    INSERT INTO tarefas
                        (numero, trilha, disciplina_id, seq_disciplina, aula,
                         qtd_exercicios_previstos, tipo, conteudo, ativo, aula_id, assunto_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    ON CONFLICT(numero) DO UPDATE SET
                        trilha                  = excluded.trilha,
                        disciplina_id           = excluded.disciplina_id,
                        seq_disciplina          = excluded.seq_disciplina,
                        aula                    = excluded.aula,
                        qtd_exercicios_previstos= excluded.qtd_exercicios_previstos,
                        tipo                    = excluded.tipo,
                        conteudo                = excluded.conteudo,
                        ativo                   = 1,
                        aula_id                 = excluded.aula_id,
                        assunto_id              = excluded.assunto_id
                    """,
                    (
                        numero,
                        converter_inteiro(row.get("Trilha")),
                        disciplina_id,
                        converter_inteiro(row.get("Seq Disciplina")),
                        aula_valor,
                        converter_inteiro(row.get("Qtd. Exercícios")),
                        tipo,
                        limpar_texto(row.get("Conteúdo")),
                        aula_id,
                        assunto_id,
                    ),
                )

            # ── Vincula novas tarefas aos alunos ativos ──
            # Para cada tarefa que não tem execução para um aluno,
            # cria um vínculo NAO_INICIADA.
            # Tarefas que já têm execução (qualquer status) são ignoradas.
            # Isso garante que o dashboard enxergue as novas tarefas
            # sem afetar nenhum registro existente.
            # ── Realinha o tipo de estudo das tarefas AINDA NÃO INICIADAS ──
            # Só toca em quem tem status NAO_INICIADA e nenhuma hora/questão
            # registrada. Qualquer tarefa com estudo lançado fica intacta:
            # o tipo escolhido pelo aluno na hora de estudar é preservado.
            conn.execute(
                """
                UPDATE execucoes SET tipo_estudo = (
                    SELECT COALESCE(t.tipo, 'Outro') FROM tarefas t
                    WHERE t.id = execucoes.tarefa_id
                )
                WHERE status = 'NAO_INICIADA'
                  AND COALESCE(ch_efetiva, 0) = 0
                  AND COALESCE(qtd_questoes_feitas, 0) = 0
                  AND NOT EXISTS (
                      SELECT 1 FROM sessoes_estudo s
                      WHERE s.aluno_id = execucoes.aluno_id
                        AND s.tarefa_id = execucoes.tarefa_id
                  )
                """
            )

            alunos_ativos_ids = [
                row[0] for row in conn.execute(
                    "SELECT id FROM alunos WHERE ativo = 1 AND perfil = 'Aluno'"
                ).fetchall()
            ]
            if alunos_ativos_ids:
                conn.execute(
                    """
                    INSERT INTO execucoes (aluno_id, tarefa_id, status, concluida, tipo_estudo)
                    SELECT a.id, t.id, 'NAO_INICIADA', 0, COALESCE(t.tipo, 'Outro')
                    FROM tarefas t
                    CROSS JOIN (
                        SELECT id FROM alunos WHERE ativo = 1 AND perfil = 'Aluno'
                    ) a
                    WHERE t.ativo = 1
                      AND NOT EXISTS (
                          SELECT 1 FROM execucoes e
                          WHERE e.aluno_id = a.id AND e.tarefa_id = t.id
                      )
                    """
                )

    except Exception as exc:
        erro_usuario("Importação cancelada.", exc)
        return False
    limpar_cache()
    return True


def vincular_tarefas_pendentes() -> int:
    """
    Cria vínculos NAO_INICIADA para tarefas que ainda não têm
    nenhuma execução registrada para os alunos ativos.

    Regras:
    - Só cria vínculo quando NÃO existe execução alguma (qualquer status).
    - Não toca em execuções existentes — histórico 100% preservado.
    - Retorna a quantidade de novos vínculos criados.
    """
    with conectar() as conn:
        antes = conn.execute("SELECT COUNT(*) FROM execucoes").fetchone()[0]
        conn.execute(
            """
            INSERT INTO execucoes (aluno_id, tarefa_id, status, concluida, tipo_estudo)
            SELECT a.id, t.id, 'NAO_INICIADA', 0, COALESCE(t.tipo, 'Outro')
            FROM tarefas t
            CROSS JOIN (
                SELECT id FROM alunos WHERE ativo = 1 AND perfil = 'Aluno'
            ) a
            WHERE t.ativo = 1
              AND NOT EXISTS (
                  SELECT 1 FROM execucoes e
                  WHERE e.aluno_id = a.id AND e.tarefa_id = t.id
              )
            """
        )
        depois = conn.execute("SELECT COUNT(*) FROM execucoes").fetchone()[0]
    limpar_cache()
    return int(depois - antes)


def _status_ciclo_para_interno(status_planilha: str) -> str:
    """Converte status da planilha Ciclo Consolidado para o status interno."""
    mapa = {
        "concluido":     STATUS_CONCLUIDA,
        "concluído":     STATUS_CONCLUIDA,
        "em andamento":  STATUS_EM_ANDAMENTO,
        "nao iniciado":  STATUS_NAO_INICIADA,
        "não iniciado":  STATUS_NAO_INICIADA,
    }
    return mapa.get(chave_texto(str(status_planilha or "")), STATUS_NAO_INICIADA)


def _data_segura(valor) -> str | None:
    """Converte datetime/string do Excel para 'YYYY-MM-DD' ou None."""
    if valor is None:
        return None
    if isinstance(valor, str) and valor.strip() in ("", "—", "-"):
        return None
    try:
        ts = pd.to_datetime(valor)
        if pd.isnull(ts):
            return None
        return str(ts.date())
    except Exception:
        return None


def importar_ciclo_consolidado(caminho_excel, modo: str = "substituir") -> dict:
    """
    Importa o arquivo 'Ciclo Consolidado' (formato com dois alunos por linha).

    Estrutura esperada — aba CICLO_CONSOLIDADO, linha de cabeçalho na linha 3:
        col 0  BLOCO
        col 1  DISCIPLINA
        col 2  OBJETIVO DO BLOCO
        col 3  STATUS          (Concluído / Em Andamento / Não Iniciado)
        col 4  DATA ESTUDO (PROGRAMADA)
        col 5  (mesclada — ignorada)
        col 6  CH EFETIVA (LILIAN)
        col 7  AULA ATUAL (LILIAN)
        col 8  QUESTÕES (L)
        col 9  ACERTOS (L)
        col 10 CH EFETIVA (JESSICA)
        col 11 AULA ATUAL (JESSICA)
        col 12 QUESTÕES (J)
        col 13 ACERTOS (J)
        col 14 CH TOTAL
        col 15 TOTAL QUESTÕES
        col 16 TOTAL ACERTOS
        col 17 DESEMPENHO (%)

    Retorna dict com {ok, registros, erros, avisos}.
    """
    caminho_excel = Path(caminho_excel)
    resultado = {"ok": False, "registros": 0, "erros": [], "avisos": []}

    if not caminho_excel.exists():
        resultado["erros"].append(f"Arquivo não encontrado: {caminho_excel}")
        return resultado

    # ── Detecta os nomes dos alunos na linha 2 (índice 1) ──
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(caminho_excel), read_only=True, data_only=True)
    except Exception as exc:
        resultado["erros"].append(f"Erro ao abrir arquivo: {exc}")
        return resultado

    if "CICLO_CONSOLIDADO" not in wb.sheetnames:
        resultado["erros"].append("Aba 'CICLO_CONSOLIDADO' não encontrada. Verifique o arquivo.")
        return resultado

    ws = wb["CICLO_CONSOLIDADO"]

    # Linha 2 (índice 1) → identificação dos alunos
    # Formato: ['IDENTIFICAÇÃO', None, ..., 'LILIAN', None, ..., 'JESSICA', None, ..., 'CONSOLIDADO', ...]
    todas_linhas = list(ws.iter_rows(values_only=True))
    if len(todas_linhas) < 4:
        resultado["erros"].append("Arquivo com poucas linhas — formato inesperado.")
        return resultado

    linha_identificacao = list(todas_linhas[1])  # linha 2
    # Detecta nomes de alunos nas colunas 6 e 10 (posições fixas do formato)
    nome_aluno_a = limpar_texto(linha_identificacao[6]) if len(linha_identificacao) > 6 else None
    nome_aluno_b = limpar_texto(linha_identificacao[10]) if len(linha_identificacao) > 10 else None

    if not nome_aluno_a:
        nome_aluno_a = "Aluno A"
        resultado["avisos"].append("Nome do primeiro aluno não detectado — usando 'Aluno A'.")
    if not nome_aluno_b:
        nome_aluno_b = "Aluno B"
        resultado["avisos"].append("Nome do segundo aluno não detectado — usando 'Aluno B'.")

    # Linhas de dados: a partir da linha 4 (índice 3), pula cabeçalho e linhas de bloco sem disciplina
    linhas_dados = []
    bloco_atual = None
    for linha in todas_linhas[3:]:
        bloco_val = limpar_texto(linha[0]) if linha[0] else None
        disc_val  = limpar_texto(linha[1]) if len(linha) > 1 else None

        # Atualiza bloco atual (ffill)
        if bloco_val and bloco_val != "TOTAIS GERAIS":
            bloco_atual = bloco_val

        # Só processa linhas com disciplina e bloco válido
        if not disc_val or not bloco_atual:
            continue
        if bloco_atual == "TOTAIS GERAIS":
            continue

        linhas_dados.append((bloco_atual, linha))

    if not linhas_dados:
        resultado["erros"].append("Nenhuma linha de dado encontrada no arquivo.")
        return resultado

    # ── Garante alunos no banco ──
    try:
        with conectar() as conn:
            for nome in [nome_aluno_a, nome_aluno_b]:
                # Verifica se já existe pelo nome antes de tentar inserir,
                # evitando conflito de unique no email em PostgreSQL.
                ja_existe = conn.execute(
                    "SELECT id FROM alunos WHERE nome = ?", (nome,)
                ).fetchone()
                if ja_existe:
                    # Apenas reativa se estiver inativo
                    conn.execute(
                        "UPDATE alunos SET ativo = 1 WHERE nome = ?", (nome,)
                    )
                else:
                    # Gera email e verifica se também já está em uso
                    email = email_local(nome)
                    email_em_uso = conn.execute(
                        "SELECT id FROM alunos WHERE email = ?", (email,)
                    ).fetchone()
                    if email_em_uso:
                        # Usa email alternativo com sufixo numérico
                        import time as _time
                        email = f"{email_local(nome).split('@')[0]}_{int(_time.time()) % 10000}@local.com"
                    conn.execute(
                        """
                        INSERT INTO alunos (nome, email, senha, perfil, ativo, force_troca_senha)
                        VALUES (?, ?, ?, 'Aluno', 1, 1)
                        """,
                        (nome, email, hash_senha("123")),
                    )

            id_a = conn.execute("SELECT id FROM alunos WHERE nome = ?", (nome_aluno_a,)).fetchone()[0]
            id_b = conn.execute("SELECT id FROM alunos WHERE nome = ?", (nome_aluno_b,)).fetchone()[0]

            if modo == "substituir":
                # Apaga também as sessões: sem isso o banco ficava inconsistente —
                # execucoes reconstruída pela planilha e sessoes_estudo intacta,
                # até a próxima sessão registrada reescrever tudo pelo somatório antigo.
                conn.execute("DELETE FROM sessoes_estudo WHERE aluno_id IN (?, ?)", (id_a, id_b))
                conn.execute("DELETE FROM execucoes WHERE aluno_id IN (?, ?)", (id_a, id_b))

            # Carrega tarefas do banco indexadas por (trilha, chave_disciplina)
            tarefas_db = conn.execute(
                """
                SELECT t.id, COALESCE(t.trilha, 0), d.nome,
                       COALESCE(t.conteudo, ''), COALESCE(t.qtd_exercicios_previstos, 0)
                FROM tarefas t
                JOIN disciplinas d ON d.id = t.disciplina_id
                WHERE t.ativo = 1 AND d.ativo = 1
                ORDER BY t.numero
                """
            ).fetchall()

            # Aliases de normalização para disciplinas com variações de nome
            aliases_disc = {
                "matematica e raciocinio logico":               "matematica e raciocinio logico",
                "matematica e raciocínio lógico":               "matematica e raciocinio logico",
                "resolucao de questoes":                        "resolucao de questoes",
                "resolução de questões":                        "resolucao de questoes",
                "resolucao de questoes erradas":                "resolucao de questoes erradas",
                "resolução de questões erradas":                "resolucao de questoes erradas",
                "administracao geral e publica":                "administracao geral e publica",
                "administração geral e pública":                "administracao geral e publica",
                "administracao financeira e orcamentaria":      "administracao financeira e orcamentaria",
                "administração financeira e orçamentária":      "administracao financeira e orcamentaria",
                "contabilidade publica":                        "contabilidade publica",
                "contabilidade pública":                        "contabilidade publica",
            }

            por_chave: dict = {}
            for tid, trilha, disc_nome, conteudo, previstos in tarefas_db:
                ck = aliases_disc.get(chave_texto(disc_nome), chave_texto(disc_nome))
                por_chave.setdefault((int(trilha or 0), ck), []).append(
                    (int(tid), int(previstos or 0), limpar_texto(conteudo) or "")
                )

            registros_gravados = 0

            for bloco, linha in linhas_dados:
                # Extrai número do bloco → trilha
                m = re.search(r"(\d+)", str(bloco))
                if not m:
                    resultado["avisos"].append(f"Bloco sem número: '{bloco}' — ignorado.")
                    continue
                trilha_num = int(m.group(1))

                disciplina_raw = limpar_texto(linha[1])
                ck_disc = aliases_disc.get(chave_texto(disciplina_raw), chave_texto(disciplina_raw))

                status_planilha = limpar_texto(linha[3]) if len(linha) > 3 else None
                status_interno  = _status_ciclo_para_interno(status_planilha or "")
                data_programada = _data_segura(linha[4] if len(linha) > 4 else None)

                # Dados Aluno A
                ch_a      = converter_horas(linha[6]  if len(linha) > 6  else None)
                aula_a    = limpar_texto(linha[7]  if len(linha) > 7  else None)
                quest_a   = converter_inteiro(linha[8]  if len(linha) > 8  else None)
                acert_a   = converter_inteiro(linha[9]  if len(linha) > 9  else None)

                # Dados Aluno B
                ch_b      = converter_horas(linha[10] if len(linha) > 10 else None)
                aula_b    = limpar_texto(linha[11] if len(linha) > 11 else None)
                quest_b   = converter_inteiro(linha[12] if len(linha) > 12 else None)
                acert_b   = converter_inteiro(linha[13] if len(linha) > 13 else None)

                # Encontra tarefas correspondentes no banco
                tarefas_match = por_chave.get((trilha_num, ck_disc), [])
                if not tarefas_match:
                    resultado["avisos"].append(
                        f"Sem tarefa no banco para Bloco {trilha_num} / {disciplina_raw} — linha ignorada."
                    )
                    continue

                # Distribui questões/acertos proporcionalmente entre as tarefas do bloco
                pesos = [p for _, p, _ in tarefas_match]

                for aluno_id, ch_aluno, quest_aluno, acert_aluno, aula_aluno in [
                    (id_a, ch_a, quest_a, acert_a, aula_a),
                    (id_b, ch_b, quest_b, acert_b, aula_b),
                ]:
                    tem_dado = ch_aluno > 0 or quest_aluno > 0 or acert_aluno > 0 or bool(aula_aluno)
                    if not tem_dado and status_interno == STATUS_NAO_INICIADA:
                        # Garante que o vínculo existe como NAO_INICIADA sem sobrescrever dados
                        for tid, _, _ in tarefas_match:
                            upsert_execucao(
                                conn, aluno_id, tid,
                                data_programada, 0, None, 0, 0,
                                None, 0, STATUS_NAO_INICIADA,
                            )
                        continue

                    qtd_tarefas = len(tarefas_match)
                    ch_por = ch_aluno / qtd_tarefas if qtd_tarefas and ch_aluno else 0
                    qdist  = distribuir_inteiro(quest_aluno, pesos)
                    adist  = distribuir_inteiro(acert_aluno, qdist)

                    for idx, (tid, _, _) in enumerate(tarefas_match):
                        upsert_execucao(
                            conn, aluno_id, tid,
                            data_programada,
                            ch_por, None, 0,
                            adist[idx],
                            aula_aluno,
                            qdist[idx],
                            status_interno,
                        )
                        registros_gravados += 1

    except Exception as exc:
        resultado["erros"].append(f"Erro durante importação: {exc}")
        if DEBUG:
            import traceback
            resultado["erros"].append(traceback.format_exc())
        return resultado

    limpar_cache()
    resultado["ok"] = True
    resultado["registros"] = registros_gravados
    return resultado


# ─────────────────────────────────────────────
# CONSULTAS COM CACHE
# ─────────────────────────────────────────────

@st.cache_data(ttl=20)
def carregar_visao_tarefas(aluno_id=None):
    filtro = ""
    params = []
    if aluno_id:
        filtro = "AND e.aluno_id = ?"
        params.append(int(aluno_id))
    return consultar(
        f"""
        SELECT
            t.id AS tarefa_id,
            t.numero AS tarefa,
            t.trilha,
            d.nome AS disciplina,
            t.disciplina_id,
            t.aula_id,
            t.assunto_id,
            COALESCE(ass.titulo, t.conteudo) AS assunto,
            t.seq_disciplina,
            COALESCE(a.aula, t.aula) AS aula,
            t.qtd_exercicios_previstos,
            CASE
                WHEN COALESCE(e.status, 'NAO_INICIADA') = 'NAO_INICIADA'
                    THEN COALESCE(t.tipo, 'Outro')
                ELSE COALESCE(NULLIF(e.tipo_estudo, 'Outro'), t.tipo, 'Outro')
            END AS tipo,
            t.conteudo,
            e.id AS execucao_id,
            e.aluno_id,
            al.nome AS aluno,
            e.data_execucao,
            COALESCE(e.ch_efetiva, 0) AS ch_efetiva,
            e.data_revisao_24h,
            COALESCE(e.ch_revisao, 0) AS ch_revisao,
            COALESCE(e.qtd_questoes_feitas, 0) AS qtd_questoes_feitas,
            COALESCE(e.qtd_acertos, 0) AS qtd_acertos,
            COALESCE(e.desempenho, 0) AS desempenho,
            COALESCE(e.status, 'NAO_INICIADA') AS status,
            e.comentario,
            e.atualizado_em,
            CASE WHEN COALESCE(e.status, 'NAO_INICIADA') = 'CONCLUIDA' THEN 1 ELSE 0 END AS concluida
        FROM tarefas t
        JOIN disciplinas d ON d.id = t.disciplina_id AND d.ativo = 1
        LEFT JOIN aulas a ON a.id = t.aula_id AND a.ativo = 1
        LEFT JOIN assuntos ass ON ass.id = t.assunto_id AND ass.ativo = 1
        LEFT JOIN execucoes e ON e.tarefa_id = t.id {filtro}
        LEFT JOIN alunos al ON al.id = e.aluno_id
        WHERE t.ativo = 1
        ORDER BY t.numero
        """,
        tuple(params),
    )


@st.cache_data(ttl=20)
def carregar_execucoes():
    """
    Execuções ordenadas pela data real de estudo, da mais recente para a
    mais antiga.

    A ordenação usa `ultima_atividade` = a data da sessão de estudo mais
    recente da tarefa, com `data_execucao` como reserva para registros
    antigos, anteriores às sessões.

    Não usa `atualizado_em`: esse campo é carimbado sempre que a linha é
    tocada, inclusive por uma importação de planilha, o que faria tarefas
    nunca estudadas subirem para o topo da lista.

    Tarefas sem nenhuma atividade vão para o fim.
    """
    return consultar(
        """
        SELECT
            e.id AS execucao_id,
            e.aluno_id,
            al.nome AS aluno,
            e.tarefa_id,
            t.numero AS tarefa,
            t.trilha,
            d.nome AS disciplina,
            t.disciplina_id,
            COALESCE(a.aula, t.aula) AS aula,
            t.aula_id,
            COALESCE(ass.titulo, t.conteudo) AS assunto,
            t.assunto_id,
            CASE
                WHEN COALESCE(e.status, 'NAO_INICIADA') = 'NAO_INICIADA'
                    THEN COALESCE(t.tipo, 'Outro')
                ELSE COALESCE(NULLIF(e.tipo_estudo, 'Outro'), t.tipo, 'Outro')
            END AS tipo,
            t.conteudo,
            t.qtd_exercicios_previstos,
            e.data_execucao,
            e.ch_efetiva,
            e.data_revisao_24h,
            e.ch_revisao,
            e.qtd_questoes_feitas,
            e.qtd_acertos,
            e.desempenho,
            COALESCE(e.status, 'NAO_INICIADA') AS status,
            e.comentario,
            e.atualizado_em,
            COALESCE((SELECT MAX(s.data_sessao) FROM sessoes_estudo s WHERE s.aluno_id = e.aluno_id AND s.tarefa_id = e.tarefa_id), e.data_execucao) AS ultima_atividade
        FROM execucoes e
        JOIN alunos al ON al.id = e.aluno_id AND al.ativo = 1 AND al.perfil = 'Aluno'
        JOIN tarefas t ON t.id = e.tarefa_id AND t.ativo = 1
        JOIN disciplinas d ON d.id = t.disciplina_id AND d.ativo = 1
        LEFT JOIN aulas a ON a.id = t.aula_id AND a.ativo = 1
        LEFT JOIN assuntos ass ON ass.id = t.assunto_id AND ass.ativo = 1
        ORDER BY
            CASE WHEN COALESCE((SELECT MAX(s.data_sessao) FROM sessoes_estudo s WHERE s.aluno_id = e.aluno_id AND s.tarefa_id = e.tarefa_id), e.data_execucao) IS NULL THEN 1 ELSE 0 END,
            COALESCE((SELECT MAX(s.data_sessao) FROM sessoes_estudo s WHERE s.aluno_id = e.aluno_id AND s.tarefa_id = e.tarefa_id), e.data_execucao) DESC,
            e.atualizado_em DESC, al.nome, t.numero
        """
    )


@st.cache_data(ttl=20)
def carregar_sessoes_dashboard() -> pd.DataFrame:
    """
    Retorna todas as sessões de estudo com info de aluno/tarefa/disciplina.
    Usado pelo dashboard para métricas temporais corretas (horas por período).
    Cada linha é uma sessão real — não o acumulado.
    """
    return consultar(
        """
        SELECT
            s.id AS sessao_id,
            s.aluno_id,
            al.nome AS aluno,
            s.tarefa_id,
            t.numero AS tarefa,
            d.nome AS disciplina,
            t.disciplina_id,
            COALESCE(a.aula, t.aula) AS aula,
            COALESCE(ass.titulo, t.conteudo) AS assunto,
            COALESCE(s.tipo_estudo, t.tipo, 'Outro') AS tipo,
            s.data_sessao AS data_execucao,
            s.ch_sessao AS ch_efetiva,
            s.qtd_questoes AS qtd_questoes_feitas,
            s.qtd_acertos,
            CASE WHEN s.qtd_questoes > 0
                 THEN ROUND(CAST(s.qtd_acertos AS NUMERIC) / s.qtd_questoes * 100, 2)
                 ELSE 0 END AS desempenho,
            s.comentario,
            e.status,
            e.atualizado_em
        FROM sessoes_estudo s
        JOIN alunos al ON al.id = s.aluno_id AND al.ativo = 1 AND al.perfil = 'Aluno'
        JOIN tarefas t ON t.id = s.tarefa_id AND t.ativo = 1
        JOIN disciplinas d ON d.id = t.disciplina_id AND d.ativo = 1
        LEFT JOIN aulas a ON a.id = t.aula_id AND a.ativo = 1
        LEFT JOIN assuntos ass ON ass.id = t.assunto_id AND ass.ativo = 1
        LEFT JOIN execucoes e ON e.aluno_id = s.aluno_id AND e.tarefa_id = s.tarefa_id
        ORDER BY s.data_sessao DESC, s.criado_em DESC
        """
    )


@st.cache_data(ttl=60)
def carregar_tarefas_base():
    return consultar(
        """
        SELECT
            t.id AS tarefa_id,
            t.numero AS tarefa,
            t.trilha,
            t.disciplina_id,
            d.nome AS disciplina,
            t.aula_id,
            COALESCE(a.aula, t.aula) AS aula,
            t.assunto_id,
            COALESCE(ass.titulo, t.conteudo) AS assunto,
            t.seq_disciplina,
            t.qtd_exercicios_previstos,
            COALESCE(t.tipo, 'Outro') AS tipo,
            t.conteudo
        FROM tarefas t
        JOIN disciplinas d ON d.id = t.disciplina_id AND d.ativo = 1
        LEFT JOIN aulas a ON a.id = t.aula_id AND a.ativo = 1
        LEFT JOIN assuntos ass ON ass.id = t.assunto_id AND ass.ativo = 1
        WHERE t.ativo = 1
        ORDER BY t.numero
        """
    )


@st.cache_data(ttl=60)
def carregar_aulas():
    return consultar(
        """
        SELECT a.id, a.disciplina_id, d.nome AS disciplina, a.aula,
               COALESCE(a.tipo_estudo, 'Outro') AS tipo_estudo,
               a.estudada_padrao, a.revisao_24h_padrao
        FROM aulas a
        JOIN disciplinas d ON d.id = a.disciplina_id
        WHERE a.ativo = 1 AND d.ativo = 1
        ORDER BY d.nome, a.aula
        """
    )


@st.cache_data(ttl=60)
def carregar_assuntos():
    return consultar(
        """
        SELECT ass.id, ass.aula_id, d.nome AS disciplina, a.aula, ass.titulo AS assunto
        FROM assuntos ass
        JOIN aulas a ON a.id = ass.aula_id
        JOIN disciplinas d ON d.id = a.disciplina_id
        WHERE ass.ativo = 1 AND a.ativo = 1 AND d.ativo = 1
        ORDER BY d.nome, a.aula, ass.titulo
        """
    )


def alunos_ativos(incluir_gestor=False):
    filtro = "" if incluir_gestor else "AND perfil = 'Aluno'"
    return consultar(f"SELECT id, nome, email, perfil FROM alunos WHERE ativo = 1 {filtro} ORDER BY nome")


def aluno_logado():
    return st.session_state["usuario"]


def disciplinas_ativas():
    return consultar("SELECT id, nome FROM disciplinas WHERE ativo = 1 ORDER BY nome")


# ─────────────────────────────────────────────
# FILTROS E MÉTRICAS
# ─────────────────────────────────────────────

def periodo_datas(periodo):
    hoje = date.today()
    if periodo == "Hoje":
        return hoje, hoje
    if periodo == "Semana":
        return hoje - timedelta(days=hoje.weekday()), hoje
    if periodo == "Mês":
        return hoje.replace(day=1), hoje
    if periodo == "Ano":
        return hoje.replace(month=1, day=1), hoje
    return None, None


def painel_filtros(df, prefixo="dash"):
    """
    Retorna quatro valores:
        df_escopo   – filtrado por aluno/disciplina/status/aula/assunto/tipo
                      SEM filtro de período. Usado para contagens estruturais:
                      total de tarefas, progresso, restantes, em andamento.
        df_periodo  – df_escopo com filtro de período e mínimo de horas aplicado.
                      Usado para métricas temporais: horas, questões, acertos.
        visao       – nome do aluno selecionado (ou "Todos")
        inicio, fim – datas do período selecionado (ou None)
    """
    st.sidebar.markdown("### Filtros")
    if st.sidebar.button("Limpar filtros", use_container_width=True, key=f"{prefixo}_limpar"):
        for chave in list(st.session_state.keys()):
            if chave.startswith(f"{prefixo}_"):
                del st.session_state[chave]
        st.rerun()

    favoritos = st.session_state.setdefault("filtros_favoritos", {})
    favoritos_opcoes = ["Nenhum"] + list(favoritos.keys())
    favorito = st.sidebar.selectbox("Filtro favorito", favoritos_opcoes, key=f"{prefixo}_favorito")
    if favorito != "Nenhum" and st.sidebar.button("Aplicar favorito", use_container_width=True, key=f"{prefixo}_aplicar"):
        for chave, valor in favoritos[favorito].items():
            st.session_state[f"{prefixo}_{chave}"] = valor
        st.rerun()

    periodo = st.sidebar.selectbox(
        "Período", ["Todos", "Hoje", "Semana", "Mês", "Ano", "Personalizado"], key=f"{prefixo}_periodo"
    )
    inicio_padrao, fim_padrao = periodo_datas(periodo)
    if periodo == "Personalizado":
        inicio = st.sidebar.date_input("Data inicial", value=inicio_padrao or date.today(), key=f"{prefixo}_inicio")
        fim    = st.sidebar.date_input("Data final",   value=fim_padrao   or date.today(), key=f"{prefixo}_fim")
    else:
        inicio, fim = inicio_padrao, fim_padrao

    usuario = aluno_logado()
    lista_alunos = sorted(df["aluno"].dropna().unique().tolist())
    if usuario["perfil"] == "Aluno":
        alunos_sel = [usuario["nome"]]
        st.sidebar.text_input("Aluno", value=usuario["nome"], disabled=True)
    else:
        alunos_sel = st.sidebar.multiselect(
            "Alunos", lista_alunos, default=[],
            placeholder="Todos os alunos",
            key=f"{prefixo}_aluno",
        )
    # visao: nome único (str) = individual | lista de nomes = comparativo
    if not alunos_sel:
        # "Todos" → comparativo com todos os alunos disponíveis no df
        todos_disponiveis = sorted(df["aluno"].dropna().unique().tolist())
        visao = todos_disponiveis if todos_disponiveis else "Todos"
    elif len(alunos_sel) == 1:
        visao = alunos_sel[0]   # individual
    else:
        visao = alunos_sel      # comparativo explícito

    disciplinas = sorted(df["disciplina"].dropna().unique().tolist())
    assuntos    = sorted(df["assunto"].dropna().unique().tolist())
    aulas       = sorted(df["aula"].dropna().unique().tolist())
    tipos       = sorted(df["tipo"].dropna().unique().tolist())

    status_escolhidos = st.sidebar.multiselect(
        "Status", STATUS_VALIDOS, default=STATUS_VALIDOS,
        format_func=lambda v: STATUS_LABELS.get(v, v), key=f"{prefixo}_status",
    )
    disciplina   = st.sidebar.multiselect("Disciplinas",     disciplinas, key=f"{prefixo}_disciplina")
    aula         = st.sidebar.multiselect("Aulas",           aulas,       key=f"{prefixo}_aula")
    assunto      = st.sidebar.multiselect("Assuntos",        assuntos,    key=f"{prefixo}_assunto")
    tipo         = st.sidebar.multiselect("Tipos de estudo", tipos,       key=f"{prefixo}_tipo")
    minimo_horas = st.sidebar.number_input(
        "Tempo mínimo estudado (h)", min_value=0.0, value=0.0,
        step=0.25, format="%.2f", key=f"{prefixo}_min_horas",
    )
    recentes = st.sidebar.toggle("Atividades recentes (15d)", value=False, key=f"{prefixo}_recentes")

    nome_filtro = st.sidebar.text_input("Nome para salvar filtro", key=f"{prefixo}_nome_filtro")
    if st.sidebar.button("Salvar filtro favorito", use_container_width=True, key=f"{prefixo}_salvar_filtro"):
        if nome_filtro:
            favoritos[nome_filtro] = {
                "periodo": periodo, "aluno": alunos_sel, "status": status_escolhidos,
                "disciplina": disciplina, "aula": aula, "assunto": assunto,
                "tipo": tipo, "min_horas": minimo_horas, "recentes": recentes,
            }
            _toast_sucesso("Filtro salvo como favorito.")

    # ── Filtros estruturais (sem período) ──
    escopo = df.copy()
    escopo["data_ref"] = pd.to_datetime(escopo["data_execucao"], errors="coerce")

    if isinstance(visao, list):
        escopo = escopo[escopo["aluno"].isin(visao)]
    elif visao != "Todos":
        escopo = escopo[escopo["aluno"] == visao]
    if status_escolhidos:
        escopo = escopo[escopo["status"].isin(status_escolhidos)]
    if disciplina:
        escopo = escopo[escopo["disciplina"].isin(disciplina)]
    if aula:
        escopo = escopo[escopo["aula"].isin(aula)]
    if assunto:
        escopo = escopo[escopo["assunto"].isin(assunto)]
    if tipo:
        escopo = escopo[escopo["tipo"].isin(tipo)]

    # ── Filtros temporais (com período) ──
    # Aplica período apenas sobre registros que têm data de execução.
    # Tarefas sem data (Não iniciadas) passam pelo filtro sem serem excluídas.
    periodo_df = escopo.copy()

    # Converte inicio/fim para Timestamp para comparação segura com datetime64
    ts_inicio = pd.Timestamp(inicio) if inicio else None
    ts_fim    = pd.Timestamp(fim)    if fim    else None

    if ts_inicio:
        sem_data = periodo_df["data_ref"].isna()
        com_data = periodo_df["data_ref"].notna() & (periodo_df["data_ref"] >= ts_inicio)
        periodo_df = periodo_df[sem_data | com_data]
    if ts_fim:
        sem_data = periodo_df["data_ref"].isna()
        com_data = periodo_df["data_ref"].notna() & (periodo_df["data_ref"] <= ts_fim + pd.Timedelta(days=1) - pd.Timedelta(seconds=1))
        periodo_df = periodo_df[sem_data | com_data]
    if minimo_horas > 0:
        periodo_df = periodo_df[
            periodo_df["data_ref"].isna() | (periodo_df["ch_efetiva"] >= minimo_horas)
        ]
    if recentes:
        limite    = pd.Timestamp.now() - pd.Timedelta(days=15)
        sem_data  = periodo_df["data_ref"].isna()
        com_data  = periodo_df["data_ref"].notna() & (periodo_df["data_ref"] >= limite)
        periodo_df = periodo_df[sem_data | com_data]

    return escopo, periodo_df, visao, inicio, fim


def base_metricas(df):
    return df[df["status"].isin(STATUS_ANALISE)].copy()


def dias_uteis_no_periodo(inicio: date | None, fim: date | None) -> int:
    """
    Conta quantos dias úteis (seg–sex) existem entre inicio e fim (inclusive).
    Retorna 0 se periodo indefinido ou inválido.
    """
    if not inicio or not fim or fim < inicio:
        return 0
    total = 0
    d = inicio
    while d <= fim:
        if d.weekday() < 5:   # 0=seg … 4=sex
            total += 1
        d += timedelta(days=1)
    return total


def _pares_aluno_tarefa(df: pd.DataFrame) -> set:
    """Conjunto de pares (aluno_id, tarefa_id) presentes num DataFrame."""
    if df is None or df.empty or not {"aluno_id", "tarefa_id"} <= set(df.columns):
        return set()
    d = df.dropna(subset=["aluno_id", "tarefa_id"])
    if d.empty:
        return set()
    return set(zip(d["aluno_id"].astype(int), d["tarefa_id"].astype(int)))


def _sessoes_do_escopo(df_escopo: pd.DataFrame, inicio=None, fim=None) -> pd.DataFrame:
    """
    Sessões de estudo restritas aos pares (aluno, tarefa) do escopo, opcionalmente
    recortadas por período.

    Por que não usar `execucoes`: lá `data_execucao` guarda a data da ÚLTIMA sessão
    da tarefa (ver `_recalcular_execucao_por_sessoes`). Somar horas por aquela coluna
    joga todo o esforço de uma tarefa no dia em que ela terminou — uma tarefa estudada
    ao longo de cinco semanas aparece inteira na última. Cada sessão tem a própria
    data, então é daqui que saem as métricas por período.
    """
    colunas = ["aluno_id", "tarefa_id", "data", "ch_efetiva",
               "qtd_questoes_feitas", "qtd_acertos"]
    try:
        sess = carregar_sessoes_dashboard()
    except Exception:
        return pd.DataFrame(columns=colunas)
    if sess is None or sess.empty:
        return pd.DataFrame(columns=colunas)

    sess = sess.copy()
    sess["data"] = pd.to_datetime(sess["data_execucao"], errors="coerce").dt.date
    sess = sess.dropna(subset=["data", "aluno_id", "tarefa_id"])

    pares = _pares_aluno_tarefa(df_escopo)
    if pares:
        chaves = zip(sess["aluno_id"].astype(int), sess["tarefa_id"].astype(int))
        sess = sess[[c in pares for c in chaves]]

    if inicio:
        sess = sess[sess["data"] >= inicio]
    if fim:
        sess = sess[sess["data"] <= fim]
    sess["data_ref"] = pd.to_datetime(sess["data"], errors="coerce")
    return sess


# Colunas que todo agregado temporal do dashboard consome
_COLS_TEMPORAL = [
    "aluno_id", "aluno", "tarefa_id", "disciplina", "disciplina_id",
    "aula", "assunto", "tipo", "status", "data_ref",
    "ch_efetiva", "qtd_questoes_feitas", "qtd_acertos",
]


def _frame_temporal(sess_per: pd.DataFrame, legado_per: pd.DataFrame) -> pd.DataFrame:
    """
    Base para tudo que agrega HORAS ao longo do tempo: uma linha por sessão, com a
    data em que aquele estudo aconteceu.

    Não serve para contar tarefas nem status — uma tarefa estudada em cinco dias
    vira cinco linhas aqui. Contagem de tarefas continua saindo de `analisavel`,
    que tem uma linha por par aluno×tarefa.
    """
    partes = []
    for origem in (sess_per, legado_per):
        if origem is None or origem.empty:
            continue
        d = origem.copy()
        for c in _COLS_TEMPORAL:
            if c not in d.columns:
                d[c] = pd.NA
        partes.append(d[_COLS_TEMPORAL])
    if not partes:
        return pd.DataFrame(columns=_COLS_TEMPORAL)
    fr = pd.concat(partes, ignore_index=True)
    fr["data_ref"] = pd.to_datetime(fr["data_ref"], errors="coerce")
    for c in ("ch_efetiva", "qtd_questoes_feitas", "qtd_acertos"):
        fr[c] = pd.to_numeric(fr[c], errors="coerce").fillna(0)
    return fr


def _sem_sessao(df: pd.DataFrame, pares_com_sessao: set) -> pd.DataFrame:
    """
    Linhas cujo par (aluno, tarefa) não tem nenhuma sessão registrada — casos de
    importação antiga, em que as horas vivem só em `execucoes`. Continuam contando
    pela data de execução para não sumirem dos totais.
    """
    if df is None or df.empty:
        return df
    if not {"aluno_id", "tarefa_id"} <= set(df.columns):
        return df
    def fora(row):
        a, t = row["aluno_id"], row["tarefa_id"]
        if pd.isna(a) or pd.isna(t):
            return True
        return (int(a), int(t)) not in pares_com_sessao
    return df[df.apply(fora, axis=1)] if not df.empty else df


def calcular_kpis_avancados(
    df_escopo: pd.DataFrame,
    df_periodo: pd.DataFrame,
    inicio_periodo=None,
    fim_periodo=None,
) -> dict:
    """
    Calcula todos os KPIs com responsabilidades separadas por DataFrame.

    df_escopo  → filtrado por aluno/disciplina/status/aula/tipo SEM período.
                 Fonte para: total de tarefas, progresso, restantes, em andamento.
                 Tarefas Não iniciadas (sem data) nunca desaparecem por filtro de período.

    df_periodo → df_escopo com período aplicado APENAS em registros com data.
                 Tarefas sem data (Não iniciadas) permanecem.
                 Fonte para: horas, questões, acertos, desempenho, concluídas no período.

    Esta separação garante coerência entre todos os cards, gráficos e KPIs.
    """
    hoje       = date.today()
    semana_ini = hoje - timedelta(days=hoje.weekday())
    semana_pas = semana_ini - timedelta(days=7)

    for _df in [df_escopo, df_periodo]:
        if "data_ref" not in _df.columns:
            _df["data_ref"] = pd.to_datetime(
                _df.get("data_execucao", pd.Series(dtype=str)), errors="coerce"
            )

    # ── Estruturais — escopo total sem período ──
    esc_conc = df_escopo[df_escopo["status"] == STATUS_CONCLUIDA]
    esc_and  = df_escopo[df_escopo["status"] == STATUS_EM_ANDAMENTO]
    esc_nao  = df_escopo[df_escopo["status"] == STATUS_NAO_INICIADA]
    esc_anal = df_escopo[df_escopo["status"].isin(STATUS_ANALISE)].copy()

    total_tarefas    = len(df_escopo)
    qtd_concluidas   = len(esc_conc)
    qtd_andamento    = len(esc_and)
    qtd_nao_iniciada = len(esc_nao)
    pct_conclusao    = qtd_concluidas / total_tarefas * 100 if total_tarefas else 0
    tarefas_rest     = qtd_andamento + qtd_nao_iniciada

    # ── Temporais — período filtrado ──
    # Horas, questões e acertos vêm de sessoes_estudo, onde cada linha tem a data
    # em que o estudo aconteceu de verdade. Ver _sessoes_do_escopo.
    per_anal = df_periodo[df_periodo["status"].isin(STATUS_ANALISE)].copy()
    per_conc = per_anal[per_anal["status"] == STATUS_CONCLUIDA]

    sess_escopo = _sessoes_do_escopo(df_escopo)
    sess_per    = _sessoes_do_escopo(df_escopo, inicio_periodo, fim_periodo)
    pares_sess  = _pares_aluno_tarefa(sess_escopo)

    legado_per = _sem_sessao(per_anal, pares_sess)
    legado_esc = _sem_sessao(esc_anal, pares_sess)

    horas_total = float(sess_per["ch_efetiva"].sum()) + float(legado_per["ch_efetiva"].sum())
    questoes    = int(sess_per["qtd_questoes_feitas"].sum()) + int(legado_per["qtd_questoes_feitas"].sum())
    acertos     = int(sess_per["qtd_acertos"].sum()) + int(legado_per["qtd_acertos"].sum())
    desempenho  = acertos / questoes * 100 if questoes else 0

    # Horas de todo o escopo (sem recorte de período) — base estável para estimativas
    horas_escopo = float(sess_escopo["ch_efetiva"].sum()) + float(legado_esc["ch_efetiva"].sum())

    # Dias ativos no PERÍODO: datas das sessões + datas de execução do legado
    datas_periodo = set(sess_per["data"].tolist())
    datas_periodo |= set(legado_per["data_ref"].dropna().dt.date.tolist())
    dias_unicos     = sorted(datas_periodo)
    qtd_dias_ativos = len(dias_unicos)

    _ini = inicio_periodo or (min(dias_unicos) if dias_unicos else None)
    _fim = fim_periodo    or (max(dias_unicos) if dias_unicos else None)
    dias_uteis = dias_uteis_no_periodo(_ini, _fim)
    media_diaria = horas_total / dias_uteis if dias_uteis > 0 else 0.0
    media_diaria_label = (
        "Não aplicável (sem dias úteis no período)"
        if dias_uteis == 0 else
        f"{dias_uteis} dia(s) útil(eis) no período"
    )

    # Semana atual e anterior — sempre sobre escopo total (para insights consistentes).
    # Horas pelas sessões; conclusões pela data de execução, que é a data em que a
    # tarefa efetivamente terminou.
    def _horas_entre(ini, fim_ex):
        """Horas das sessões do escopo em [ini, fim_ex)."""
        if sess_escopo.empty:
            base = 0.0
        else:
            m = (sess_escopo["data"] >= ini) & (sess_escopo["data"] < fim_ex)
            base = float(sess_escopo.loc[m, "ch_efetiva"].sum())
        if legado_esc is not None and not legado_esc.empty:
            leg = legado_esc.dropna(subset=["data_ref"])
            if not leg.empty:
                d = leg["data_ref"].dt.date
                base += float(leg.loc[(d >= ini) & (d < fim_ex), "ch_efetiva"].sum())
        return base

    prox_semana   = semana_ini + timedelta(days=7)
    horas_semana  = _horas_entre(semana_ini, prox_semana)
    horas_sem_pas = _horas_entre(semana_pas, semana_ini)

    # Limite superior evita que data de execução no futuro entre como "esta semana"
    exec_semana   = esc_anal[
        (esc_anal["data_ref"].dt.date >= semana_ini) &
        (esc_anal["data_ref"].dt.date < prox_semana)
    ]
    conc_semana   = int((exec_semana["status"] == STATUS_CONCLUIDA).sum())
    exec_sem_pas  = esc_anal[
        (esc_anal["data_ref"].dt.date >= semana_pas) &
        (esc_anal["data_ref"].dt.date < semana_ini)
    ]
    conc_sem_pas  = int((exec_sem_pas["status"] == STATUS_CONCLUIDA).sum())
    conc_periodo  = len(per_conc)

    # Janelas móveis de 7 dias — comparação justa em qualquer dia da semana.
    # A semana-calendário compara 1 dia de segunda contra 7 dias da semana anterior;
    # os cards mantêm o recorte calendário, os insights usam estas janelas.
    horas_7d     = _horas_entre(hoje - timedelta(days=6),  hoje + timedelta(days=1))
    horas_7d_ant = _horas_entre(hoje - timedelta(days=13), hoje - timedelta(days=6))
    delta_7d     = horas_7d - horas_7d_ant

    # Sequência, dias ativos e dias sem estudar
    # Fonte primária: sessoes_estudo (tarefas EM_ANDAMENTO ou CONCLUIDAS)
    # Fallback: data_execucao de esc_anal
    # dias_sem_estudar nunca negativo (datas futuras → max 0)
    datas_execucao = set(sess_escopo["data"].tolist())
    datas_execucao |= set(esc_anal["data_ref"].dropna().dt.date.tolist())

    todas_datas = sorted(datas_execucao)
    conjunto_datas = set(todas_datas)
    # qtd_dias_ativos já calculado acima com base no período

    # A sequência conta a partir do ÚLTIMO dia com registro, não de hoje.
    # Contar de hoje zerava o valor toda manhã antes da primeira sessão do dia:
    # numa segunda-feira, cinco dias seguidos de estudo apareciam como "0 dias".
    sequencia = 0
    if todas_datas:
        d = max(todas_datas)
        while d in conjunto_datas:
            sequencia += 1
            d -= timedelta(days=1)

    sequencia_ativa = bool(todas_datas) and (hoje - max(todas_datas)).days <= 1
    dias_sem_estudar = max(0, (hoje - max(todas_datas)).days) if todas_datas else 0

    # Dias ÚTEIS sem estudar: sexta → segunda são 3 dias corridos, mas só 1 útil.
    # O alerta de pausa usa este valor para não disparar por causa do fim de semana.
    if todas_datas and dias_sem_estudar > 0:
        dias_uteis_sem_estudar = dias_uteis_no_periodo(
            max(todas_datas) + timedelta(days=1), hoje
        )
    else:
        dias_uteis_sem_estudar = 0

    # Numerador e denominador precisam vir do MESMO recorte. Usar as concluídas do
    # escopo (sem período) contra as horas do período dava 7,99 tarefas/hora num
    # filtro de 7 dias, contra 0,78 reais.
    produtividade = conc_periodo / horas_total if horas_total else 0

    # Previsão — ritmo baseado no período, restantes baseados no escopo
    previsao      = None
    ritmo_semanal = 0.0
    previsao_base = ""
    conc_com_data = per_conc.dropna(subset=["data_ref"]).copy()
    if not conc_com_data.empty:
        conc_com_data["semana"] = conc_com_data["data_ref"].dt.to_period("W")
        por_semana = conc_com_data.groupby("semana").size()
        # O groupby só devolve semanas que TIVERAM conclusão. Sem reindexar sobre o
        # calendário, as semanas paradas somem da média e o ritmo fica otimista —
        # 50% acima do real numa aluna, 88% na outra.
        fim_serie = fim_periodo or hoje
        inicio_serie = conc_com_data["data_ref"].min().date()
        calendario = pd.period_range(
            start=inicio_serie, end=max(fim_serie, inicio_serie), freq="W"
        )
        por_semana = por_semana.reindex(calendario, fill_value=0)
        ultimas = por_semana.tail(8)
        if len(ultimas) >= 1:
            ritmo_semanal = float(ultimas.mean())

    if tarefas_rest == 0:
        previsao_base = "Todas as tarefas do escopo já estão concluídas."
    elif ritmo_semanal > 0:
        semanas_rest  = tarefas_rest / ritmo_semanal
        previsao      = hoje + timedelta(weeks=semanas_rest)
        previsao_base = (
            f"Ritmo médio (período): {ritmo_semanal:.1f} tarefas/semana "
            f"· {tarefas_rest} restante(s) "
            f"· Estimativa: {previsao.strftime('%d/%m/%Y')}."
        )
    else:
        previsao_base = "Dados insuficientes: sem tarefas concluídas com data no período selecionado."

    # Custo médio por tarefa: horas e conclusões do mesmo recorte. Quando o período
    # não tem conclusões, cai para o histórico inteiro do escopo — caso contrário a
    # estimativa de esforço restante zeraria e leria como "não falta nada".
    if conc_periodo:
        media_h_tarefa = horas_total / conc_periodo
    elif qtd_concluidas:
        media_h_tarefa = horas_escopo / qtd_concluidas
    else:
        media_h_tarefa = 0
    horas_restantes = tarefas_rest * media_h_tarefa

    media_por_dia_ativo = horas_total / qtd_dias_ativos if qtd_dias_ativos else 0.0

    return {
        "analisavel":           per_anal,
        "temporal":             _frame_temporal(sess_per, legado_per),
        "temporal_escopo":      _frame_temporal(sess_escopo, legado_esc),
        "concluidas_df":        per_conc,
        "andamento_df":         esc_and,
        "escopo_analisavel":    esc_anal,
        "total_tarefas":        total_tarefas,
        "qtd_concluidas":       qtd_concluidas,
        "qtd_andamento":        qtd_andamento,
        "qtd_nao_iniciada":     qtd_nao_iniciada,
        "pct_conclusao":        pct_conclusao,
        "tarefas_restantes":    tarefas_rest,
        "horas_total":          horas_total,
        "horas_escopo":         horas_escopo,
        "conc_periodo":         conc_periodo,
        "horas_semana":         horas_semana,
        "horas_sem_pas":        horas_sem_pas,
        "delta_horas_semana":   horas_semana - horas_sem_pas,
        "horas_7d":             horas_7d,
        "horas_7d_ant":         horas_7d_ant,
        "delta_7d":             delta_7d,
        "conc_semana":          conc_semana,
        "conc_sem_pas":         conc_sem_pas,
        "delta_conc_semana":    conc_semana - conc_sem_pas,
        "questoes":             questoes,
        "acertos":              acertos,
        "desempenho":           desempenho,
        "qtd_dias_ativos":      qtd_dias_ativos,
        "dias_uteis":           dias_uteis,
        "media_diaria":         media_diaria,
        "media_diaria_label":   media_diaria_label,
        "media_por_dia_ativo":  media_por_dia_ativo,
        "sequencia":            sequencia,
        "sequencia_ativa":      sequencia_ativa,
        "dias_sem_estudar":     dias_sem_estudar,
        "dias_uteis_sem_estudar": dias_uteis_sem_estudar,
        "produtividade":        produtividade,
        "ritmo_semanal":        ritmo_semanal,
        "previsao_conclusao":   previsao,
        "previsao_base":        previsao_base,
        "horas_restantes":      horas_restantes,
        "dias_unicos":          todas_datas,   # todas as datas de estudo do escopo
    }

def calcular_metricas(df_escopo, df_periodo=None, inicio_periodo=None, fim_periodo=None):
    """Compatível com código legado. Se df_periodo não for fornecido, usa df_escopo."""
    if df_periodo is None:
        df_periodo = df_escopo
    k = calcular_kpis_avancados(df_escopo, df_periodo, inicio_periodo, fim_periodo)
    return {
        "analisavel": k["analisavel"], "concluidas_df": k["concluidas_df"],
        "andamento_df": k["andamento_df"], "total": k["total_tarefas"],
        "concluidas": k["qtd_concluidas"], "andamento": k["qtd_andamento"],
        "nao_iniciadas": k["qtd_nao_iniciada"], "progresso": k["pct_conclusao"],
        "horas": k["horas_total"], "questoes": k["questoes"], "acertos": k["acertos"],
        "desempenho": k["desempenho"], "media_dia": k["media_diaria"],
        "produtividade": k["produtividade"],
    }


def gerar_insights(df: pd.DataFrame, kpis: dict, nome_aluno: str = "") -> list:
    """Gera insights analíticos contextuais baseados nos dados do aluno."""
    insights = []
    ana = kpis["analisavel"]
    nome = nome_aluno or "O aluno"
    if ana.empty:
        return [{"tipo":"info","icone":"📭","titulo":"Sem dados suficientes","texto":"Registre atividades para obter análise personalizada."}]

    # Tendência de horas — janelas móveis de 7 dias.
    # A semana-calendário comparava 1 dia de segunda contra 7 da semana anterior,
    # então "estudou 100% menos esta semana" disparava toda segunda de manhã.
    h7  = kpis.get("horas_7d", 0.0)
    h7a = kpis.get("horas_7d_ant", 0.0)
    if h7a > 0:
        pct_d = (h7 - h7a) / h7a * 100
        if pct_d < -20:
            insights.append({"dim":"tendencia","tipo":"warning","icone":"📉","titulo":"Queda de ritmo",
                "texto":f"{nome} estudou {abs(pct_d):.0f}% menos nos últimos 7 dias ({horas_para_hm(h7)}) que nos 7 anteriores ({horas_para_hm(h7a)})."})
        elif pct_d > 20:
            insights.append({"dim":"tendencia","tipo":"success","icone":"📈","titulo":"Ritmo em alta",
                "texto":f"{nome} aumentou {pct_d:.0f}% as horas nos últimos 7 dias ({horas_para_hm(h7)} vs {horas_para_hm(h7a)}). Excelente ritmo!"})

    # Sequência
    seq  = kpis["sequencia"]
    viva = kpis.get("sequencia_ativa", True)
    dsu  = kpis.get("dias_uteis_sem_estudar", kpis["dias_sem_estudar"])
    if seq >= 7 and viva:
        insights.append({"dim":"sequencia","tipo":"success","icone":"🔥","titulo":f"Sequência de {seq} dias!",
            "texto":f"{nome} está estudando há {seq} dias consecutivos. Consistência é fundamental para aprovação."})
    elif dsu >= 3:
        # dias ÚTEIS: sexta → segunda não conta como pausa
        insights.append({"dim":"sequencia","tipo":"warning","icone":"⏰","titulo":"Dias sem estudar",
            "texto":f"{nome} não registra atividade há {dsu} dia(s) útil(eis) "
                    f"({kpis['dias_sem_estudar']} dias corridos). Retome a rotina para não perder o ritmo."})

    # Disciplinas
    if "disciplina" in ana.columns and len(ana["disciplina"].unique()) > 1:
        disc_agg = ana.groupby("disciplina", as_index=False).agg(
            questoes=("qtd_questoes_feitas","sum"), acertos=("qtd_acertos","sum"),
            horas=("ch_efetiva","sum"), tarefas=("tarefa_id","count"),
            concluidas=("status", lambda s:(s==STATUS_CONCLUIDA).sum()))
        disc_agg["desempenho"] = disc_agg.apply(lambda r: r["acertos"]/r["questoes"]*100 if r["questoes"] else 0, axis=1)
        disc_agg["pct_conc"]   = disc_agg.apply(lambda r: r["concluidas"]/r["tarefas"]*100 if r["tarefas"] else 0, axis=1)

        # Denominador do progresso: TODAS as tarefas da disciplina, não só as
        # iniciadas. Sobre as iniciadas, 10 de 10 em Informática dava "100%
        # concluído" numa disciplina de 55 tarefas — 18% de verdade.
        if "disciplina" in df.columns:
            total_disc = df.groupby("disciplina")["tarefa_id"].count()
            disc_agg["tarefas_disciplina"] = (
                disc_agg["disciplina"].map(total_disc).fillna(disc_agg["tarefas"])
            )
        else:
            disc_agg["tarefas_disciplina"] = disc_agg["tarefas"]
        disc_agg["pct_conc"] = disc_agg.apply(
            lambda r: r["concluidas"] / r["tarefas_disciplina"] * 100
            if r["tarefas_disciplina"] else 0, axis=1)

        criticas = disc_agg[(disc_agg["questoes"]>0) & (disc_agg["desempenho"]<60)].sort_values("desempenho")
        for _, r in criticas.head(2).iterrows():
            insights.append({"dim":f"desempenho:{r['disciplina']}","tipo":"danger","icone":"🚨","titulo":f"Atenção: {r['disciplina']}",
                "texto":f"Desempenho de {r['desempenho']:.1f}% em {r['disciplina']} abaixo de 60%. Dedique sessões de revisão e questões comentadas."})

        total_h = disc_agg["horas"].sum()
        if total_h > 0:
            # horas > 0: disciplina sem nenhuma hora no período não é "negligenciada",
            # só não foi tocada — num filtro de 7 dias isso vale para quase todas.
            neg = disc_agg[(disc_agg["horas"] > 0) &
                           (disc_agg["horas"]/total_h < 0.05) &
                           (disc_agg["tarefas"] > 0)]
            for _, r in neg.head(2).iterrows():
                insights.append({"dim":f"tempo:{r['disciplina']}","tipo":"warning","icone":"📌","titulo":f"Disciplina negligenciada: {r['disciplina']}",
                    "texto":f"Apenas {horas_para_hm(r['horas'])} em {r['disciplina']} (<5% do tempo). Reequilibre a distribuição."})

        avancadas = disc_agg[disc_agg["pct_conc"]>=80].sort_values("pct_conc", ascending=False)
        if not avancadas.empty:
            nomes = ", ".join(
                f"{r['disciplina']} ({r['pct_conc']:.0f}%)"
                for _, r in avancadas.head(3).iterrows()
            )
            insights.append({"dim":"progresso","tipo":"success","icone":"🏆","titulo":"Disciplinas avançadas",
                "texto":f"Ótimo progresso em: {nomes}. Continue com revisões."})

        # Onde o esforço já começou mas o avanço na disciplina ainda é pequeno
        atrasadas = disc_agg[(disc_agg["tarefas"] >= 3) & (disc_agg["pct_conc"] < 25)] \
            .sort_values("pct_conc")
        for _, r in atrasadas.head(2).iterrows():
            insights.append({"dim":f"progresso:{r['disciplina']}","tipo":"info","icone":"📊",
                "titulo":f"Avanço inicial: {r['disciplina']}",
                "texto":f"{int(r['concluidas'])} de {int(r['tarefas_disciplina'])} tarefas concluídas "
                        f"({r['pct_conc']:.0f}%) em {r['disciplina']}, com {int(r['tarefas'])} já iniciadas."})

    # Risco de atraso
    prev = kpis.get("previsao_conclusao")
    if prev and kpis["tarefas_restantes"] > 0:
        dias_p = (prev - date.today()).days
        ritmo = kpis.get("ritmo_semanal", 0)
        base_ritmo = f" Ritmo considerado: {ritmo:.1f} tarefa(s)/semana." if ritmo else ""
        if dias_p > 365:
            insights.append({"dim":"previsao","tipo":"danger","icone":"⚠️","titulo":"Risco de atraso alto",
                "texto":f"No ritmo atual, conclusão estimada em {prev.strftime('%d/%m/%Y')} ({dias_p} dias).{base_ritmo} Considere aumentar a carga semanal."})
        elif dias_p > 180:
            insights.append({"dim":"previsao","tipo":"warning","icone":"📅","titulo":"Previsão de conclusão",
                "texto":f"Estimativa: {prev.strftime('%d/%m/%Y')} ({dias_p} dias).{base_ritmo} Mantenha o ritmo."})

    # Produtividade — concluídas e horas do MESMO recorte de período
    prod = kpis["produtividade"]
    if prod >= 1.5:
        insights.append({"dim":"produtividade","tipo":"success","icone":"⚡","titulo":"Alta produtividade",
            "texto":f"{prod:.2f} tarefas concluídas por hora estudada no período — excelente eficiência."})
    elif 0 < prod < 0.3:
        insights.append({"dim":"produtividade","tipo":"info","icone":"🕐","titulo":"Sessões longas, pouco avanço",
            "texto":f"{prod:.2f} tarefas concluídas por hora estudada no período. Experimente sessões mais curtas com foco (técnica Pomodoro)."})

    # Consistência — horas por dia EM QUE HOUVE ESTUDO.
    # O card "Média diária" divide por dias úteis do período e é outra medida;
    # aqui o texto fala em "dia ativo", então o número tem que ser o de dia ativo.
    mda = kpis.get("media_por_dia_ativo", 0.0)
    if mda >= 4:
        insights.append({"dim":"consistencia","tipo":"success","icone":"📚","titulo":"Sessões longas e produtivas",
            "texto":f"Média de {horas_para_hm(mda)} nos dias em que estuda — ritmo sólido."})
    elif 0 < mda < 1:
        insights.append({"dim":"consistencia","tipo":"warning","icone":"⏱️","titulo":"Sessões curtas",
            "texto":f"Média de {horas_para_hm(mda)} nos dias em que estuda. Para concursos competitivos, recomenda-se ≥4h por dia de estudo."})

    # Frequência — quantos dias úteis do período tiveram estudo
    du, dat = kpis.get("dias_uteis", 0), kpis.get("qtd_dias_ativos", 0)
    if du >= 10 and dat > 0:
        freq = dat / du * 100
        if freq < 40:
            insights.append({"dim":"frequencia","tipo":"warning","icone":"📆","titulo":"Frequência baixa",
                "texto":f"{dat} dia(s) com estudo em {du} dias úteis do período ({freq:.0f}%). "
                        f"A carga por sessão está boa; falta regularidade."})

    if not insights:
        insights.append({"dim":"geral","tipo":"info","icone":"✅","titulo":"Estudos em dia",
            "texto":f"{nome} não apresenta alertas críticos. Continue monitorando o progresso."})

    return _resolver_conflitos(insights)


_PESO_INSIGHT = {"danger": 3, "warning": 2, "info": 1, "success": 0}


def _resolver_conflitos(insights: list) -> list:
    """
    Impede que a mesma dimensão apareça ao mesmo tempo como sucesso e como alerta.
    Mantém o item mais severo de cada dimensão e devolve a lista ordenada por
    severidade, para o que exige ação ficar no topo.
    """
    por_dim = {}
    for i, ins in enumerate(insights):
        dim = ins.get("dim") or f"_{i}"
        atual = por_dim.get(dim)
        if atual is None or _PESO_INSIGHT.get(ins["tipo"], 0) > _PESO_INSIGHT.get(atual["tipo"], 0):
            por_dim[dim] = ins
    ordenados = sorted(
        por_dim.values(),
        key=lambda x: -_PESO_INSIGHT.get(x["tipo"], 0),
    )
    return ordenados


def fig_layout(fig, height=320):
    """
    Estilo visual padrão de qualquer figura Plotly.

    A legenda fica ACIMA da área de plotagem. Antes ficava em y=-0.35 com margem
    inferior de 10px: como a margem não reservava espaço, a legenda era desenhada
    por cima das barras e dos rótulos do eixo x. Com um traço só ela é escondida,
    porque o título já diz o que está sendo mostrado.
    """
    n_series = len([tr for tr in fig.data if getattr(tr, "showlegend", None) is not False])
    com_legenda = n_series > 1

    fig.update_layout(
        showlegend=com_legenda,
        margin=dict(l=12, r=18, t=76 if com_legenda else 52, b=12),
        height=height,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#0f172a", family="Inter, sans-serif", size=12),
        title=dict(font=dict(size=14.5, color="#0f172a"),
                   x=0, xanchor="left", y=0.98, yanchor="top"),
        legend=dict(orientation="h", yanchor="bottom", y=1.0,
                    xanchor="right", x=1, title_text="",
                    font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
        hoverlabel=dict(bgcolor="white", font_size=12, bordercolor="#e2e8f0"),
        bargap=0.28,
    )
    # automargin evita que rótulo longo de eixo seja cortado
    fig.update_xaxes(showgrid=False, linecolor="#e2e8f0", tickfont=dict(size=10),
                     automargin=True)
    fig.update_yaxes(showgrid=True, gridcolor="#f1f5f9", linecolor="rgba(0,0,0,0)",
                     tickfont=dict(size=10), automargin=True, zeroline=False)
    return fig


def grafico_vazio(msg="Sem dados suficientes para este gráfico."):
    """Retorna figura vazia com mensagem amigável."""
    fig = go.Figure()
    fig.add_annotation(text=msg, x=0.5, y=0.5, xref="paper", yref="paper",
        showarrow=False, font=dict(size=13, color="#94a3b8"))
    fig.update_layout(height=260, paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)", xaxis=dict(visible=False), yaxis=dict(visible=False))
    return fig


def preparar_tabela(df):
    tabela = df.copy()
    if "data_execucao" in tabela.columns:
        tabela["data_execucao"] = tabela["data_execucao"].apply(formatar_data_br)
    if "status" in tabela.columns:
        tabela["status_label"] = tabela["status"].map(STATUS_LABELS)
    if "ch_efetiva" in tabela.columns:
        tabela["tempo"] = tabela["ch_efetiva"].apply(horas_para_hm)
    return tabela


# ─────────────────────────────────────────────
# TELA: LOGIN
# ─────────────────────────────────────────────

def tela_login():
    render_html(f"""
        <div class="hero">
          <h1>{APP_NAME}</h1>
          <p>Acompanhamento de estudos, produtividade, tarefas educacionais e desempenho por aluno.</p>
        </div>
    """)
    with st.expander("Usuário inicial"):
        st.write(f"Gestor: **{ADMIN_EMAIL}** / senha **{ADMIN_PASSWORD}**")
    with st.form("login"):
        email = st.text_input("E-mail").strip().lower()
        senha = st.text_input("Senha", type="password")
        entrar = st.form_submit_button("Entrar", use_container_width=True)
    if entrar:
        usuario = consultar("SELECT * FROM alunos WHERE email = ? AND ativo = 1", (email,))
        if usuario.empty or not verificar_senha(senha, usuario.iloc[0]["senha"]):
            st.error("Usuário ou senha inválidos.")
            return
        st.session_state["usuario"] = usuario.iloc[0].to_dict()
        st.rerun()


def tela_troca_obrigatoria():
    usuario = aluno_logado()
    render_html(f"""
        <div class="hero">
          <h1>{APP_NAME}</h1>
          <p>Por segurança, altere sua senha inicial antes de acessar o sistema.</p>
        </div>
    """)
    with st.form("troca_obrigatoria"):
        nova = st.text_input("Nova senha", type="password")
        confirmar = st.text_input("Confirmar nova senha", type="password")
        salvar = st.form_submit_button("Alterar senha e entrar", use_container_width=True)
    if salvar:
        if nova != confirmar:
            st.error("A confirmação não confere.")
        elif not senha_valida(nova):
            st.error("Use uma senha com pelo menos 8 caracteres, contendo letras e números.")
        else:
            atualizar_senha_usuario(usuario["id"], nova, 0)
            usuario_atualizado = consultar("SELECT * FROM alunos WHERE id = ?", (int(usuario["id"]),))
            st.session_state["usuario"] = usuario_atualizado.iloc[0].to_dict()
            _toast_sucesso("Senha alterada com sucesso. Faça login novamente.")
            st.rerun()


# ─────────────────────────────────────────────
# TELA: DASHBOARD — funções auxiliares
# ─────────────────────────────────────────────

def _render_kpis_produtividade(kpis: dict):
    """Bloco: Total de horas, média diária por dias úteis, sequência, dias sem estudar, produtividade."""
    horas  = kpis["horas_total"]
    seq    = kpis["sequencia"]
    dsem   = kpis["dias_sem_estudar"]
    du     = kpis["dias_uteis"]
    md_lbl = kpis["media_diaria_label"]
    cols   = st.columns(5)
    cards  = [
        kpi_card(
            "Total de horas estudadas", horas_para_hm(horas),
            f"{kpis['qtd_dias_ativos']} dia(s) com registro no período",
            tooltip=(
                "Soma de todas as horas de estudo (ch_efetiva) das atividades "
                "Em andamento e Concluídas dentro do período e filtros selecionados. "
                "Filtros que impactam: aluno, disciplina, tipo, período de datas. "
                "Quanto maior, mais tempo dedicado — mas compare com desempenho para avaliar eficiência."
            ),
        ),
        kpi_card(
            "Média diária (dias úteis)", horas_para_hm(kpis["media_diaria"]),
            md_lbl,
            tooltip=(
                "Média de horas de estudo por dia útil (segunda a sexta-feira) do período selecionado. "
                f"Fórmula: total de horas ÷ número de dias úteis do período. "
                f"Dias úteis no período atual: {du}. "
                "Sábados e domingos NÃO entram no divisor. "
                "Se o período não tiver dias úteis, exibe 'Não aplicável'. "
                "Referência: ≥ 4h/dia útil = ritmo consistente para concursos."
            ),
        ),
        kpi_card(
            "Sequência atual", f"{seq} dia{'s' if seq!=1 else ''}",
            "dias consecutivos com registro",
            tooltip=(
                "Quantos dias seguidos (contando de hoje para trás) houve pelo menos 1 execução. "
                "Reinicia quando há um dia sem registro. "
                "⚠️ Não é afetado pelo filtro de período — considera todo o histórico do aluno. "
                "Sequências longas indicam hábito de estudo consolidado."
            ),
        ),
        kpi_card(
            "Dias sem estudar", f"{dsem}",
            "desde o último registro",
            delta=f"{dsem}d de pausa" if dsem > 0 else "",
            delta_pos=False,
            tooltip=(
                "Quantidade de dias desde o último registro de execução. "
                "0 = estudou hoje. Acima de 3 dias = alerta de perda de ritmo. "
                "⚠️ Não é afetado pelo filtro de período — sempre considera a data mais recente no histórico total."
            ),
        ),
        kpi_card(
            "Produtividade", f"{kpis['produtividade']:.2f}",
            "tarefas concluídas / hora estudada",
            tooltip=(
                "Eficiência: quantas tarefas são concluídas por hora de estudo. "
                "Fórmula: tarefas concluídas ÷ total de horas. "
                "> 1,0 = boa taxa de conclusão. "
                "< 0,3 pode indicar sessões longas sem finalizar tarefas. "
                "Filtros ativos: aluno, disciplina, período."
            ),
        ),
    ]
    for col, card in zip(cols, cards):
        with col:
            render_html(card)


def _render_kpis_avanco(kpis: dict):
    """Bloco: Progresso geral, Em andamento, Tarefas restantes, Horas restantes, Previsão."""
    pct      = kpis["pct_conclusao"]
    prev     = kpis.get("previsao_conclusao")
    prev_str = prev.strftime("%d/%m/%Y") if prev else "Dados insuficientes"
    prev_sub = kpis.get("previsao_base", "")
    total    = kpis["total_tarefas"]
    conc     = kpis["qtd_concluidas"]
    and_     = kpis["qtd_andamento"]
    nao      = kpis["qtd_nao_iniciada"]
    rest     = kpis["tarefas_restantes"]

    cols = st.columns(5)
    cards = [
        kpi_card(
            "Progresso geral", f"{pct:.1f}%",
            f"{conc} concluídas de {total} tarefas totais",
            tooltip=(
                "Percentual de conclusão em relação a TODAS as tarefas do aluno no período filtrado "
                "(concluídas, em andamento e não iniciadas). "
                f"Fórmula: ({conc} concluídas ÷ {total} total) × 100. "
                "Filtros ativos: aluno, disciplina, período. "
                "100% = plano concluído."
            ),
        ),
        kpi_card(
            "Em andamento", f"{and_}",
            f"de {total} tarefas totais",
            tooltip=(
                "Quantidade de tarefas com status EM_ANDAMENTO dentro dos filtros selecionados. "
                "Não confunde com Não iniciadas (essas são contadas separadamente). "
                f"Não iniciadas: {nao}. "
                "Muitas tarefas em andamento simultâneas podem indicar falta de foco."
            ),
        ),
        kpi_card(
            "Tarefas restantes", f"{rest}",
            f"{and_} em andamento + {nao} não iniciadas",
            tooltip=(
                "Total de tarefas ainda não concluídas dentro dos filtros ativos. "
                f"Fórmula: em andamento ({and_}) + não iniciadas ({nao}). "
                "Base usada para calcular previsão de conclusão e horas estimadas restantes."
            ),
        ),
        kpi_card(
            "Horas restantes (est.)", horas_para_hm(kpis["horas_restantes"]),
            "estimativa pela média atual",
            tooltip=(
                "Estimativa de horas necessárias para concluir as tarefas restantes. "
                f"Fórmula: (horas totais ÷ tarefas concluídas) × tarefas restantes ({rest}). "
                "Baseado no ritmo médio atual. Varia conforme a dedicação futura."
            ),
        ),
        kpi_card(
            "Previsão de conclusão", prev_str,
            prev_sub[:80] + ("…" if len(prev_sub) > 80 else ""),
            tooltip=(
                "Data estimada de término do plano de estudos. "
                "Fórmula: hoje + (tarefas restantes ÷ ritmo médio semanal) semanas. "
                "Ritmo = média de tarefas concluídas por semana nas últimas semanas com dado. "
                f"Detalhes: {prev_sub}"
            ),
        ),
    ]
    for col, card in zip(cols, cards):
        with col:
            render_html(card)


def _render_kpis_desempenho(kpis: dict):
    """Bloco: Desempenho, Questões feitas, Concluídas no período, Dias ativos."""
    des   = kpis["desempenho"]
    q     = kpis["questoes"]
    ac    = kpis["acertos"]
    conc  = kpis["conc_periodo"]   # todas concluídas no período filtrado
    total = kpis["total_tarefas"]

    cols = st.columns(4)
    cards = [
        kpi_card(
            "Desempenho geral", f"{des:.1f}%",
            f"{ac} acertos em {q} questões",
            tooltip=(
                "Taxa de acerto nas questões feitas em atividades do período filtrado. "
                "Fórmula: (acertos ÷ questões feitas) × 100. "
                "≥ 70% = satisfatório para concursos. "
                "Filtros ativos: aluno, disciplina, tipo, período."
            ),
        ),
        kpi_card(
            "Questões feitas", f"{q:,}".replace(",", "."),
            "no período e filtros selecionados",
            tooltip=(
                "Soma de qtd_questoes_feitas em todas as execuções iniciadas/concluídas "
                "dentro dos filtros selecionados. "
                "Quanto mais questões, maior o treino e a familiarização com o estilo das provas."
            ),
        ),
        kpi_card(
            "Concluídas", f"{conc}",
            f"de {total} tarefas no período filtrado",
            tooltip=(
                "Quantidade de tarefas com status CONCLUIDA dentro do período e filtros selecionados. "
                "Fórmula: contagem de execuções com status=CONCLUIDA. "
                "Filtros ativos: aluno, disciplina, tipo, período de datas."
            ),
        ),
        kpi_card(
            "Dias ativos", f"{kpis['qtd_dias_ativos']}",
            "dias com pelo menos 1 registro (histórico total)",
            tooltip=(
                "Número de dias distintos com pelo menos uma execução no escopo filtrado. "
                "⚠️ Não é afetado pelo filtro de período — considera todo o histórico do aluno. "
                "Maior número = hábito de estudo mais sólido e frequência consistente."
            ),
        ),
    ]
    for col, card in zip(cols, cards):
        with col:
            render_html(card)


def _aba_visao_geral(df_filtrado, analisavel, temporal=None):
    status_df = df_filtrado.groupby("status", as_index=False)["tarefa_id"].count().rename(columns={"tarefa_id":"qtd"})
    status_df["label"] = status_df["status"].map(STATUS_LABELS)
    col_a, col_b = st.columns(2)
    with col_a:
        render_html(_tooltip_grafico(
            "Barras verticais mostrando quantas tarefas estão em cada status. "
            "Não iniciada = fila pendente · Em andamento = em execução · Concluída = finalizada. "
            "Quanto mais barras verdes, melhor o avanço geral."
        ))
        fig = go.Figure(go.Bar(
            x=status_df["label"], y=status_df["qtd"],
            marker_color=[STATUS_CORES.get(s,"#94a3b8") for s in status_df["status"]],
            text=status_df["qtd"], textposition="outside"))
        fig.update_layout(title="Distribuição por status", showlegend=False)
        st.plotly_chart(fig_layout(fig, 300), use_container_width=True)
    with col_b:
        if not analisavel.empty and "aluno" in analisavel.columns:
            # horas pelas sessões, conclusões pelas execuções
            base_pa = temporal if temporal is not None and not temporal.empty else analisavel
            pa = base_pa.groupby("aluno", as_index=False).agg(horas=("ch_efetiva","sum"))
            pa = pa.merge(
                analisavel.groupby("aluno", as_index=False).agg(
                    concluidas=("status", lambda s:(s==STATUS_CONCLUIDA).sum())),
                on="aluno", how="outer",
            ).fillna(0).sort_values("horas", ascending=True)
            render_html(_tooltip_grafico(
                "Horas estudadas por aluno. O rótulo de cada barra traz também quantas "
                "tarefas ele concluiu nesse tempo e quantas horas custou cada conclusão — "
                "quanto menor esse número, mais tarefas por hora de estudo."
            ))
            # Um eixo só, uma unidade só. Horas e contagem de tarefas em escalas
            # sobrepostas eram ilegíveis: os rótulos "41" e "104h" caíam no mesmo
            # ponto e as barras não tinham base de comparação.
            pa["h_por_tarefa"] = pa.apply(
                lambda r: r["horas"] / r["concluidas"] if r["concluidas"] else 0, axis=1)
            rotulos = [
                f"{horas_para_hm(h)} · {int(c)} concluída(s)"
                + (f" · {horas_para_hm(hp)}/tarefa" if c else "")
                for h, c, hp in zip(pa["horas"], pa["concluidas"], pa["h_por_tarefa"])
            ]
            fig = go.Figure(go.Bar(
                y=pa["aluno"], x=pa["horas"], orientation="h",
                marker_color="#3b82f6",
                text=rotulos, textposition="outside",
                textfont=dict(size=11),
                hovertemplate="%{y}<br>%{x:.2f} horas<extra></extra>",
            ))
            fig.update_layout(
                title="Horas estudadas e tarefas concluídas por aluno",
                xaxis=dict(title="Horas", rangemode="tozero"),
            )
            # espaço à direita para o rótulo não sair do gráfico
            fig.update_xaxes(range=[0, float(pa["horas"].max()) * 1.75])
            st.plotly_chart(fig_layout(fig, max(220, len(pa) * 90)), use_container_width=True)
        else:
            st.plotly_chart(grafico_vazio("Sem atividades para comparação."), use_container_width=True)


def _aba_disciplinas(analisavel, df_total=None, temporal=None):
    """
    df_total   = df filtrado completo (todos os status) → denominador do progresso.
    analisavel = Em andamento + Concluídas, uma linha por tarefa → contagens.
    temporal   = uma linha por sessão, com a data real → horas, questões e acertos.
    """
    if analisavel.empty:
        st.info("Sem dados analisáveis."); return

    # Usa df_total quando disponível para progresso correto; senão cai em analisavel
    df_prog = df_total if df_total is not None and not df_total.empty else analisavel
    base_h  = temporal if temporal is not None and not temporal.empty else analisavel

    # Progresso: todos os status por disciplina
    prog = df_prog.groupby("disciplina", as_index=False).agg(
        total_disc=("tarefa_id", "count"),
        concl_disc=("status", lambda s: (s == STATUS_CONCLUIDA).sum()),
    )
    prog["progresso"] = prog.apply(
        lambda r: r["concl_disc"] / r["total_disc"] * 100 if r["total_disc"] else 0, axis=1
    )

    # Horas / questões / desempenho: nível de sessão, para o período recortar certo
    perf = base_h.groupby("disciplina", as_index=False).agg(
        horas=("ch_efetiva", "sum"),
        questoes=("qtd_questoes_feitas", "sum"),
        acertos=("qtd_acertos", "sum"),
    )
    # Tarefas analisáveis continuam vindo de analisavel: no frame temporal uma
    # tarefa estudada em vários dias apareceria várias vezes.
    perf = perf.merge(
        analisavel.groupby("disciplina", as_index=False).agg(tarefas_anal=("tarefa_id", "count")),
        on="disciplina", how="outer",
    ).fillna(0)
    perf["desempenho"] = perf.apply(
        lambda r: r["acertos"] / r["questoes"] * 100 if r["questoes"] else 0, axis=1
    )

    disc = prog.merge(perf, on="disciplina", how="left").fillna(0)

    col_a, col_b = st.columns(2)
    with col_a:
        render_html(_tooltip_grafico(
            "Barra horizontal mostrando o % de tarefas concluídas por disciplina. "
            "Denominador = TODAS as tarefas da disciplina (concluídas + em andamento + não iniciadas). "
            "Fórmula: (concluídas ÷ total) × 100. "
            "Verde ≥ 80% · Amarelo 40–79% · Vermelho < 40%."
        ))
        d = disc.sort_values("progresso")
        fig = go.Figure(go.Bar(
            x=d["progresso"], y=d["disciplina"], orientation="h",
            text=[f"{p:.0f}%  ({int(c)}/{int(t)})"
                  for p, c, t in zip(d["progresso"], d["concl_disc"], d["total_disc"])],
            textposition="outside", textfont=dict(size=10),
            marker_color=d["progresso"].map(
                lambda v: "#22c55e" if v >= 80 else ("#f59e0b" if v >= 40 else "#ef4444")
            ),
            hovertemplate="%{y}<br>%{x:.1f}% concluído<extra></extra>",
        ))
        # Escala pelo maior valor, com folga para o rótulo. Fixar 0–100 achataria
        # todas as barras num plano em que o progresso real mal passa de 20%.
        topo = min(105, max(10, float(d["progresso"].max()) * 1.9))
        fig.update_layout(
            title="Progresso por disciplina (% de tarefas concluídas)",
            xaxis=dict(range=[0, topo], title="% concluído"),
        )
        st.plotly_chart(fig_layout(fig, max(300, len(disc) * 34)), use_container_width=True)

    with col_b:
        render_html(_tooltip_grafico(
            "Cada bolha é uma disciplina. Eixo horizontal = horas estudadas, "
            "eixo vertical = taxa de acerto, tamanho = volume de questões. "
            "Canto inferior direito = muitas horas e pouco acerto, onde o estudo "
            "está rendendo menos. A linha tracejada marca 70% de acerto. "
            "Só aparecem disciplinas com questões respondidas."
        ))
        # Antes eram barras de horas + linha de % com eixo duplo e os nomes das
        # disciplinas no eixo x. Com nomes longos os rótulos se sobrepunham e a
        # relação "muitas horas, pouco acerto" — o que o gráfico quer mostrar —
        # exigia cruzar duas escalas de cabeça. A dispersão mostra isso direto.
        bolhas = disc[disc["questoes"] > 0].copy()
        if bolhas.empty:
            st.plotly_chart(grafico_vazio("Nenhuma disciplina com questões respondidas."),
                            use_container_width=True)
        else:
            fig = go.Figure(go.Scatter(
                x=bolhas["horas"], y=bolhas["desempenho"],
                mode="markers+text",
                text=bolhas["disciplina"].map(lambda s: quebrar_texto(str(s), 22)),
                textposition="top center", textfont=dict(size=10, color="#475569"),
                marker=dict(
                    size=bolhas["questoes"], sizemode="area",
                    sizeref=2.0 * max(bolhas["questoes"].max(), 1) / (46.0 ** 2),
                    sizemin=8, color="#3b82f6", opacity=.75,
                    line=dict(color="#1e40af", width=1),
                ),
                hovertemplate=("<b>%{customdata[0]}</b><br>%{x:.1f} horas"
                               "<br>%{y:.1f}% de acerto<br>%{customdata[1]} questões<extra></extra>"),
                customdata=list(zip(bolhas["disciplina"], bolhas["questoes"].astype(int))),
            ))
            fig.add_hline(y=70, line=dict(color="#f59e0b", width=1, dash="dot"),
                          annotation_text="70%", annotation_position="right",
                          annotation_font=dict(size=10, color="#b45309"))
            fig.update_layout(
                title="Onde as horas estão rendendo",
                xaxis=dict(title="Horas estudadas", rangemode="tozero"),
                yaxis=dict(title="Taxa de acerto (%)", range=[0, 105]),
            )
            st.plotly_chart(fig_layout(fig, 380), use_container_width=True)

    # Só disciplinas com questões respondidas. Com o denominador zerado o
    # desempenho vira 0% e a disciplina aparecia no radar como se o aluno
    # tivesse errado tudo — eram 7 de 12 no caso real, retraindo o polígono
    # inteiro por falta de dado, não por falta de acerto.
    disc_desemp = disc[disc["questoes"] > 0]
    if len(disc_desemp) >= 3:
        render_html(_tooltip_grafico(
            "Radar comparando a taxa de acerto (%) de cada disciplina. "
            "Disciplinas mais distantes do centro têm melhor desempenho. "
            "Só entram disciplinas com questões respondidas — sem questões não há "
            "taxa de acerto, e um 0% ali significaria ausência de dado, não erro."
        ))
        fig_r = go.Figure(go.Scatterpolar(
            r=disc_desemp["desempenho"].tolist() + [disc_desemp["desempenho"].tolist()[0]],
            theta=disc_desemp["disciplina"].tolist() + [disc_desemp["disciplina"].tolist()[0]],
            fill="toself", fillcolor="rgba(59,130,246,0.15)",
            line=dict(color="#3b82f6", width=2),
        ))
        fig_r.update_layout(
            title="Radar de desempenho",
            polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
            showlegend=False,
        )
        st.plotly_chart(fig_layout(fig_r, 360), use_container_width=True)

    st.dataframe(
        disc.rename(columns={
            "progresso": "Progresso (%)", "desempenho": "Desempenho (%)",
            "horas": "Horas", "questoes": "Questões", "acertos": "Acertos",
            "total_disc": "Total tarefas", "concl_disc": "Concluídas",
        }),
        use_container_width=True, hide_index=True,
    )


def _aba_evolucao(analisavel, temporal=None):
    """
    Todo gráfico desta aba é série temporal, então a base são as SESSÕES: cada uma
    com a data em que aconteceu. Usar `analisavel` empilhava as horas inteiras de
    uma tarefa no dia em que ela terminou.
    As conclusões por semana continuam vindo de `analisavel`, porque ali a data de
    execução é justamente a data em que a tarefa foi concluída.
    """
    base = temporal if temporal is not None and not temporal.empty else analisavel
    if base.empty or "data_ref" not in base.columns:
        st.info("Sem dados de evolução."); return
    evo = base.dropna(subset=["data_ref"]).copy()
    if evo.empty:
        st.info("Sem datas de execução."); return
    evo["dia"] = evo["data_ref"].dt.date
    diario = evo.groupby("dia", as_index=False).agg(
        horas=("ch_efetiva","sum"), tarefas=("tarefa_id","nunique"))

    # Preenche os dias sem estudo com zero antes da média móvel.
    # Sem isto o rolling(7) andava sobre LINHAS, não sobre dias: 42 dias com
    # registro espalhados por 206 dias corridos faziam a "média de 7 dias"
    # cobrir semanas inteiras e ignorar todas as pausas, inflando a linha.
    diario["dia_ts"] = pd.to_datetime(diario["dia"])
    calendario = pd.date_range(diario["dia_ts"].min(), diario["dia_ts"].max(), freq="D")
    diario = (diario.set_index("dia_ts")
                    .reindex(calendario)
                    .rename_axis("dia_ts")
                    .reset_index())
    diario["horas"] = diario["horas"].fillna(0)
    diario["tarefas"] = diario["tarefas"].fillna(0)
    diario["media7"] = diario["horas"].rolling(7, min_periods=1).mean()
    render_html(_tooltip_grafico(
        "Barras = horas estudadas por dia. Linha pontilhada = média móvel de 7 dias corridos. "
        "Dias sem estudo entram como zero, então a linha cai durante as pausas — "
        "é a média por dia de calendário, não por dia estudado. "
        "Quando a média sobe: ritmo crescente. Quando desce: possível desaceleração."
    ))
    fig = go.Figure()
    fig.add_trace(go.Bar(x=diario["dia_ts"], y=diario["horas"], name="Horas/dia", marker_color="#3b82f6", opacity=0.7))
    fig.add_trace(go.Scatter(x=diario["dia_ts"], y=diario["media7"], name="Média 7 dias", mode="lines", line=dict(color="#f59e0b", width=2, dash="dot")))
    fig.update_layout(title="Evolução diária de horas + média móvel 7 dias")
    st.plotly_chart(fig_layout(fig, 300), use_container_width=True)
    col_a, col_b = st.columns(2)
    with col_a:
        sem = (evo.set_index("data_ref").resample("W")
                  .agg(horas=("ch_efetiva","sum")).reset_index())
        conc_sem = analisavel.dropna(subset=["data_ref"]).copy()
        if not conc_sem.empty:
            conc_sem = (conc_sem.set_index("data_ref").resample("W")
                        .agg(concluidas=("status", lambda x:(x==STATUS_CONCLUIDA).sum()))
                        .reset_index())
            sem = sem.merge(conc_sem, on="data_ref", how="left")
        if "concluidas" not in sem.columns:
            sem["concluidas"] = 0
        sem["concluidas"] = sem["concluidas"].fillna(0)
        if not sem.empty:
            render_html(_tooltip_grafico(
                "Barras = horas por semana (eixo esquerdo). "
                "Linha = tarefas concluídas na mesma semana (eixo direito). "
                "Semanas com muitas horas e poucas conclusões indicam tarefas longas "
                "ou dificuldade de fechar o que foi começado."
            ))
            # Eixo de datas em vez de rótulos "Sem dd/mm": com 30 semanas os textos
            # se sobrepunham. Num eixo temporal o Plotly escolhe a densidade de ticks.
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(
                x=sem["data_ref"], y=sem["horas"], name="Horas", marker_color="#3b82f6",
                hovertemplate="Semana de %{x|%d/%m}<br>%{y:.2f}h<extra></extra>"))
            fig2.add_trace(go.Scatter(
                x=sem["data_ref"], y=sem["concluidas"], name="Concluídas",
                mode="lines+markers", yaxis="y2", marker_color="#22c55e", line=dict(width=2),
                hovertemplate="Semana de %{x|%d/%m}<br>%{y:.0f} concluída(s)<extra></extra>"))
            fig2.update_layout(
                title="Horas × Conclusões semanais",
                xaxis=dict(type="date", tickformat="%d/%m", title="Semana"),
                yaxis=dict(title="Horas"),
                yaxis2=dict(overlaying="y", side="right", rangemode="tozero",
                            showgrid=False, title="Concluídas"),
            )
            st.plotly_chart(fig_layout(fig2, 320), use_container_width=True)
    with col_b:
        ordem_dias = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        labels_dias = {"Monday":"Seg","Tuesday":"Ter","Wednesday":"Qua","Thursday":"Qui","Friday":"Sex","Saturday":"Sáb","Sunday":"Dom"}
        evo["dia_semana"] = evo["data_ref"].dt.day_name()
        # Junto das horas vai a contagem de DIAS distintos. Sem ela, "1h" na segunda
        # se lê como "estudou pouco nas segundas", quando o caso real era "estudou
        # numa única segunda-feira em sete meses". São diagnósticos diferentes.
        ds = evo.groupby("dia_semana", as_index=False).agg(
            horas=("ch_efetiva", "sum"),
            dias=("data_ref", lambda s: s.dt.date.nunique()),
        )
        ds = ds[ds["dia_semana"].isin(ordem_dias)]
        # Dias da semana sem nenhum estudo precisam aparecer como zero
        faltando = [d for d in ordem_dias if d not in set(ds["dia_semana"])]
        if faltando:
            ds = pd.concat([ds, pd.DataFrame({"dia_semana": faltando, "horas": 0, "dias": 0})],
                           ignore_index=True)
        ds["dia_semana"] = pd.Categorical(ds["dia_semana"], categories=ordem_dias, ordered=True)
        ds = ds.sort_values("dia_semana")
        ds["label"] = ds["dia_semana"].map(labels_dias)
        ds["media"] = ds.apply(lambda r: r["horas"] / r["dias"] if r["dias"] else 0, axis=1)
        if not ds.empty:
            render_html(_tooltip_grafico(
                "Horas acumuladas em cada dia da semana ao longo de todo o período — "
                "não é a média de uma semana. O número entre parênteses é em quantos "
                "dias distintos daquele dia da semana houve estudo. "
                "Uma barra baixa com poucos dias significa que o aluno raramente estuda "
                "naquele dia; baixa com muitos dias significa sessões curtas. "
                "A barra verde é o dia com mais horas acumuladas."
            ))
            fig3 = go.Figure(go.Bar(
                x=ds["label"], y=ds["horas"],
                marker_color=["#22c55e" if h == ds["horas"].max() and h > 0 else "#3b82f6"
                              for h in ds["horas"]],
                text=[f"{horas_para_hm(h)}<br><span style='font-size:9px;color:#64748b'>"
                      f"{int(d)} dia(s)</span>" if h > 0 else ""
                      for h, d in zip(ds["horas"], ds["dias"])],
                textposition="outside",
                customdata=list(zip(ds["dias"].astype(int), ds["media"])),
                hovertemplate=("%{x}<br>%{y:.2f}h no total"
                               "<br>%{customdata[0]} dia(s) com estudo"
                               "<br>média de %{customdata[1]:.2f}h por dia estudado<extra></extra>"),
            ))
            fig3.update_layout(
                title="Distribuição por dia da semana (acumulado do período)",
                yaxis=dict(title="Horas acumuladas"),
            )
            fig3.update_yaxes(range=[0, float(ds["horas"].max() or 1) * 1.28])
            st.plotly_chart(fig_layout(fig3, 330), use_container_width=True)


def _aba_gestao_tempo(analisavel, temporal=None):
    """Distribuição de horas — base em sessões, cada uma com o próprio tipo e data."""
    base = temporal if temporal is not None and not temporal.empty else analisavel
    analisavel = base
    if analisavel.empty:
        st.info("Sem dados para análise de tempo."); return
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        tipos = analisavel.groupby("tipo", as_index=False).agg(horas=("ch_efetiva","sum"))
        tipos = tipos[tipos["horas"]>0].sort_values("horas", ascending=False)
        render_html(_tooltip_grafico(
            "Pizza mostrando como as horas de estudo estão distribuídas entre os diferentes tipos de atividade "
            "(Leitura, Exercícios, Revisão, Videoaula etc.). "
            "Ideal para identificar se há desequilíbrio — ex: muita teoria e pouca prática."
        ))
        fig = px.pie(tipos, names="tipo", values="horas", color_discrete_sequence=px.colors.qualitative.Set3)
        fig.update_traces(textposition="inside", textinfo="percent+label")
        fig.update_layout(title="Distribuição por tipo de estudo", showlegend=False)
        st.plotly_chart(fig_layout(fig, 300), use_container_width=True)
    with col_b:
        bd = analisavel.groupby("disciplina", as_index=False).agg(horas=("ch_efetiva","sum")).sort_values("horas", ascending=True)
        tot = bd["horas"].sum() or 1
        bd["pct"] = bd["horas"]/tot*100
        render_html(_tooltip_grafico(
            "Barras horizontais com horas acumuladas por disciplina. "
            "O % indica a proporção de cada disciplina no tempo total de estudo. "
            "Disciplinas com % muito baixo podem estar sendo negligenciadas."
        ))
        fig2 = go.Figure(go.Bar(x=bd["horas"], y=bd["disciplina"], orientation="h",
            text=bd["pct"].map(lambda v: f"{v:.0f}%"), textposition="outside", marker_color="#3b82f6"))
        fig2.update_layout(title="Horas por disciplina")
        st.plotly_chart(fig_layout(fig2, max(280, len(bd)*36)), use_container_width=True)
    with col_c:
        ef = analisavel.groupby("tipo", as_index=False).agg(questoes=("qtd_questoes_feitas","sum"), horas=("ch_efetiva","sum"))
        ef = ef[ef["horas"]>0]
        ef["eficiencia"] = ef["questoes"]/ef["horas"]
        ef = ef.sort_values("eficiencia", ascending=False)
        if not ef.empty:
            render_html(_tooltip_grafico(
                "Quantas questões são feitas por hora em cada tipo de atividade. "
                "Fórmula: questões feitas ÷ horas de estudo. "
                "A barra verde destaca o tipo mais eficiente em volume de prática."
            ))
            fig3 = go.Figure(go.Bar(x=ef["tipo"], y=ef["eficiencia"],
                text=ef["eficiencia"].map(lambda v: f"{v:.1f}"), textposition="outside",
                marker_color=["#22c55e" if v==ef["eficiencia"].max() else "#3b82f6" for v in ef["eficiencia"]]))
            fig3.update_layout(title="Questões/hora por tipo")
            st.plotly_chart(fig_layout(fig3, 300), use_container_width=True)
        else:
            st.plotly_chart(grafico_vazio("Sem dados de questões."), use_container_width=True)


def _aba_ranking(analisavel, temporal=None):
    """
    Horas, questões e dias ativos saem das sessões; conclusões saem de `analisavel`,
    que tem uma linha por tarefa. Misturar as duas coisas numa agregação só faria a
    contagem de tarefas repetir a cada sessão.
    """
    if analisavel.empty:
        st.info("Sem dados para ranking."); return
    base = temporal if temporal is not None and not temporal.empty else analisavel

    ag = base.groupby("aluno", as_index=False).agg(
        horas=("ch_efetiva","sum"),
        questoes=("qtd_questoes_feitas","sum"), acertos=("qtd_acertos","sum"),
        dias=("data_ref", lambda s: s.dropna().dt.date.nunique()),
    )
    conc = analisavel.groupby("aluno", as_index=False).agg(
        concluidas=("status", lambda s:(s==STATUS_CONCLUIDA).sum()),
    )
    ag = ag.merge(conc, on="aluno", how="outer").fillna(0)
    ag["desempenho"]   = ag.apply(lambda r: r["acertos"]/r["questoes"]*100 if r["questoes"] else 0, axis=1)
    ag["produtividade"] = ag.apply(lambda r: r["concluidas"]/r["horas"] if r["horas"] else 0, axis=1)

    def _rank_rows(df_s, col, fmt, cor):
        mx = df_s[col].max() or 1; html = ""
        medals = {0:"🥇",1:"🥈",2:"🥉"}
        for i,(_, r) in enumerate(df_s.iterrows()):
            pos = medals.get(i, str(i+1))
            pct = r[col]/mx*100
            val = fmt.format(r[col])
            html += (f'<div class="rank-row"><div class="rank-pos">{pos}</div>'
                f'<div class="rank-name">{escape_html(r["aluno"])}</div>'
                f'<div class="rank-bar-wrap"><div class="rank-bar" style="width:{pct:.0f}%;background:{cor}"></div></div>'
                f'<div class="rank-val">{escape_html(val)}</div></div>')
        return html

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        render_html('<div class="section-title">🏆 Produtividade (t/h)</div>')
        render_html(_rank_rows(ag.sort_values("produtividade",ascending=False), "produtividade", "{:.2f}", "#3b82f6"))
    with col_b:
        render_html('<div class="section-title">🔥 Consistência (dias)</div>')
        render_html(_rank_rows(ag.sort_values("dias",ascending=False), "dias", "{:.0f}", "#22c55e"))
    with col_c:
        render_html('<div class="section-title">📊 Desempenho (%)</div>')
        render_html(_rank_rows(ag.sort_values("desempenho",ascending=False), "desempenho", "{:.1f}%", "#f59e0b"))
    st.markdown("---")
    st.dataframe(ag.sort_values("produtividade",ascending=False).rename(columns={
        "horas":"Horas","concluidas":"Concluídas","questoes":"Questões","acertos":"Acertos",
        "dias":"Dias ativos","desempenho":"Desempenho (%)","produtividade":"Produtividade (t/h)"}),
        use_container_width=True, hide_index=True)


def _aba_analise_ia(df_escopo, df_periodo, visao, kpis, inicio_periodo=None, fim_periodo=None):
    analisavel = kpis["analisavel"]
    if analisavel.empty:
        st.info("Sem atividades para análise.")
        return

    # visao pode ser "Todos", um nome (str) ou uma lista de nomes
    if isinstance(visao, list):
        alunos_lista = visao
    elif visao == "Todos":
        alunos_lista = sorted(analisavel["aluno"].dropna().unique().tolist())
    else:
        alunos_lista = [visao]

    for nome_aluno in alunos_lista:
        grupo_esc = df_escopo[df_escopo["aluno"] == nome_aluno].copy()
        grupo_per = df_periodo[df_periodo["aluno"] == nome_aluno].copy()
        grupo     = grupo_esc.copy()
        grupo["data_ref"] = pd.to_datetime(
            grupo.get("data_execucao", pd.Series(dtype=str)), errors="coerce"
        )
        # O período precisa ser repassado: sem ele o divisor de dias úteis é outro
        # e a aba mostrava média diária diferente da dos cards, para o mesmo aluno.
        kpis_al  = calcular_kpis_avancados(
            grupo_esc, grupo_per, inicio_periodo, fim_periodo
        )
        insights = gerar_insights(grupo, kpis_al, nome_aluno)

        with st.expander(f"🧠 {nome_aluno}", expanded=(len(alunos_lista) == 1)):
            c1, c2, c3, c4 = st.columns(4)
            dh = kpis_al["delta_horas_semana"]
            delta_h = ("+" if dh >= 0 else "-") + horas_para_hm(abs(dh)) + " vs sem. ant."
            c1.metric("Horas esta semana", horas_para_hm(kpis_al["horas_semana"]), delta=delta_h)
            c2.metric("Progresso geral", f"{kpis_al['pct_conclusao']:.1f}%")
            c3.metric("Desempenho", f"{kpis_al['desempenho']:.1f}%")
            c4.metric(
                "Sequência", f"{kpis_al['sequencia']} dias",
                delta=(f"{kpis_al['dias_sem_estudar']}d sem estudar"
                       if kpis_al["dias_sem_estudar"] > 0 else "Estudou hoje"),
                help="Dias consecutivos com registro, contados a partir do último "
                     "dia estudado. O delta mostra há quantos dias foi esse último registro.",
            )
            render_html('<div class="section-title">Análise automática</div>')
            for ins in insights:
                render_html(insight_card(ins["tipo"], ins["icone"], ins["titulo"], ins["texto"]))

            ana_al = kpis_al["analisavel"]
            tmp_al = kpis_al.get("temporal")
            base_al = tmp_al if tmp_al is not None and not tmp_al.empty else ana_al
            if not base_al.empty and "data_ref" in base_al.columns:
                evo = base_al.dropna(subset=["data_ref"]).copy()
                if not evo.empty:
                    # horas pelas sessões, conclusões pelas execuções
                    sem = (evo.set_index("data_ref").resample("W")
                             .agg(horas=("ch_efetiva", "sum")).reset_index())
                    ca = ana_al.dropna(subset=["data_ref"])
                    if not ca.empty:
                        ca = (ca.set_index("data_ref").resample("W")
                                .agg(concluidas=("status", lambda x: (x == STATUS_CONCLUIDA).sum()))
                                .reset_index())
                        sem = sem.merge(ca, on="data_ref", how="left")
                    if "concluidas" not in sem.columns:
                        sem["concluidas"] = 0
                    sem["concluidas"] = sem["concluidas"].fillna(0)
                    if len(sem) > 1:
                        fig = go.Figure()
                        fig.add_trace(go.Bar(
                            x=sem["data_ref"], y=sem["horas"], name="Horas",
                            marker_color="#3b82f6",
                            hovertemplate="Semana de %{x|%d/%m}<br>%{y:.2f}h<extra></extra>"))
                        fig.add_trace(go.Scatter(
                            x=sem["data_ref"], y=sem["concluidas"], name="Concluídas",
                            mode="lines+markers", yaxis="y2", marker_color="#22c55e",
                            line=dict(width=2),
                            hovertemplate="Semana de %{x|%d/%m}<br>%{y:.0f} concluída(s)<extra></extra>"))
                        fig.update_layout(
                            title=f"Evolução semanal — {nome_aluno}",
                            xaxis=dict(type="date", tickformat="%d/%m"),
                            yaxis=dict(title="Horas"),
                            yaxis2=dict(overlaying="y", side="right", rangemode="tozero",
                                        showgrid=False),
                        )
                        st.plotly_chart(fig_layout(fig, 280), use_container_width=True)

            if not ana_al.empty and "disciplina" in ana_al.columns:
                # questões e acertos pelas sessões do período; contagens por tarefa
                da = base_al.groupby("disciplina", as_index=False).agg(
                    questoes=("qtd_questoes_feitas", "sum"),
                    acertos=("qtd_acertos", "sum"),
                )
                da = da.merge(
                    ana_al.groupby("disciplina", as_index=False).agg(
                        iniciadas=("tarefa_id", "count"),
                        concluidas=("status", lambda s: (s == STATUS_CONCLUIDA).sum()),
                    ),
                    on="disciplina", how="outer",
                ).fillna(0)
                da["desempenho"] = da.apply(
                    lambda r: r["acertos"] / r["questoes"] * 100 if r["questoes"] else 0, axis=1)
                # Progresso sobre o total da disciplina no escopo do aluno, não
                # sobre as tarefas que ele já iniciou.
                total_disc = grupo_esc.groupby("disciplina")["tarefa_id"].count()
                da["tarefas"] = da["disciplina"].map(total_disc).fillna(da["iniciadas"]).astype(int)
                da["progresso"]  = da.apply(
                    lambda r: r["concluidas"] / r["tarefas"] * 100 if r["tarefas"] else 0, axis=1)
                frageis  = da[(da["questoes"] > 0) & (da["desempenho"] < 70)].sort_values("desempenho")
                criticas = da[(da["progresso"] < 30) & (da["iniciadas"] >= 3)].sort_values("progresso")
                if not frageis.empty or not criticas.empty:
                    render_html('<div class="section-title">Disciplinas que precisam de atenção</div>')
                    fc1, fc2 = st.columns(2)
                    with fc1:
                        if not frageis.empty:
                            st.caption("🔴 Baixo desempenho em questões (<70%)")
                            st.dataframe(
                                frageis[["disciplina", "desempenho", "questoes", "acertos"]].rename(
                                    columns={"disciplina": "Disciplina", "desempenho": "Desempenho (%)",
                                             "questoes": "Questões", "acertos": "Acertos"}),
                                hide_index=True, use_container_width=True,
                            )
                    with fc2:
                        if not criticas.empty:
                            st.caption("⚠️ Baixo progresso (<30% concluído)")
                            st.dataframe(
                                criticas[["disciplina", "progresso", "tarefas", "concluidas"]].rename(
                                    columns={"disciplina": "Disciplina", "progresso": "Progresso (%)",
                                             "tarefas": "Tarefas", "concluidas": "Concluídas"}),
                                hide_index=True, use_container_width=True,
                            )


# ─────────────────────────────────────────────
# TELA: DASHBOARD
# ─────────────────────────────────────────────

def _tooltip_grafico(texto: str) -> str:
    """Retorna HTML de um ícone ? com tooltip flutuante para gráficos."""
    return (
        f'<div class="kpi-ttip-wrap" style="display:inline-block;margin-left:6px;vertical-align:middle">'
        f'<div class="kpi-ttip-icon">?</div>'
        f'<div class="kpi-ttip-box">{escape_html(texto)}</div>'
        f'</div>'
    )


def _titulo_secao(label: str, tooltip: str = "", periodo: str = "") -> None:
    badge = ""
    if periodo:
        if "Comparativo" in periodo:
            cor, borda, txt = "#f5f3ff", "#ddd6fe", "#5b21b6"  # roxo — comparativo
        elif "15" in periodo:
            cor, borda, txt = "#fef3c7", "#fde68a", "#92400e"   # âmbar
        elif "7" in periodo:
            cor, borda, txt = "#eff6ff", "#bfdbfe", "#1d4ed8"   # azul
        else:
            cor, borda, txt = "#f0fdf4", "#bbf7d0", "#166534"   # verde
        badge = (
            f'<span style="background:{cor};color:{txt};border:1px solid {borda};'
            f'border-radius:999px;padding:2px 10px;font-size:.67rem;font-weight:800;'
            f'margin-left:8px;vertical-align:middle">{periodo}</span>'
        )
    tip = _tooltip_grafico(tooltip) if tooltip else ""
    render_html(
        f'<div style="font-size:.88rem;font-weight:800;color:#0f172a;margin:18px 0 6px;'
        f'display:flex;align-items:center;gap:4px">'
        f'{escape_html(label)}{badge}{tip}</div>'
    )


def _resumo_15dias(df_total: pd.DataFrame) -> None:
    """
    KPIs dos últimos 15 dias. Horas, questões, acertos e dias ativos vêm das
    sessões da janela; conclusões vêm das execuções, porque `data_execucao` é a
    data em que a tarefa foi de fato concluída.
    """
    hoje  = date.today()
    ini15 = hoje - timedelta(days=14)
    df15  = df_total[df_total["data_ref"].dt.date >= ini15].copy()
    ana15 = df15[df15["status"].isin(STATUS_ANALISE)]

    sess15 = _sessoes_do_escopo(df_total, ini15, hoje)
    pares  = _pares_aluno_tarefa(_sessoes_do_escopo(df_total))
    leg15  = _sem_sessao(ana15, pares)
    base15 = _frame_temporal(sess15, leg15)

    h15    = float(base15["ch_efetiva"].sum())
    q15    = int(base15["qtd_questoes_feitas"].sum())
    ac15   = int(base15["qtd_acertos"].sum())
    des15  = ac15 / q15 * 100 if q15 else 0
    conc15 = int((ana15["status"] == STATUS_CONCLUIDA).sum())
    dias15 = base15["data_ref"].dropna().dt.date.nunique()

    _titulo_secao(
        "Últimos 15 dias",
        "Indicadores calculados exclusivamente com execuções dos últimos 15 dias (hoje inclusive). "
        "Permite monitorar o ritmo recente de forma mais representativa que uma semana. "
        "Não é afetado pelo filtro de período — sempre considera os 15 dias anteriores a hoje. "
        "Filtros de aluno e disciplina são respeitados.",
        "Últimos 15 dias",
    )
    cols  = st.columns(5)
    cards = [
        kpi_card(
            "Horas (15d)", horas_para_hm(h15),
            f"{dias15} dia(s) com registro",
            tooltip=(
                "Total de horas de estudo registradas nos últimos 15 dias. "
                "Inclui atividades Em andamento e Concluídas. "
                "Filtros de aluno e disciplina são respeitados; filtro de período não se aplica aqui. "
                "Fórmula: soma das sessões de estudo com data ≥ hoje − 14 dias."
            ),
        ),
        kpi_card(
            "Questões (15d)", f"{q15}",
            "nos últimos 15 dias",
            tooltip=(
                "Soma de todas as questões feitas em execuções com data nos últimos 15 dias. "
                "Fórmula: soma de qtd_questoes_feitas das execuções do período de 15 dias."
            ),
        ),
        kpi_card(
            "Acertos (15d)", f"{ac15}",
            f"de {q15} questões",
            tooltip=(
                "Total de questões acertadas nos últimos 15 dias. "
                "Fórmula: soma de qtd_acertos das execuções dos últimos 15 dias."
            ),
        ),
        kpi_card(
            "Desempenho (15d)", f"{des15:.1f}%",
            "taxa de acerto — 15 dias",
            tooltip=(
                "Taxa de acerto nas questões feitas nos últimos 15 dias. "
                "Fórmula: (acertos ÷ questões) × 100. "
                "Acima de 70% é satisfatório para concursos públicos."
            ),
        ),
        kpi_card(
            "Concluídas (15d)", f"{conc15}",
            "tarefas concluídas em 15 dias",
            tooltip=(
                "Quantidade de tarefas marcadas como Concluída "
                "com data de execução nos últimos 15 dias."
            ),
        ),
    ]
    for col, card in zip(cols, cards):
        with col:
            render_html(card)


def _card_comparativo(label: str, valor: str, subtexto: str = "", tooltip: str = "") -> str:
    """Card simples para uso na grade comparativa."""
    tip = ""
    if tooltip:
        tip = (
            '<div class="kpi-ttip-wrap" style="position:absolute;top:6px;right:6px">'
            '<div class="kpi-ttip-icon">?</div>'
            f'<div class="kpi-ttip-box">{escape_html(tooltip)}</div>'
            '</div>'
        )
    return (
        f'<div class="kpi-card" style="min-height:80px">{tip}'
        f'<div class="kpi-label">{escape_html(label)}</div>'
        f'<div class="kpi-value" style="font-size:1.2rem">{escape_html(valor)}</div>'
        f'<div class="kpi-sub">{escape_html(subtexto)}</div>'
        f'</div>'
    )


def _dashboard_comparativo(
    df_escopo: pd.DataFrame,
    df_periodo: pd.DataFrame,
    alunos_sel: list,
    inicio_periodo,
    fim_periodo,
):
    """
    Modo comparativo: KPIs e gráficos separados por aluno, sem somar dados.
    Ativado quando mais de um aluno é selecionado no filtro.
    """
    import plotly.graph_objects as go

    CORES_ALUNOS = [
        "#3b82f6","#22c55e","#f59e0b","#ef4444",
        "#8b5cf6","#06b6d4","#ec4899","#84cc16",
    ]

    render_html(
        '<div class="insight-card info" style="margin-bottom:16px">'
        '<div class="insight-icon">👥</div>'
        '<div class="insight-body">'
        '<div class="insight-title">Modo comparativo ativo</div>'
        '<p class="insight-text">'
        f'Exibindo <strong>{len(alunos_sel)} aluno(s)</strong> em paralelo — '
        'dados calculados <strong>individualmente por aluno</strong>, sem somar. '
        'Para ver o dashboard individual, selecione apenas um aluno no filtro lateral. '
        'Filtros de disciplina, período e status são aplicados individualmente por aluno.'
        '</p></div></div>'
    )

    # ── KPIs individuais por aluno ──
    dados_alunos = []
    for nome in alunos_sel:
        esc_al = df_escopo[df_escopo["aluno"] == nome].copy()
        per_al = df_periodo[df_periodo["aluno"] == nome].copy()
        if esc_al.empty and per_al.empty:
            continue
        k = calcular_kpis_avancados(esc_al, per_al, inicio_periodo, fim_periodo)
        dados_alunos.append({"nome": nome, "kpis": k, "esc": esc_al, "per": per_al})

    if not dados_alunos:
        st.info("Sem dados para os alunos selecionados com os filtros aplicados.")
        return

    # ── Grade de cards por aluno ──
    _titulo_secao(
        "KPIs por aluno",
        "Cada coluna é um aluno — valores calculados individualmente, sem somar. "
        "Contagens estruturais (progresso, restantes) ignoram filtro de período. "
        "Horas e desempenho respeitam o período selecionado.",
        "Comparativo",
    )

    cols = st.columns(min(len(dados_alunos), 4))
    for i, d in enumerate(dados_alunos):
        k    = d["kpis"]
        nome = d["nome"]
        prev = k["previsao_conclusao"]
        cor  = CORES_ALUNOS[i % len(CORES_ALUNOS)]

        with cols[i % len(cols)]:
            # Cabeçalho do aluno
            render_html(
                f'<div style="background:{cor};color:#fff;border-radius:10px 10px 0 0;'
                f'padding:10px 14px;font-weight:800;font-size:.88rem;text-align:center;'
                f'margin-bottom:0">'
                f'{escape_html(nome)}</div>'
            )
            kpis_itens = [
                ("Total de horas",       horas_para_hm(k["horas_total"]),
                 "no período",
                 "Soma de ch_efetiva das execuções com data no período. Não inclui tarefas sem data."),
                ("Progresso",            f"{k['pct_conclusao']:.1f}%",
                 f"{k['qtd_concluidas']} / {k['total_tarefas']} tarefas",
                 "Concluídas ÷ total de tarefas do escopo (sem filtro de período)."),
                ("Em andamento",         str(k["qtd_andamento"]),
                 f"{k['qtd_nao_iniciada']} não iniciadas",
                 "Tarefas com status Em andamento. Não inclui filtro de período."),
                ("Restantes",            str(k["tarefas_restantes"]),
                 "andamento + não iniciadas",
                 "Total de tarefas ainda não concluídas no escopo do aluno."),
                ("Média diária (úteis)", horas_para_hm(k["media_diaria"]),
                 k["media_diaria_label"],
                 "Horas do período ÷ dias úteis (seg–sex) do período selecionado."),
                ("Desempenho",           f"{k['desempenho']:.1f}%",
                 f"{k['acertos']} acertos / {k['questoes']} questões",
                 "Taxa de acerto nas questões do período selecionado."),
                ("Produtividade",        f"{k['produtividade']:.2f}",
                 "tarefas concluídas / hora",
                 "Tarefas concluídas (escopo) ÷ horas estudadas (período)."),
                ("Previsão de conclusão",
                 prev.strftime("%d/%m/%Y") if prev else "Dados insuficientes",
                 k["previsao_base"][:55] + ("…" if len(k["previsao_base"]) > 55 else ""),
                 k["previsao_base"]),
            ]
            for lbl, val, sub, tip in kpis_itens:
                render_html(_card_comparativo(lbl, val, sub, tip))

    st.markdown("---")

    # Pre-computa per_anal_todos (usado em Evolução e Ranking)
    per_anal_todos = df_periodo[df_periodo["status"].isin(STATUS_ANALISE)].copy()
    if "data_ref" not in per_anal_todos.columns:
        per_anal_todos["data_ref"] = pd.to_datetime(
            per_anal_todos.get("data_execucao", pd.Series(dtype=str)), errors="coerce")
    per_anal_todos = per_anal_todos[per_anal_todos["aluno"].isin(alunos_sel)]

    # Base temporal do grupo: uma linha por sessão, já recortada por escopo e período
    # dentro de calcular_kpis_avancados de cada aluno.
    _temporais = [d["kpis"]["temporal"] for d in dados_alunos
                  if d["kpis"].get("temporal") is not None and not d["kpis"]["temporal"].empty]
    temporal_todos = (pd.concat(_temporais, ignore_index=True)
                      if _temporais else pd.DataFrame(columns=_COLS_TEMPORAL))

    # ── Gráficos ──
    abas = st.tabs([
        "📊 Horas & Progresso",
        "📚 Disciplinas",
        "📅 Evolução",
        "🎯 Desempenho",
        "🏆 Ranking",
        "🧠 Análise IA",
        "📋 Atividades",
    ])

    df_resumo = pd.DataFrame([
        {
            "Aluno":             d["nome"],
            "Horas":             round(d["kpis"]["horas_total"], 2),
            "Concluídas":        d["kpis"]["qtd_concluidas"],
            "Progresso (%)":     round(d["kpis"]["pct_conclusao"], 1),
            "Em andamento":      d["kpis"]["qtd_andamento"],
            "Restantes":         d["kpis"]["tarefas_restantes"],
            "Desempenho (%)":    round(d["kpis"]["desempenho"], 1),
            "Questões":          d["kpis"]["questoes"],
            "Acertos":           d["kpis"]["acertos"],
            "Produtividade (t/h)": round(d["kpis"]["produtividade"], 2),
        }
        for d in dados_alunos
    ])

    with abas[0]:
        render_html(_tooltip_grafico(
            "Barras comparando horas e tarefas por aluno. "
            "Horas = período selecionado. Concluídas = escopo total. "
            "Valores não somados — cada barra representa um aluno."
        ))
        col_a, col_b = st.columns(2)
        with col_a:
            fig = go.Figure()
            for i, d in enumerate(dados_alunos):
                fig.add_trace(go.Bar(
                    name=d["nome"], x=[d["nome"]], y=[d["kpis"]["horas_total"]],
                    marker_color=CORES_ALUNOS[i % len(CORES_ALUNOS)],
                    text=[horas_para_hm(d["kpis"]["horas_total"])],
                    textposition="outside",
                ))
            fig.update_layout(title="Total de horas por aluno", showlegend=False)
            st.plotly_chart(fig_layout(fig, 300), use_container_width=True)

        with col_b:
            fig = go.Figure()
            for i, d in enumerate(dados_alunos):
                fig.add_trace(go.Bar(
                    name=d["nome"], x=[d["nome"]],
                    y=[d["kpis"]["qtd_concluidas"]],
                    marker_color=CORES_ALUNOS[i % len(CORES_ALUNOS)],
                    text=[str(d["kpis"]["qtd_concluidas"])], textposition="outside",
                ))
            fig.update_layout(title="Tarefas concluídas por aluno", showlegend=False)
            st.plotly_chart(fig_layout(fig, 300), use_container_width=True)

        # Progresso em barras horizontais com label do aluno
        render_html(_tooltip_grafico(
            "Progresso geral de cada aluno: tarefas concluídas ÷ total de tarefas do escopo. "
            "Não é afetado pelo filtro de período."
        ))
        fig = go.Figure()
        for i, d in enumerate(dados_alunos):
            pct = d["kpis"]["pct_conclusao"]
            fig.add_trace(go.Bar(
                name=d["nome"], y=[d["nome"]], x=[pct],
                orientation="h",
                marker_color=("#22c55e" if pct>=80 else "#f59e0b" if pct>=40 else "#ef4444"),
                text=[f"{pct:.1f}%"], textposition="outside",
            ))
        fig.update_layout(title="Progresso geral por aluno (%)", showlegend=False,
            xaxis_range=[0, 120])
        st.plotly_chart(fig_layout(fig, max(220, len(dados_alunos)*50)), use_container_width=True)
        st.dataframe(df_resumo[["Aluno","Horas","Concluídas","Progresso (%)","Em andamento","Restantes"]],
            use_container_width=True, hide_index=True)

    with abas[1]:
        render_html(_tooltip_grafico(
            "Progresso por disciplina separado por aluno. "
            "Denominador = todas as tarefas da disciplina no escopo do aluno (sem filtro de período). "
            "Barras agrupadas: cada cor é um aluno."
        ))
        disc_rows = []
        for i, d in enumerate(dados_alunos):
            esc_a   = d["esc"]
            per_a   = d["per"]
            per_an  = per_a[per_a["status"].isin(STATUS_ANALISE)]
            dp = esc_a.groupby("disciplina", as_index=False).agg(
                total=("tarefa_id","count"),
                concluidas=("status", lambda s:(s==STATUS_CONCLUIDA).sum()),
            )
            tmp_a = d["kpis"].get("temporal")
            base_a = tmp_a if tmp_a is not None and not tmp_a.empty else per_an
            dperf = base_a.groupby("disciplina", as_index=False).agg(
                horas=("ch_efetiva","sum"),
                questoes=("qtd_questoes_feitas","sum"),
                acertos=("qtd_acertos","sum"),
            )
            dm = dp.merge(dperf, on="disciplina", how="left").fillna(0)
            dm["progresso"]  = dm.apply(lambda r: r["concluidas"]/r["total"]*100 if r["total"] else 0, axis=1)
            dm["desempenho"] = dm.apply(lambda r: r["acertos"]/r["questoes"]*100 if r["questoes"] else 0, axis=1)
            dm["Aluno"]      = d["nome"]
            disc_rows.append(dm)

        if disc_rows:
            df_disc = pd.concat(disc_rows, ignore_index=True)
            disciplinas_u = sorted(df_disc["disciplina"].unique().tolist())

            fig = go.Figure()
            for i, d in enumerate(dados_alunos):
                sub = df_disc[df_disc["Aluno"] == d["nome"]]
                fig.add_trace(go.Bar(
                    name=d["nome"], x=sub["disciplina"], y=sub["progresso"],
                    marker_color=CORES_ALUNOS[i % len(CORES_ALUNOS)],
                    text=sub["progresso"].map(lambda v: f"{v:.0f}%"),
                    textposition="outside",
                ))
            # Escala adaptativa: com progresso real abaixo de 25%, um eixo fixo
            # em 0–125 deixava as barras rentes ao chão e 80% do gráfico vazio.
            topo_prog = min(125, max(12, float(df_disc["progresso"].max()) * 1.35))
            fig.update_layout(
                title="Progresso por disciplina — por aluno (%)",
                barmode="group",
                yaxis=dict(range=[0, topo_prog], title="% concluído"),
            )
            st.plotly_chart(fig_layout(fig, max(300, len(disciplinas_u)*60)), use_container_width=True)

            # Desempenho por disciplina — só onde houve questões respondidas.
            # Sem esse recorte, disciplina sem questão virava barra de 0% e lia-se
            # como desempenho péssimo em vez de ausência de medida.
            df_des = df_disc[df_disc["questoes"] > 0]
            if not df_des.empty:
                render_html(_tooltip_grafico(
                    "Taxa de acerto por disciplina, separada por aluno. "
                    "Aparecem apenas disciplinas em que o aluno já respondeu questões."
                ))
                fig2 = go.Figure()
                for i, d in enumerate(dados_alunos):
                    sub = df_des[df_des["Aluno"] == d["nome"]]
                    if sub.empty:
                        continue
                    fig2.add_trace(go.Bar(
                        name=d["nome"], x=sub["disciplina"], y=sub["desempenho"],
                        marker_color=CORES_ALUNOS[i % len(CORES_ALUNOS)],
                        text=sub["desempenho"].map(lambda v: f"{v:.0f}%"),
                        textposition="outside",
                    ))
                topo_des = min(125, max(40, float(df_des["desempenho"].max()) * 1.28))
                fig2.update_layout(
                    title="Desempenho por disciplina — por aluno (%)",
                    barmode="group",
                    yaxis=dict(range=[0, topo_des], title="Taxa de acerto (%)"),
                )
                st.plotly_chart(fig_layout(fig2, max(300, df_des["disciplina"].nunique()*60)),
                                use_container_width=True)

            st.dataframe(
                df_disc[["Aluno","disciplina","progresso","desempenho","horas","concluidas","total"]].rename(
                    columns={"disciplina":"Disciplina","progresso":"Progresso (%)","desempenho":"Desempenho (%)",
                             "horas":"Horas","concluidas":"Concluídas","total":"Total tarefas"}),
                use_container_width=True, hide_index=True,
            )

    with abas[2]:
        render_html(_tooltip_grafico(
            "Ritmo de estudo por aluno. A linha é a média móvel de 7 dias corridos: "
            "dias sem estudo entram como zero, então a linha desce durante as pausas. "
            "Ligar só os dias estudados desenharia uma reta por cima de semanas paradas "
            "e daria a impressão de ritmo constante onde não houve estudo. "
            "Abaixo, o total de horas por semana."
        ))
        evo = (temporal_todos if not temporal_todos.empty else per_anal_todos)
        evo = evo.dropna(subset=["data_ref"]).copy()
        if not evo.empty:
            evo["dia"] = evo["data_ref"].dt.date
            diario = evo.groupby(["dia","aluno"], as_index=False).agg(horas=("ch_efetiva","sum"))
            calend = pd.date_range(
                pd.to_datetime(diario["dia"]).min(), pd.to_datetime(diario["dia"]).max(), freq="D"
            )
            fig = go.Figure()
            for i, d in enumerate(dados_alunos):
                sub = diario[diario["aluno"] == d["nome"]].sort_values("dia")
                if sub.empty:
                    continue
                serie = (sub.set_index(pd.to_datetime(sub["dia"]))["horas"]
                            .reindex(calend, fill_value=0))
                fig.add_trace(go.Scatter(
                    x=calend, y=serie.rolling(7, min_periods=1).mean(), mode="lines",
                    name=d["nome"], line=dict(color=CORES_ALUNOS[i%len(CORES_ALUNOS)], width=2),
                ))
            fig.update_layout(
                title="Ritmo diário por aluno — média móvel de 7 dias",
                yaxis_title="Horas/dia",
            )
            st.plotly_chart(fig_layout(fig, 320), use_container_width=True)

            semanal = (evo.groupby(["aluno","data_ref"])
                .agg(horas=("ch_efetiva","sum"))
                .reset_index())
            semanal = semanal.set_index("data_ref").groupby("aluno").resample("W")["horas"].sum().reset_index()
            fig2 = go.Figure()
            for i, d in enumerate(dados_alunos):
                sub = semanal[semanal["aluno"] == d["nome"]]
                if not sub.empty:
                    fig2.add_trace(go.Bar(
                        name=d["nome"], x=sub["data_ref"], y=sub["horas"],
                        marker_color=CORES_ALUNOS[i%len(CORES_ALUNOS)],
                        hovertemplate=("<b>"+str(d["nome"])+"</b><br>Semana de %{x|%d/%m}"
                                       "<br>%{y:.2f}h<extra></extra>"),
                    ))
            fig2.update_layout(
                title="Horas semanais por aluno", barmode="group",
                xaxis=dict(type="date", tickformat="%d/%m", title="Semana"),
                yaxis=dict(title="Horas"),
            )
            st.plotly_chart(fig_layout(fig2, 320), use_container_width=True)
        else:
            st.info("Sem dados de evolução para o período selecionado.")

    with abas[3]:
        render_html(_tooltip_grafico(
            "Desempenho, questões e produtividade por aluno. "
            "Calculado sobre execuções com data no período selecionado."
        ))
        col_a, col_b = st.columns(2)
        with col_a:
            fig = go.Figure()
            for i, d in enumerate(dados_alunos):
                v = d["kpis"]["desempenho"]
                fig.add_trace(go.Bar(
                    name=d["nome"], x=[d["nome"]], y=[v],
                    marker_color=("#22c55e" if v>=70 else "#f59e0b" if v>=50 else "#ef4444"),
                    text=[f"{v:.1f}%"], textposition="outside",
                ))
            fig.update_layout(title="Desempenho por aluno (%)", showlegend=False,
                yaxis_range=[0,115])
            st.plotly_chart(fig_layout(fig, 300), use_container_width=True)

        with col_b:
            fig = go.Figure()
            for i, d in enumerate(dados_alunos):
                fig.add_trace(go.Bar(
                    name=d["nome"]+" — Questões", x=[d["nome"]],
                    y=[d["kpis"]["questoes"]], marker_color=CORES_ALUNOS[i%len(CORES_ALUNOS)],
                ))
                fig.add_trace(go.Bar(
                    name=d["nome"]+" — Acertos", x=[d["nome"]],
                    y=[d["kpis"]["acertos"]],
                    marker_color=CORES_ALUNOS[i%len(CORES_ALUNOS)], opacity=0.5,
                ))
            fig.update_layout(title="Questões vs. Acertos por aluno", barmode="group")
            st.plotly_chart(fig_layout(fig, 300), use_container_width=True)

        st.dataframe(
            df_resumo[["Aluno","Desempenho (%)","Questões","Acertos","Produtividade (t/h)"]],
            use_container_width=True, hide_index=True,
        )

    with abas[4]:
        _titulo_secao(
            "Ranking comparativo",
            "Classificação dos alunos por produtividade, consistência e desempenho. "
            "Valores calculados individualmente — não somados.",
            "Comparativo",
        )
        _aba_ranking(
            per_anal_todos if not per_anal_todos.empty else df_escopo[df_escopo["status"].isin(STATUS_ANALISE)],
            temporal=temporal_todos,
        )

    with abas[5]:
        _titulo_secao(
            "Análise inteligente por aluno",
            "Insights automáticos gerados individualmente para cada aluno. "
            "Os dados não são somados — cada análise é independente por aluno.",
            "Comparativo",
        )
        render_html(
            '<div class="insight-card info" style="margin-bottom:12px">'
            '<div class="insight-icon">🧠</div>'
            '<div class="insight-body">'
            '<div class="insight-title">Análise comparativa de IA</div>'
            '<p class="insight-text">'
            'Cada expander abaixo contém a análise individual do aluno. '
            'A IA identifica padrões, riscos de atraso, disciplinas críticas e '
            'gera recomendações personalizadas com base no histórico de cada um.'
            '</p></div></div>'
        )
        kpis_grupo = calcular_kpis_avancados(df_escopo, df_periodo, inicio_periodo, fim_periodo)
        _aba_analise_ia(df_escopo, df_periodo, alunos_sel, kpis_grupo, inicio_periodo, fim_periodo)

        if len(dados_alunos) > 1:
            st.markdown("---")
            render_html('<div class="section-title">📊 Comparativo geral do grupo</div>')
            mais_avancado  = max(dados_alunos, key=lambda d: d["kpis"]["pct_conclusao"])
            menos_avancado = min(dados_alunos, key=lambda d: d["kpis"]["pct_conclusao"])
            mais_horas     = max(dados_alunos, key=lambda d: d["kpis"]["horas_total"])
            mais_produtivo = max(dados_alunos, key=lambda d: d["kpis"]["produtividade"])
            menor_seq      = min(dados_alunos, key=lambda d: d["kpis"]["sequencia"])
            insights_grupo = [
                ("success", "🏆",
                 f"{mais_avancado['nome']} está mais avançado",
                 f"Progresso de {mais_avancado['kpis']['pct_conclusao']:.1f}% — "
                 f"{mais_avancado['kpis']['qtd_concluidas']} tarefas concluídas de "
                 f"{mais_avancado['kpis']['total_tarefas']}."),
                ("warning", "⚠️",
                 f"{menos_avancado['nome']} precisa de atenção",
                 f"Progresso de {menos_avancado['kpis']['pct_conclusao']:.1f}% — "
                 f"{menos_avancado['kpis']['tarefas_restantes']} tarefas restantes."),
                ("info", "⏱️",
                 f"{mais_horas['nome']} dedicou mais horas no período",
                 f"{horas_para_hm(mais_horas['kpis']['horas_total'])} de estudo registradas."),
                ("success", "⚡",
                 f"{mais_produtivo['nome']} é o mais produtivo",
                 f"{mais_produtivo['kpis']['produtividade']:.2f} tarefas concluídas por hora."),
            ]
            if menor_seq["kpis"]["sequencia"] == 0 or menor_seq["kpis"]["dias_sem_estudar"] >= 3:
                insights_grupo.append((
                    "danger", "🔴",
                    f"{menor_seq['nome']} com baixa frequência",
                    f"Sequência atual: {menor_seq['kpis']['sequencia']} dia(s). "
                    f"{menor_seq['kpis']['dias_sem_estudar']} dia(s) sem registro."
                ))
            for tipo, icone, titulo, texto in insights_grupo:
                render_html(insight_card(tipo, icone, titulo, texto))

    with abas[6]:
        tabela = preparar_tabela(df_escopo[df_escopo["aluno"].isin(alunos_sel)])
        colunas = ["aluno","tarefa","disciplina","aula","assunto","tipo","status_label",
                   "data_execucao","tempo","qtd_questoes_feitas","qtd_acertos","desempenho","comentario"]
        st.dataframe(tabela[[c for c in colunas if c in tabela.columns]],
            use_container_width=True, hide_index=True, row_height=72)

def dashboard():
    render_html(f"""
        <div class="hero">
          <h1>📊 {APP_NAME}</h1>
          <p>Dashboard analítico: produtividade, avanço, desempenho, evolução e análise inteligente.</p>
        </div>
    """)
    df        = carregar_execucoes()
    df_sessoes = carregar_sessoes_dashboard()
    if df.empty:
        st.info("Cadastre alunos, tarefas e registros para iniciar o acompanhamento.")
        return

    # painel_filtros chamado UMA única vez — retorna escopo estrutural e período
    df_escopo, df_periodo_exec, visao, inicio_periodo, fim_periodo = painel_filtros(df, "dash")

    # Aplica os mesmos filtros de aluno/disciplina/período ao df de sessões
    # sem chamar painel_filtros novamente (evita widgets duplicados)
    df_periodo = df_sessoes.copy()
    df_periodo["data_ref"] = pd.to_datetime(df_periodo["data_execucao"], errors="coerce")

    # Filtro de aluno
    if isinstance(visao, list):
        df_periodo = df_periodo[df_periodo["aluno"].isin(visao)]
    elif visao != "Todos":
        df_periodo = df_periodo[df_periodo["aluno"] == visao]

    # Filtro de período (apenas em registros com data)
    ts_inicio = pd.Timestamp(inicio_periodo) if inicio_periodo else None
    ts_fim    = pd.Timestamp(fim_periodo)    if fim_periodo    else None
    if ts_inicio:
        sem_data   = df_periodo["data_ref"].isna()
        com_data   = df_periodo["data_ref"].notna() & (df_periodo["data_ref"] >= ts_inicio)
        df_periodo = df_periodo[sem_data | com_data]
    if ts_fim:
        sem_data   = df_periodo["data_ref"].isna()
        com_data   = df_periodo["data_ref"].notna() & (
            df_periodo["data_ref"] <= ts_fim + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
        )
        df_periodo = df_periodo[sem_data | com_data]

    # Filtro de disciplina (se aplicado nos filtros da sidebar)
    disc_filtro = st.session_state.get("dash_disciplina", [])
    if disc_filtro:
        df_periodo = df_periodo[df_periodo["disciplina"].isin(disc_filtro)]
    if df_escopo.empty:
        st.info("Nenhum registro encontrado com os filtros selecionados.")
        return

    # ── Detecta modo ──
    # individual  → visao é str (nome de um único aluno)
    # comparativo → visao é list (múltiplos alunos OU "Todos os alunos")
    modo_comparativo = isinstance(visao, list)

    if modo_comparativo:
        alunos_no_escopo = sorted(df_escopo["aluno"].dropna().unique().tolist())
        # Restringe à lista de alunos que têm dados no escopo filtrado
        alunos_comp = [a for a in visao if a in alunos_no_escopo]
        if not alunos_comp:
            st.info("Nenhum dado encontrado para os alunos selecionados com os filtros aplicados.")
            return
        _dashboard_comparativo(df_escopo, df_periodo, alunos_comp, inicio_periodo, fim_periodo)
        return

    # ── Modo individual ──
    kpis       = calcular_kpis_avancados(df_escopo, df_periodo, inicio_periodo, fim_periodo)
    analisavel = kpis["analisavel"]

    st.caption(
        "📌 Passe o mouse sobre ? para ver fórmula e interpretação. "
        "Âmbar = últimos 15 dias · Verde = escopo completo do filtro. "
        "Contagens estruturais (total, restantes, progresso) ignoram filtro de período. "
        "Para ver todos os alunos em paralelo, limpe o filtro Alunos."
    )

    _resumo_15dias(df_escopo)
    st.markdown("---")

    _titulo_secao("Histórico completo",
        "Indicadores calculados sobre todo o período disponível nos filtros selecionados.",
        "Histórico completo")
    _render_kpis_produtividade(kpis)

    _titulo_secao("Avanço no plano",
        "Métricas de progresso em relação ao total de tarefas do plano de estudos.", "Histórico completo")
    _render_kpis_avanco(kpis)

    _titulo_secao("Desempenho acadêmico",
        "Indicadores de performance nas questões e tarefas concluídas.", "Histórico completo")
    _render_kpis_desempenho(kpis)

    st.markdown("---")

    abas = st.tabs([
        "📊 Visão geral", "📚 Disciplinas", "📅 Evolução",
        "⏱️ Gestão do tempo", "🏆 Rankings", "🧠 Análise IA", "📋 Atividades",
    ])

    with abas[0]:
        _titulo_secao("Distribuição por status",
            "Quantas tarefas estão em cada status. Permite visualizar a fila de trabalho e o progresso geral.")
        _aba_visao_geral(df_escopo, analisavel, temporal=kpis["temporal"])

    with abas[1]:
        _titulo_secao("Análise por disciplina",
            "Progresso usa o escopo total (sem filtro de período). "
            "Horas e desempenho respeitam o período selecionado.")
        _aba_disciplinas(analisavel, df_total=df_escopo, temporal=kpis["temporal"])

    with abas[2]:
        _titulo_secao("Evolução temporal",
            "Gráficos de horas e conclusões ao longo do tempo.")
        _aba_evolucao(analisavel, temporal=kpis["temporal"])

    with abas[3]:
        _titulo_secao("Gestão do tempo",
            "Como as horas de estudo estão distribuídas entre tipos de atividade e disciplinas.")
        _aba_gestao_tempo(analisavel, temporal=kpis["temporal"])

    with abas[4]:
        _titulo_secao("Rankings comparativos",
            "Classificação dos alunos por produtividade, consistência e desempenho.")
        _aba_ranking(analisavel, temporal=kpis["temporal"])

    with abas[5]:
        _titulo_secao("Análise inteligente por aluno",
            "Insights automáticos: padrões, riscos, disciplinas frágeis e recomendações.")
        _aba_analise_ia(df_escopo, df_periodo, visao, kpis, inicio_periodo, fim_periodo)

    with abas[6]:
        _titulo_secao("Todas as atividades",
            "Tabela detalhada com todos os registros de execução do período filtrado.")
        tabela  = preparar_tabela(df_escopo)
        colunas = ["aluno","tarefa","disciplina","aula","assunto","tipo","status_label",
                   "data_execucao","tempo","qtd_questoes_feitas","qtd_acertos","desempenho","comentario"]
        st.dataframe(tabela[[c for c in colunas if c in tabela.columns]],
            use_container_width=True, hide_index=True, row_height=72)


# ─────────────────────────────────────────────
# TELA: REGISTRO RÁPIDO
# ─────────────────────────────────────────────

def ultima_atividade(aluno_id=None):
    """Retorna o registro mais recente iniciado ou concluído do aluno."""
    filtro = ""
    params = []
    if aluno_id:
        filtro = "AND e.aluno_id = ?"
        params.append(int(aluno_id))
    df = consultar(
        f"""
        SELECT
            e.id AS execucao_id, al.nome AS aluno, t.numero AS tarefa, d.nome AS disciplina,
            COALESCE(ass.titulo, t.conteudo) AS assunto, COALESCE(t.tipo, 'Outro') AS tipo,
            e.status, e.data_execucao, e.ch_efetiva, e.comentario, e.atualizado_em,
            e.tarefa_id, e.aluno_id
        FROM execucoes e
        JOIN alunos al ON al.id = e.aluno_id AND al.ativo = 1
        JOIN tarefas t ON t.id = e.tarefa_id AND t.ativo = 1
        JOIN disciplinas d ON d.id = t.disciplina_id AND d.ativo = 1
        LEFT JOIN assuntos ass ON ass.id = t.assunto_id
        WHERE e.status IN ('EM_ANDAMENTO','CONCLUIDA') {filtro}
        ORDER BY e.atualizado_em DESC
        LIMIT 1
        """,
        tuple(params),
    )
    return None if df.empty else df.iloc[0].to_dict()


def exibir_card_ultima(atividade):
    if not atividade:
        st.info("Ainda não há atividade recente iniciada ou concluída.")
        return
    classe = status_card_class(atividade["status"])
    render_html(f"""
        <div class="quick-card {classe}">
          <div class="quick-label">Última atividade registrada</div>
          <h3>Tarefa {escape_html(atividade['tarefa'])} · {escape_html(atividade['disciplina'])}</h3>
          <div>{status_badge(atividade['status'])}</div>
          <div class="quick-grid">
            <div class="quick-item"><div class="quick-label">Aluno</div>
              <div class="quick-value">{escape_html(atividade['aluno'])}</div></div>
            <div class="quick-item"><div class="quick-label">Assunto</div>
              <div class="quick-value">{escape_html(atividade['assunto'] or '-')}</div></div>
            <div class="quick-item"><div class="quick-label">Tipo</div>
              <div class="quick-value">{escape_html(atividade['tipo'])}</div></div>
            <div class="quick-item"><div class="quick-label">Tempo</div>
              <div class="quick-value">{horas_para_hm(float(atividade["ch_efetiva"] or 0))}</div></div>
            <div class="quick-item"><div class="quick-label">Data</div>
              <div class="quick-value">{escape_html(formatar_data_br(atividade['data_execucao']))}</div></div>
            <div class="quick-item"><div class="quick-label">Atualização</div>
              <div class="quick-value">{escape_html(str(atividade['atualizado_em']))}</div></div>
            <div class="quick-item" style="grid-column: span 2;"><div class="quick-label">Observações</div>
              <div class="quick-value">{escape_html(atividade['comentario'] or '-')}</div></div>
          </div>
        </div>
    """)


def _label_tarefa(row) -> str:
    return f"Tarefa {int(row['tarefa'])} — {row['disciplina']} | {str(row['assunto'] or row['conteudo'] or '')[:60]}"


def _verificar_regra_andamento(tarefas_df: pd.DataFrame, tarefa_selecionada, novo_status: str, disciplina_id: int) -> tuple[bool, str]:
    """
    Regra: uma tarefa NAO_INICIADA só pode ser iniciada se não existir outra
    tarefa em andamento para a mesma disciplina.
    Retorna (pode_salvar, mensagem_de_erro).
    """
    status_atual = str(tarefa_selecionada.get("status", STATUS_NAO_INICIADA))
    if status_atual != STATUS_NAO_INICIADA:
        return True, ""
    if novo_status == STATUS_NAO_INICIADA:
        return True, ""

    # Há alguma outra tarefa em andamento na mesma disciplina?
    outras_em_andamento = tarefas_df[
        (tarefas_df["status"] == STATUS_EM_ANDAMENTO)
        & (tarefas_df["disciplina_id"] == disciplina_id)
        & (tarefas_df["tarefa_id"] != tarefa_selecionada["tarefa_id"])
    ]
    if outras_em_andamento.empty:
        return True, ""

    tarefa_bloqueante = outras_em_andamento.iloc[0]
    msg = (
        f"❌ Não é possível iniciar esta tarefa. "
        f"A Tarefa {int(tarefa_bloqueante['tarefa'])} ({tarefa_bloqueante['disciplina']}) "
        f"está em andamento. Conclua ou atualize essa tarefa antes de iniciar outra na mesma disciplina."
    )
    return False, msg


def _painel_tarefa(tarefa) -> None:
    """
    Exibe o conteúdo completo da tarefa selecionada antes do formulário:
    disciplina, aula, assunto, descrição e exercícios previstos.
    Texto nunca é cortado.
    """
    disciplina = escape_html(str(tarefa.get("disciplina") or ""))
    aula       = escape_html(str(tarefa.get("aula") or ""))
    assunto    = escape_html(str(tarefa.get("assunto") or ""))
    conteudo   = str(tarefa.get("conteudo") or "").strip()
    previstos  = int(tarefa.get("qtd_exercicios_previstos") or 0)
    tipo       = escape_html(str(tarefa.get("tipo") or ""))
    num_tarefa = int(tarefa.get("tarefa", 0))

    # Cabeçalho da tarefa
    render_html(f"""
    <div style="
        background:#f8fafc;
        border:1px solid #e2e8f0;
        border-left:4px solid #3b82f6;
        border-radius:10px;
        padding:16px 20px;
        margin-bottom:14px;
    ">
      <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;margin-bottom:10px">
        <div>
          <div style="font-size:.68rem;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em">
            Tarefa {num_tarefa}
          </div>
          <div style="font-size:1.05rem;font-weight:800;color:#0f172a;margin-top:2px">{disciplina}</div>
        </div>
        <div style="display:flex;gap:8px;flex-wrap:wrap">
          <span style="background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;
            border-radius:999px;padding:3px 10px;font-size:.73rem;font-weight:700">{tipo}</span>
          {f'<span style="background:#f0fdf4;color:#166534;border:1px solid #bbf7d0;border-radius:999px;padding:3px 10px;font-size:.73rem;font-weight:700">📝 {previstos} exercícios previstos</span>' if previstos else ''}
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:{'12px' if conteudo else '0'}">
        <div>
          <div style="font-size:.66rem;font-weight:800;color:#94a3b8;text-transform:uppercase;margin-bottom:3px">Aula</div>
          <div style="font-size:.87rem;color:#0f172a;font-weight:500;line-height:1.45">{aula or '—'}</div>
        </div>
        <div>
          <div style="font-size:.66rem;font-weight:800;color:#94a3b8;text-transform:uppercase;margin-bottom:3px">Assunto</div>
          <div style="font-size:.87rem;color:#0f172a;font-weight:500;line-height:1.45">{assunto or '—'}</div>
        </div>
      </div>
      {f'''<div style="border-top:1px solid #e2e8f0;padding-top:10px;margin-top:4px">
        <div style="font-size:.66rem;font-weight:800;color:#94a3b8;text-transform:uppercase;margin-bottom:6px">
          Descrição / Objetivos do bloco
        </div>
        <div style="font-size:.84rem;color:#334155;line-height:1.65;white-space:pre-wrap">{escape_html(conteudo)}</div>
      </div>''' if conteudo else ''}
    </div>
    """)


def _toast_sucesso(msg: str) -> None:
    """
    Exibe mensagem de sucesso de forma garantida em todas as páginas,
    mesmo após st.rerun().

    Estratégia:
    1. Tenta st.toast (nativo, Streamlit ≥ 1.28) — aparece como notificação flutuante.
    2. Também salva em session_state para o banner ser renderizado no próximo ciclo
       via _render_banner_sucesso(), chamada no início de main().
    """
    # Guarda no session_state para persistir após rerun
    st.session_state["_sucesso_msg"] = msg
    # Tenta exibir imediatamente via toast
    try:
        st.toast(msg, icon="✅")
    except (AttributeError, Exception):
        pass   # será exibido via banner no próximo ciclo


def _render_banner_sucesso() -> None:
    """
    Renderiza o banner de sucesso persistente (salvo em session_state).
    Chamada UMA VEZ no início de main(). CSS responsivo, sem overflow.
    """
    msg = st.session_state.pop("_sucesso_msg", None)
    if not msg:
        return
    render_html(f"""
    <style>
    @keyframes rr_fadeIn {{
      from {{ opacity:0; transform:translateY(-8px); }}
      to   {{ opacity:1; transform:translateY(0); }}
    }}
    .rr-banner {{
        display:flex; align-items:flex-start; gap:12px;
        background:#f0fdf4; border:1px solid #86efac;
        border-left:5px solid #16a34a; border-radius:10px;
        padding:14px 18px; margin-bottom:16px;
        box-sizing:border-box; width:100%; max-width:100%;
        overflow:visible; animation:rr_fadeIn .35s ease;
        position:relative; z-index:999;
    }}
    .rr-banner-icon {{ font-size:1.3rem; flex-shrink:0; padding-top:1px; }}
    .rr-banner-text {{
        font-size:.9rem; font-weight:700; color:#166534;
        line-height:1.5; word-break:break-word;
        overflow-wrap:anywhere; flex:1; min-width:0;
    }}
    @media(max-width:600px){{
        .rr-banner {{ padding:10px 12px; gap:8px; }}
        .rr-banner-text {{ font-size:.82rem; }}
    }}
    </style>
    <div class="rr-banner">
      <span class="rr-banner-icon">&#x2705;</span>
      <span class="rr-banner-text">{escape_html(msg)}</span>
    </div>
    """)


# ─────────────────────────────────────────────────────────────────
# REGISTRO RÁPIDO — helpers de seleção persistente
# ─────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────
# REGISTRO RÁPIDO — estado e helpers
# ─────────────────────────────────────────────────────────────────

# Keys de ARMAZENAMENTO no session_state (não usadas como key de widget)
_RR_STORE = {
    STATUS_NAO_INICIADA: "rr_store_nao",
    STATUS_EM_ANDAMENTO: "rr_store_and",
    STATUS_CONCLUIDA:    "rr_store_con",
}
# Keys de WIDGET do selectbox (diferentes das de armazenamento)
_RR_KEY = {
    STATUS_NAO_INICIADA: "rr_wgt_nao",
    STATUS_EM_ANDAMENTO: "rr_wgt_and",
    STATUS_CONCLUIDA:    "rr_wgt_con",
}
_RR_ABA_KEY = "rr_aba_ativa"


def _rr_init(aluno_id: int) -> None:
    """Inicializa as chaves de session_state do Registro Rápido."""
    if st.session_state.get("rr_aluno_anterior") != aluno_id:
        for k in list(_RR_STORE.values()) + list(_RR_KEY.values()):
            st.session_state.pop(k, None)
        st.session_state["rr_aluno_anterior"] = aluno_id


def _rr_get_tarefa_id(status: str) -> int | None:
    """Retorna o tarefa_id salvo para o status (lê da chave de armazenamento)."""
    return st.session_state.get(_RR_STORE[status])


def _rr_set_tarefa_id(status: str, tarefa_id: int | None) -> None:
    """Salva o tarefa_id na chave de ARMAZENAMENTO (nunca na key do widget)."""
    store_key = _RR_STORE[status]
    if tarefa_id is None:
        st.session_state.pop(store_key, None)
    else:
        st.session_state[store_key] = int(tarefa_id)


def _rr_limpar_outros(status_ativo: str) -> None:
    """Limpa seleções das outras duas abas."""
    for s in _RR_STORE:
        if s != status_ativo:
            st.session_state.pop(_RR_STORE[s], None)


def _rr_aba_ativa() -> str:
    return st.session_state.get(_RR_ABA_KEY, STATUS_NAO_INICIADA)


def _rr_set_aba(status: str) -> None:
    st.session_state[_RR_ABA_KEY] = status


def _rr_selectbox(grupo: pd.DataFrame, status: str, label: str) -> pd.Series | None:
    """
    Selectbox com persistência de seleção via session_state.

    Separação obrigatória:
    - _RR_STORE[status] → chave de ARMAZENAMENTO (lida/escrita pelo nosso código)
    - _RR_KEY[status]   → key do WIDGET (gerenciada exclusivamente pelo Streamlit)

    Nunca escrevemos em session_state[widget_key] após o widget ser instanciado.
    """
    if grupo.empty:
        return None

    # Garante ordem numérica do plano, independente da ordem de chegada do DataFrame
    grupo = grupo.sort_values("tarefa").reset_index(drop=True)
    ids      = grupo["tarefa_id"].tolist()
    id_salvo = _rr_get_tarefa_id(status)

    # Calcula índice inicial pela chave de armazenamento
    if id_salvo in ids:
        idx = ids.index(id_salvo)
    else:
        idx = 0

    # Widget: key é exclusiva desta aba, gerenciada pelo Streamlit
    sel_id = st.selectbox(
        label,
        ids,
        index=idx,
        format_func=lambda v: _label_tarefa(grupo[grupo["tarefa_id"] == v].iloc[0]),
        key=_RR_KEY[status],
    )

    # Persiste na chave de ARMAZENAMENTO (separada da key do widget)
    _rr_set_tarefa_id(status, sel_id)

    matches = grupo[grupo["tarefa_id"] == sel_id]
    return None if matches.empty else matches.iloc[0]


def _rr_formulario(
    tarefas_df: pd.DataFrame,
    tarefa: pd.Series,
    aluno_id: int,
    status_alvo: str,
) -> None:
    """
    Formulário unificado: status + sessão em um único submit.
    Se horas e questões forem zero, apenas o status é atualizado.
    """
    tarefa_id    = int(tarefa["tarefa_id"])
    eh_concluida = (status_alvo == STATUS_CONCLUIDA)

    _painel_tarefa(tarefa)

    if eh_concluida:
        render_html(
            '<div class="rule-warning" style="margin-bottom:14px">'
            '⚠️ <strong>Atenção:</strong> Esta tarefa já está <strong>Concluída</strong>. '
            'Você ainda pode registrar novas sessões ou alterar o status.'
            '</div>'
        )

    status_atual = str(tarefa.get("status", STATUS_NAO_INICIADA))
    tipo_atual   = str(tarefa.get("tipo") or "Outro")
    tipo_idx     = TIPOS_ESTUDO.index(tipo_atual) if tipo_atual in TIPOS_ESTUDO else TIPOS_ESTUDO.index("Outro")

    render_html(
        '<div class="section-title">✏️ Atualizar status e registrar sessão</div>'
        '<div class="insight-card info" style="margin-bottom:10px">'
        '<div class="insight-icon">ℹ️</div>'
        '<div class="insight-body">'
        '<p class="insight-text" style="margin:0">'
        'Atualize o status e, opcionalmente, registre as horas desta sessão. '
        'Se não houver horas nem questões, deixe os campos zerados — '
        'apenas o status será atualizado.'
        '</p></div></div>'
    )

    with st.form(f"rr_form_{status_alvo}_{tarefa_id}"):
        col1, col2 = st.columns(2)
        novo_status = col1.selectbox(
            "Status da tarefa", STATUS_VALIDOS,
            index=STATUS_VALIDOS.index(status_atual),
            format_func=lambda v: STATUS_LABELS[v],
        )
        tipo_estudo = col2.selectbox("Tipo de estudo", TIPOS_ESTUDO, index=tipo_idx)

        st.markdown("---")
        render_html('<div style="font-size:.8rem;font-weight:700;color:#475569;margin-bottom:4px">⏱️ Sessão de estudo (opcional)</div>')

        col1, col2, col3 = st.columns(3)
        data_sessao = col1.date_input("Data da sessão", value=date.today())
        ch_h = col2.number_input("Horas", min_value=0, value=0, step=1)
        ch_m = col3.number_input("Minutos", min_value=0, max_value=59, value=0, step=5)

        col4, col5 = st.columns(2)
        questoes = col4.number_input("Questões desta sessão", min_value=0, value=0, step=1)
        acertos  = col5.number_input("Acertos desta sessão",  min_value=0, value=0, step=1)

        comentario = st.text_area(
            "Observações desta sessão",
            placeholder="O que foi estudado, dificuldades, conteúdos… (opcional)",
            height=80,
        )

        if eh_concluida:
            confirmado = st.checkbox(
                "✅ Confirmo que desejo alterar este registro já concluído", value=False
            )
        else:
            confirmado = True

        col_b1, col_b2 = st.columns([3, 1])
        salvar   = col_b1.form_submit_button("💾 Salvar", type="primary", use_container_width=True)
        cancelar = col_b2.form_submit_button("✖ Cancelar", use_container_width=True)

    if cancelar:
        st.info("Nenhuma alteração foi salva.")
        _rr_historico(aluno_id, tarefa_id)
        return

    if not salvar:
        _rr_historico(aluno_id, tarefa_id)
        return

    # Validações
    ch = hm_para_horas(ch_h, ch_m)
    erros: list[str] = []
    if eh_concluida and not confirmado:
        erros.append("Para alterar uma tarefa já concluída, marque a confirmação acima.")
    if acertos > questoes:
        erros.append("O número de acertos não pode ser maior que o número de questões.")
    pode, msg_bloqueio = _verificar_regra_andamento(
        tarefas_df, tarefa, novo_status, int(tarefa["disciplina_id"]))
    if not pode:
        erros.append(msg_bloqueio)

    if erros:
        for e in erros:
            render_html(f'<div class="rule-error">❌ {escape_html(e)}</div>')
        _rr_historico(aluno_id, tarefa_id)
        return

    # Gravação
    try:
        with st.spinner("Salvando…"):
            with conectar() as conn:
                # Upsert, não UPDATE: se a tarefa ainda não tem vínculo com o
                # aluno (planilha importada sem "Vincular tarefas pendentes"),
                # um UPDATE afetaria zero linhas e o registro se perderia.
                conn.execute(
                    """INSERT INTO execucoes
                           (aluno_id, tarefa_id, status, tipo_estudo, concluida, atualizado_em)
                       VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                       ON CONFLICT(aluno_id, tarefa_id) DO UPDATE SET
                           status        = excluded.status,
                           tipo_estudo   = excluded.tipo_estudo,
                           concluida     = excluded.concluida,
                           atualizado_em = CURRENT_TIMESTAMP""",
                    (aluno_id, tarefa_id, novo_status, tipo_estudo,
                     1 if novo_status == STATUS_CONCLUIDA else 0),
                )
                tem_sessao = ch > 0 or questoes > 0
                if tem_sessao:
                    conn.execute(
                        """INSERT INTO sessoes_estudo
                           (aluno_id,tarefa_id,data_sessao,ch_sessao,
                            qtd_questoes,qtd_acertos,tipo_estudo,comentario)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (aluno_id, tarefa_id, str(data_sessao),
                         float(ch), int(questoes), int(acertos),
                         normalizar_tipo_estudo(tipo_estudo),
                         limpar_texto(comentario)),
                    )
                    _recalcular_execucao_por_sessoes(conn, aluno_id, tarefa_id)

        limpar_cache()

        status_ant = STATUS_LABELS.get(status_atual, status_atual)
        status_nov = STATUS_LABELS.get(novo_status, novo_status)
        partes = []
        if status_atual != novo_status:
            partes.append(f"Status: {status_ant} → {status_nov}.")
        else:
            partes.append("Status mantido.")
        if tem_sessao:
            partes.append(f"Sessão: {horas_para_hm(ch)}.")
            if questoes > 0:
                partes.append(f"{acertos}/{questoes} acertos.")
        else:
            partes.append("Nenhuma sessão registrada.")

        if status_atual != novo_status:
            _rr_set_tarefa_id(status_alvo, None)

        _toast_sucesso(" ".join(partes))
        st.rerun()

    except Exception as exc:
        erro_usuario("❌ Não foi possível salvar.", exc)

    _rr_historico(aluno_id, tarefa_id)


def _rr_historico(aluno_id: int, tarefa_id: int) -> None:
    """Histórico de sessões com totais e exclusão."""
    st.markdown("---")
    render_html('<div class="section-title">📅 Histórico de sessões</div>')
    sessoes = carregar_sessoes(aluno_id, tarefa_id)

    if sessoes.empty:
        st.info("Nenhuma sessão registrada para esta tarefa.")
        return

    total_h = float(sessoes["ch_sessao"].sum())
    total_q = int(sessoes["qtd_questoes"].sum())
    total_a = int(sessoes["qtd_acertos"].sum())
    des_tot = round(total_a / total_q * 100, 1) if total_q else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total de horas", horas_para_hm(total_h))
    c2.metric("Total de questões", str(total_q))
    c3.metric("Total de acertos", str(total_a))
    c4.metric("Desempenho geral", f"{des_tot}%")

    st.dataframe(
        sessoes.rename(columns={
            "data_sessao":"Data","ch_sessao":"Tempo (h)",
            "qtd_questoes":"Questões","qtd_acertos":"Acertos",
            "desempenho_sessao":"Desempenho (%)","tipo_estudo":"Tipo","comentario":"Observações",
        })[["Data","Tempo (h)","Questões","Acertos","Desempenho (%)","Tipo","Observações"]],
        use_container_width=True, hide_index=True,
    )

    with st.expander("🗑️ Excluir uma sessão"):
        render_html(
            '<div class="rule-warning" style="margin-bottom:8px">'
            '⚠️ A exclusão reduz os totais acumulados da tarefa. Ação irreversível.'
            '</div>'
        )
        ids_sessoes = sessoes["id"].tolist()
        sel_id = st.selectbox(
            "Selecione a sessão", ids_sessoes,
            format_func=lambda v: (
                f"{sessoes.loc[sessoes['id']==v,'data_sessao'].iloc[0]} — "
                f"{horas_para_hm(float(sessoes.loc[sessoes['id']==v,'ch_sessao'].iloc[0]))} — "
                f"{int(sessoes.loc[sessoes['id']==v,'qtd_questoes'].iloc[0])} questões"
            ),
            key=f"del_sessao_{tarefa_id}",
        )
        conf_del = st.checkbox("Confirmo a exclusão desta sessão", key=f"conf_del_{tarefa_id}")
        if st.button("🗑️ Excluir sessão", disabled=not conf_del, key=f"btn_del_{tarefa_id}"):
            try:
                excluir_sessao(sel_id, aluno_id, tarefa_id)
                _toast_sucesso("Sessão excluída. Totais recalculados.")
                st.rerun()
            except Exception as exc:
                erro_usuario("❌ Não foi possível excluir a sessão.", exc)

def tela_registro_rapido():
    render_html(
        '<div class="hero">'
        '<h1>⚡ Registro Rápido</h1>'
        '<p>Selecione uma tarefa para ver seu conteúdo completo e registrar o progresso. '
        'O formulário é pré-preenchido com os dados do último registro salvo.</p>'
        '</div>'
    )

    usuario = aluno_logado()
    alunos  = alunos_ativos()

    # ── Seleção do aluno ──
    if usuario["perfil"] == "Aluno":
        aluno_id = int(usuario["id"])
        render_html(
            f'<div style="display:inline-flex;align-items:center;gap:8px;'
            f'background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;'
            f'padding:8px 14px;margin-bottom:14px">'
            f'<span style="font-size:1.1rem">👤</span>'
            f'<div>'
            f'<div style="font-size:.66rem;font-weight:800;color:#94a3b8;text-transform:uppercase">Aluno</div>'
            f'<div style="font-size:.9rem;font-weight:700;color:#0f172a">{escape_html(usuario["nome"])}</div>'
            f'</div></div>'
        )
    else:
        if alunos.empty:
            st.info("Nenhum aluno cadastrado. Vá em **Alunos** para cadastrar.")
            return
        aluno_id = st.selectbox(
            "Aluno",
            alunos["id"].tolist(),
            format_func=lambda v: alunos.loc[alunos["id"] == v, "nome"].iloc[0],
            key="rr_aluno_sel",
        )

    # Inicializa estado por aluno
    _rr_init(aluno_id)

    # Última atividade
    exibir_card_ultima(ultima_atividade(aluno_id))

    # Carrega tarefas do aluno
    tarefas = carregar_visao_tarefas(aluno_id)
    if tarefas.empty:
        st.info("Nenhuma tarefa vinculada a este aluno. Vá em **Tarefas** para vincular.")
        return

    grupo_nao  = tarefas[tarefas["status"] == STATUS_NAO_INICIADA].sort_values("tarefa").reset_index(drop=True)
    grupo_and  = tarefas[tarefas["status"] == STATUS_EM_ANDAMENTO].sort_values("tarefa").reset_index(drop=True)
    grupo_conc = tarefas[tarefas["status"] == STATUS_CONCLUIDA].sort_values("tarefa").reset_index(drop=True)

    qtd_nao  = len(grupo_nao)
    qtd_and  = len(grupo_and)
    qtd_conc = len(grupo_conc)

    # ── Seletor de aba — controla qual painel exibir ──
    # Usa radio horizontal para evitar o problema de st.tabs (todas renderizadas juntas)
    aba_opcoes = {
        STATUS_NAO_INICIADA: f"🔘 Não iniciadas ({qtd_nao})",
        STATUS_EM_ANDAMENTO: f"🟡 Em andamento ({qtd_and})",
        STATUS_CONCLUIDA:    f"🟢 Concluídas ({qtd_conc})",
    }
    aba_atual = _rr_aba_ativa()
    # Garante que a aba ativa é válida
    if aba_atual not in aba_opcoes:
        aba_atual = STATUS_NAO_INICIADA

    aba_escolhida = st.radio(
        "Filtrar por status",
        list(aba_opcoes.keys()),
        index=list(aba_opcoes.keys()).index(aba_atual),
        format_func=lambda v: aba_opcoes[v],
        horizontal=True,
        key="rr_aba_radio",
    )

    # Se o usuário trocou de aba: salva a nova aba ativa e limpa seleções das outras
    if aba_escolhida != aba_atual:
        _rr_set_aba(aba_escolhida)
        _rr_limpar_outros(aba_escolhida)
        aba_atual = aba_escolhida

    st.divider()

    # ── Renderiza o painel da aba ativa ──
    if aba_atual == STATUS_NAO_INICIADA:
        render_html(
            '<div class="insight-card info" style="margin-bottom:14px">'
            '<div class="insight-icon">ℹ️</div>'
            '<div class="insight-body">'
            '<div class="insight-title">Regra: tarefas não iniciadas</div>'
            '<p class="insight-text">Uma tarefa não iniciada só pode ser iniciada se '
            '<strong>não houver outra tarefa em andamento na mesma disciplina</strong>. '
            'Conclua ou atualize a tarefa em andamento primeiro.</p>'
            '</div></div>'
        )
        if grupo_nao.empty:
            st.info("Nenhuma tarefa não iniciada.")
        else:
            tarefa = _rr_selectbox(grupo_nao, STATUS_NAO_INICIADA, "Tarefa não iniciada")
            if tarefa is not None:
                _rr_formulario(tarefas, tarefa, aluno_id, STATUS_NAO_INICIADA)

    elif aba_atual == STATUS_EM_ANDAMENTO:
        if grupo_and.empty:
            st.info("Nenhuma tarefa em andamento.")
        else:
            tarefa = _rr_selectbox(grupo_and, STATUS_EM_ANDAMENTO, "Tarefa em andamento")
            if tarefa is not None:
                _rr_formulario(tarefas, tarefa, aluno_id, STATUS_EM_ANDAMENTO)

    else:  # CONCLUIDA
        render_html(
            '<div class="insight-card warning" style="margin-bottom:14px">'
            '<div class="insight-icon">⚠️</div>'
            '<div class="insight-body">'
            '<div class="insight-title">Atenção: tarefas concluídas</div>'
            '<p class="insight-text">Alterações em tarefas concluídas modificam o histórico de estudos. '
            'O sistema exigirá confirmação explícita antes de salvar.</p>'
            '</div></div>'
        )
        if grupo_conc.empty:
            st.info("Nenhuma tarefa concluída.")
        else:
            tarefa = _rr_selectbox(grupo_conc, STATUS_CONCLUIDA, "Tarefa concluída")
            if tarefa is not None:
                _rr_formulario(tarefas, tarefa, aluno_id, STATUS_CONCLUIDA)

    # ── Histórico recente ──
    # Histórico recente: últimas sessões reais do aluno (uma linha por sessão)
    recentes = carregar_sessoes_aluno(aluno_id).head(10)
    if not recentes.empty:
        st.markdown("---")
        render_html('<div class="section-title">📋 Histórico recente de sessões</div>')
        # carregar_sessoes_aluno retorna: data_sessao, ch_sessao, qtd_questoes, qtd_acertos,
        # tipo_estudo, comentario, tarefa, disciplina, assunto
        recentes["tempo"] = recentes["ch_sessao"].apply(horas_para_hm)
        recentes["desempenho"] = recentes.apply(
            lambda r: f"{round(r['qtd_acertos']/r['qtd_questoes']*100,1)}%"
            if r["qtd_questoes"] > 0 else "—", axis=1
        )
        st.dataframe(
            recentes.rename(columns={
                "data_sessao": "Data", "tarefa": "Tarefa",
                "disciplina": "Disciplina", "assunto": "Assunto",
                "tipo_estudo": "Tipo", "tempo": "Tempo",
                "qtd_questoes": "Questões", "qtd_acertos": "Acertos",
                "desempenho": "Desempenho", "comentario": "Observações",
            })[[
                "Data","Tarefa","Disciplina","Assunto","Tipo",
                "Tempo","Questões","Acertos","Desempenho","Observações",
            ]],
            use_container_width=True, hide_index=True, row_height=64,
        )

# ─────────────────────────────────────────────

def tela_tarefas():
    st.header("Gestão de tarefas educacionais")
    usuario    = aluno_logado()
    alunos     = alunos_ativos()
    disciplinas = disciplinas_ativas()
    aulas      = carregar_aulas()
    assuntos   = carregar_assuntos()
    tarefas    = carregar_tarefas_base()

    abas = st.tabs(["Execução e status", "Criar tarefa", "Editar tarefas", "Vincular alunos"])

    # ── Aba 1: Execução e status ──
    with abas[0]:
        if usuario["perfil"] == "Aluno":
            aluno_id = int(usuario["id"])
            st.text_input("Aluno", value=usuario["nome"], disabled=True)
        else:
            if alunos.empty:
                st.info("Cadastre um aluno.")
                return
            aluno_id = st.selectbox(
                "Aluno", alunos["id"].tolist(),
                format_func=lambda v: alunos.loc[alunos["id"] == v, "nome"].iloc[0],
            )
        df = carregar_visao_tarefas(aluno_id)
        col1, col2, col3 = st.columns(3)
        disciplina_f = col1.multiselect("Disciplina", sorted(df["disciplina"].dropna().unique().tolist()))
        tipo_f       = col2.multiselect("Tipo de estudo", sorted(df["tipo"].dropna().unique().tolist()))
        status_f     = col3.multiselect("Status", STATUS_VALIDOS, default=STATUS_VALIDOS, format_func=lambda v: STATUS_LABELS[v])
        filtrado = df.copy()
        if disciplina_f: filtrado = filtrado[filtrado["disciplina"].isin(disciplina_f)]
        if tipo_f:       filtrado = filtrado[filtrado["tipo"].isin(tipo_f)]
        if status_f:     filtrado = filtrado[filtrado["status"].isin(status_f)]

        colunas_editor = ["tarefa_id", "tarefa", "disciplina", "aula", "assunto", "tipo", "status",
                          "data_execucao", "ch_horas", "ch_minutos", "qtd_questoes_feitas", "qtd_acertos", "desempenho", "comentario"]
        editor = filtrado.copy()
        editor["data_execucao"] = editor["data_execucao"].fillna("")
        # Desmembra ch_efetiva em horas e minutos para edição amigável
        editor["ch_horas"]   = editor["ch_efetiva"].apply(lambda v: int(float(v or 0)))
        editor["ch_minutos"] = editor["ch_efetiva"].apply(lambda v: int(round((float(v or 0) % 1) * 60)))
        editor = editor[[c for c in colunas_editor if c in editor.columns]]
        editado = st.data_editor(
            editor,
            use_container_width=True,
            hide_index=True,
            row_height=72,
            disabled=["tarefa_id", "tarefa", "disciplina", "aula", "assunto", "desempenho"],
            column_config={
                "tarefa_id": None,
                "tarefa": "Tarefa",
                "tipo": st.column_config.SelectboxColumn("Tipo de estudo", options=TIPOS_ESTUDO),
                "status": st.column_config.SelectboxColumn("Status", options=STATUS_VALIDOS, required=True),
                "data_execucao": st.column_config.TextColumn("Data (aaaa-mm-dd)"),
                "ch_horas":   st.column_config.NumberColumn("Horas", min_value=0, step=1),
                "ch_minutos": st.column_config.NumberColumn("Minutos", min_value=0, max_value=59, step=5),
                "qtd_questoes_feitas": st.column_config.NumberColumn("Questões feitas", min_value=0, step=1),
                "qtd_acertos": st.column_config.NumberColumn("Acertos", min_value=0, step=1),
                "comentario": st.column_config.TextColumn("Observações", width="large"),
            },
            key="editor_execucoes",
        )
        if st.button("Salvar alterações de execução", type="primary"):
            try:
                with st.spinner("Salvando…"):
                 with conectar() as conn:
                    # Tarefas cujas horas são derivadas de sessões registradas.
                    # Editá-las aqui criaria divergência: o painel soma as sessões,
                    # a grade mostraria outro número, e a próxima sessão registrada
                    # sobrescreveria a edição sem avisar.
                    com_sessao = {
                        (int(t), float(h or 0), int(q or 0), int(a or 0))
                        for t, h, q, a in conn.execute(
                            """SELECT tarefa_id, COALESCE(SUM(ch_sessao),0),
                                      COALESCE(SUM(qtd_questoes),0), COALESCE(SUM(qtd_acertos),0)
                               FROM sessoes_estudo WHERE aluno_id = ?
                               GROUP BY tarefa_id""",
                            (aluno_id,),
                        ).fetchall()
                    }
                    derivadas = {t: (h, q, a) for t, h, q, a in com_sessao}

                    for _, row in editado.iterrows():
                        if converter_inteiro(row["qtd_acertos"]) > converter_inteiro(row["qtd_questoes_feitas"]):
                            raise ValueError(f"Tarefa {row['tarefa']}: acertos maiores que questões.")
                        ch_float = hm_para_horas(int(row.get("ch_horas", 0)), int(row.get("ch_minutos", 0)))

                        tid = int(row["tarefa_id"])
                        if tid in derivadas:
                            h_s, q_s, a_s = derivadas[tid]
                            q_e = converter_inteiro(row["qtd_questoes_feitas"])
                            a_e = converter_inteiro(row["qtd_acertos"])
                            if abs(ch_float - h_s) > 0.005 or q_e != q_s or a_e != a_s:
                                raise ValueError(
                                    f"Tarefa {row['tarefa']}: horas, questões e acertos vêm das sessões "
                                    f"registradas ({horas_para_hm(h_s)}, {q_s} questões, {a_s} acertos). "
                                    "Edite a sessão em Registro rápido; aqui altere só status, tipo, data ou observações."
                                )
                        upsert_execucao(
                            conn, aluno_id, int(row["tarefa_id"]),
                            converter_data(row["data_execucao"]),
                            ch_float, None, 0,
                            converter_inteiro(row["qtd_acertos"]),
                            limpar_texto(row["comentario"]),
                            converter_inteiro(row["qtd_questoes_feitas"]),
                            row["status"], row["tipo"],
                        )
                limpar_cache()
                _toast_sucesso("Registros de execução salvos com sucesso.")
                st.rerun()
            except Exception as exc:
                erro_usuario("Não foi possível salvar.", exc)

    # ── Aba 2: Criar tarefa ──
    with abas[1]:
        if usuario["perfil"] != "Gestor":
            st.info("Somente gestores podem criar tarefas.")
        elif disciplinas.empty:
            st.info("Cadastre disciplinas/aulas primeiro.")
        else:
            with st.form("nova_tarefa", clear_on_submit=True):
                col1, col2, col3 = st.columns(3)
                numero       = col1.number_input("Número da tarefa", min_value=1, step=1)
                trilha       = col2.number_input("Trilha", min_value=0, step=1)
                disciplina_id = col3.selectbox("Disciplina", disciplinas["id"].tolist(), format_func=lambda v: disciplinas.loc[disciplinas["id"] == v, "nome"].iloc[0])
                aulas_disc   = aulas[aulas["disciplina_id"] == disciplina_id]
                aula_id      = st.selectbox("Aula", aulas_disc["id"].tolist(), format_func=lambda v: aulas_disc.loc[aulas_disc["id"] == v, "aula"].iloc[0]) if not aulas_disc.empty else None
                assuntos_aula = assuntos[assuntos["aula_id"] == aula_id] if aula_id else assuntos.iloc[0:0]
                assunto_id   = st.selectbox("Assunto", assuntos_aula["id"].tolist(), format_func=lambda v: assuntos_aula.loc[assuntos_aula["id"] == v, "assunto"].iloc[0]) if not assuntos_aula.empty else None
                col4, col5   = st.columns(2)
                tipo         = col4.selectbox("Tipo de estudo", TIPOS_ESTUDO)
                previstos    = col5.number_input("Exercícios previstos", min_value=0, step=1)
                conteudo     = st.text_area("Conteúdo")
                alunos_vinc  = st.multiselect("Vincular alunos", alunos["id"].tolist(), format_func=lambda v: alunos.loc[alunos["id"] == v, "nome"].iloc[0])
                salvar       = st.form_submit_button("Criar tarefa", use_container_width=True)
            if salvar:
                try:
                    aula_nome = aulas.loc[aulas["id"] == aula_id, "aula"].iloc[0] if aula_id else "Aula não informada"
                    with conectar() as conn:
                        conn.execute(
                            "INSERT INTO tarefas (numero, trilha, disciplina_id, aula_id, assunto_id, aula, qtd_exercicios_previstos, tipo, conteudo, ativo) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)",
                            (int(numero), int(trilha), int(disciplina_id), aula_id, assunto_id, aula_nome, int(previstos), tipo, limpar_texto(conteudo)),
                        )
                        tarefa_id = ultimo_id(conn)
                        for av in alunos_vinc:
                            upsert_execucao(conn, int(av), tarefa_id, None, 0, None, 0, 0, None, 0, STATUS_NAO_INICIADA)
                    limpar_cache()
                    _toast_sucesso("Tarefa criada com sucesso.")
                    st.rerun()
                except (IntegrityError, UniqueViolation):
                    st.error("Já existe uma tarefa com esse número.")

    # ── Aba 3: Editar tarefas ──
    with abas[2]:
        if usuario["perfil"] != "Gestor":
            st.info("Somente gestores podem editar tarefas.")
        elif tarefas.empty:
            st.info("Nenhuma tarefa cadastrada.")
        else:
            colunas_edit = ["tarefa_id", "tarefa", "trilha", "disciplina", "aula", "assunto", "tipo", "qtd_exercicios_previstos", "conteudo"]
            edit    = tarefas[[c for c in colunas_edit if c in tarefas.columns]].copy()
            editado = st.data_editor(
                edit, use_container_width=True, hide_index=True, row_height=72,
                disabled=["tarefa_id", "disciplina", "aula", "assunto"],
                column_config={
                    "tarefa_id": None,
                    "tipo": st.column_config.SelectboxColumn("Tipo", options=TIPOS_ESTUDO),
                    "conteudo": st.column_config.TextColumn("Conteúdo", width="large"),
                    "qtd_exercicios_previstos": st.column_config.NumberColumn("Exercícios previstos", min_value=0, step=1),
                },
                key="editor_tarefas_crud",
            )
            col1, col2 = st.columns(2)
            if col1.button("Salvar tarefas", type="primary"):
                with conectar() as conn:
                    for _, row in editado.iterrows():
                        conn.execute(
                            "UPDATE tarefas SET numero = ?, trilha = ?, tipo = ?, qtd_exercicios_previstos = ?, conteudo = ? WHERE id = ?",
                            (int(row["tarefa"]), converter_inteiro(row["trilha"]), row["tipo"], converter_inteiro(row["qtd_exercicios_previstos"]), limpar_texto(row["conteudo"]), int(row["tarefa_id"])),
                        )
                limpar_cache()
                _toast_sucesso("Alterações nas tarefas salvas com sucesso.")
                st.rerun()
            excluir = col2.selectbox("Excluir tarefa", tarefas["tarefa_id"].tolist(), format_func=lambda v: f"Tarefa {int(tarefas.loc[tarefas['tarefa_id'] == v, 'tarefa'].iloc[0])}")
            if col2.button("Excluir tarefa selecionada"):
                executar("UPDATE tarefas SET ativo = 0 WHERE id = ?", (int(excluir),))
                limpar_cache()
                _toast_sucesso("Tarefa excluída com sucesso.")
                st.rerun()

    # ── Aba 4: Vincular alunos ──
    with abas[3]:
        if usuario["perfil"] != "Gestor":
            st.info("Somente gestores podem vincular alunos.")
        elif tarefas.empty or alunos.empty:
            st.info("Cadastre alunos e tarefas.")
        else:
            tarefa_id = st.selectbox(
                "Tarefa", tarefas["tarefa_id"].tolist(),
                format_func=lambda v: f"Tarefa {int(tarefas.loc[tarefas['tarefa_id'] == v, 'tarefa'].iloc[0])} - {tarefas.loc[tarefas['tarefa_id'] == v, 'disciplina'].iloc[0]}",
            )
            vincular = st.multiselect("Alunos", alunos["id"].tolist(), format_func=lambda v: alunos.loc[alunos["id"] == v, "nome"].iloc[0])
            if st.button("Vincular alunos à tarefa", type="primary"):
                with conectar() as conn:
                    for av in vincular:
                        upsert_execucao(conn, int(av), int(tarefa_id), None, 0, None, 0, 0, None, 0, STATUS_NAO_INICIADA)
                limpar_cache()
                _toast_sucesso("Alunos vinculados à tarefa com sucesso.")
                st.rerun()


# ─────────────────────────────────────────────
# TELA: AULAS E ASSUNTOS
# ─────────────────────────────────────────────

def tela_aulas():
    st.header("Aulas e assuntos")
    usuario     = aluno_logado()
    disciplinas = disciplinas_ativas()
    aulas       = carregar_aulas()
    assuntos    = carregar_assuntos()
    abas        = st.tabs(["Aulas", "Assuntos"])

    with abas[0]:
        if usuario["perfil"] == "Gestor" and not disciplinas.empty:
            with st.form("nova_aula", clear_on_submit=True):
                col1, col2, col3 = st.columns(3)
                disciplina_id = col1.selectbox("Disciplina", disciplinas["id"].tolist(), format_func=lambda v: disciplinas.loc[disciplinas["id"] == v, "nome"].iloc[0])
                aula          = col2.text_input("Nome da aula")
                tipo_estudo   = col3.selectbox("Tipo de estudo", TIPOS_ESTUDO)
                salvar        = st.form_submit_button("Criar aula")
            if salvar:
                if not limpar_texto(aula):
                    st.error("Informe o nome da aula.")
                else:
                    with conectar() as conn:
                        upsert_aula(conn, disciplina_id, aula, tipo_estudo=tipo_estudo)
                    limpar_cache()
                    _toast_sucesso("Aula criada com sucesso.")
                    st.rerun()

        if aulas.empty:
            st.info("Nenhuma aula cadastrada.")
        else:
            editado = st.data_editor(
                aulas, use_container_width=True, hide_index=True,
                disabled=["id", "disciplina_id", "disciplina"],
                column_config={"tipo_estudo": st.column_config.SelectboxColumn("Tipo de estudo", options=TIPOS_ESTUDO)},
                key="editor_aulas",
            )
            if usuario["perfil"] == "Gestor":
                col1, col2 = st.columns(2)
                if col1.button("Salvar alterações de aulas", type="primary"):
                    with conectar() as conn:
                        for _, row in editado.iterrows():
                            conn.execute(
                                "UPDATE aulas SET aula = ?, tipo_estudo = ?, estudada_padrao = ?, revisao_24h_padrao = ? WHERE id = ?",
                                (limpar_texto(row["aula"]), normalizar_tipo_estudo(row["tipo_estudo"]), limpar_texto(row["estudada_padrao"]), limpar_texto(row["revisao_24h_padrao"]), int(row["id"])),
                            )
                    limpar_cache()
                    _toast_sucesso("Alterações nas aulas salvas com sucesso.")
                    st.rerun()
                excluir = col2.selectbox("Excluir aula", aulas["id"].tolist(), format_func=lambda v: aulas.loc[aulas["id"] == v, "aula"].iloc[0])
                if col2.button("Excluir aula selecionada"):
                    executar("UPDATE aulas SET ativo = 0 WHERE id = ?", (int(excluir),))
                    limpar_cache()
                    _toast_sucesso("Aula excluída com sucesso.")
                    st.rerun()

    with abas[1]:
        if usuario["perfil"] == "Gestor" and not aulas.empty:
            with st.form("novo_assunto", clear_on_submit=True):
                aula_id = st.selectbox(
                    "Aula", aulas["id"].tolist(),
                    format_func=lambda v: f"{aulas.loc[aulas['id'] == v, 'disciplina'].iloc[0]} — {aulas.loc[aulas['id'] == v, 'aula'].iloc[0]}",
                )
                titulo = st.text_input("Assunto")
                salvar = st.form_submit_button("Criar assunto")
            if salvar:
                if not limpar_texto(titulo):
                    st.error("Informe o assunto.")
                else:
                    with conectar() as conn:
                        upsert_assunto(conn, aula_id, titulo)
                    limpar_cache()
                    _toast_sucesso("Assunto criado com sucesso.")
                    st.rerun()

        if assuntos.empty:
            st.info("Nenhum assunto cadastrado.")
        else:
            editado = st.data_editor(
                assuntos, use_container_width=True, hide_index=True,
                disabled=["id", "aula_id", "disciplina", "aula"],
                key="editor_assuntos",
            )
            if usuario["perfil"] == "Gestor":
                col1, col2 = st.columns(2)
                if col1.button("Salvar alterações de assuntos", type="primary"):
                    with conectar() as conn:
                        for _, row in editado.iterrows():
                            conn.execute("UPDATE assuntos SET titulo = ? WHERE id = ?", (limpar_texto(row["assunto"]), int(row["id"])))
                    limpar_cache()
                    _toast_sucesso("Alterações nos assuntos salvas com sucesso.")
                    st.rerun()
                excluir = col2.selectbox("Excluir assunto", assuntos["id"].tolist(), format_func=lambda v: assuntos.loc[assuntos["id"] == v, "assunto"].iloc[0])
                if col2.button("Excluir assunto selecionado"):
                    executar("UPDATE assuntos SET ativo = 0 WHERE id = ?", (int(excluir),))
                    limpar_cache()
                    _toast_sucesso("Assunto excluído com sucesso.")
                    st.rerun()


# ─────────────────────────────────────────────
# TELA: DISCIPLINAS
# ─────────────────────────────────────────────

def tela_disciplinas():
    st.header("Disciplinas")
    usuario = aluno_logado()
    df      = disciplinas_ativas()

    if usuario["perfil"] == "Gestor":
        with st.form("nova_disciplina", clear_on_submit=True):
            nome   = st.text_input("Nome da disciplina")
            salvar = st.form_submit_button("Criar disciplina")
        if salvar:
            try:
                executar("INSERT INTO disciplinas (nome, ativo) VALUES (?, 1)", (limpar_texto(nome),))
                limpar_cache()
                _toast_sucesso("Disciplina criada com sucesso.")
                st.rerun()
            except (IntegrityError, UniqueViolation):
                st.error("Já existe uma disciplina com esse nome.")

        if not df.empty:
            editado = st.data_editor(df, use_container_width=True, hide_index=True, disabled=["id"], key="editor_disciplinas")
            col1, col2 = st.columns(2)
            if col1.button("Salvar disciplinas", type="primary"):
                with st.spinner("Salvando…"):
                 with conectar() as conn:
                    for _, row in editado.iterrows():
                        conn.execute("UPDATE disciplinas SET nome = ? WHERE id = ?", (limpar_texto(row["nome"]), int(row["id"])))
                limpar_cache()
                _toast_sucesso("Alterações nas disciplinas salvas com sucesso.")
                st.rerun()
            excluir = col2.selectbox("Excluir disciplina", df["id"].tolist(), format_func=lambda v: df.loc[df["id"] == v, "nome"].iloc[0])
            if col2.button("Excluir disciplina selecionada"):
                executar("UPDATE disciplinas SET ativo = 0 WHERE id = ?", (int(excluir),))
                limpar_cache()
                _toast_sucesso("Disciplina excluída com sucesso.")
                st.rerun()

    st.dataframe(df, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────
# TELA: ALUNOS
# ─────────────────────────────────────────────

def _excluir_aluno_cascade(aluno_id: int) -> str:
    """
    Remove permanentemente um aluno e todos os dados relacionados:
    execuções, vínculos e o próprio registro na tabela alunos.
    Retorna o nome do aluno excluído.
    """
    with conectar() as conn:
        row = conn.execute("SELECT nome FROM alunos WHERE id = ?", (int(aluno_id),)).fetchone()
        nome = row[0] if row else str(aluno_id)
        # sessoes_estudo não tem FK para alunos: sem este DELETE as sessões
        # ficavam no banco para sempre, apontando para um aluno inexistente.
        conn.execute("DELETE FROM sessoes_estudo WHERE aluno_id = ?", (int(aluno_id),))
        conn.execute("DELETE FROM execucoes WHERE aluno_id = ?", (int(aluno_id),))
        conn.execute("DELETE FROM alunos WHERE id = ? AND perfil = 'Aluno'", (int(aluno_id),))
    limpar_cache()
    return nome


def tela_alunos():
    st.header("Alunos")
    usuario = aluno_logado()
    if usuario["perfil"] != "Gestor":
        st.info("Somente gestores podem administrar alunos.")
        return

    abas = st.tabs(["👥 Alunos ativos", "➕ Novo aluno", "🧹 Limpeza de usuários"])

    # ── Aba 1: Alunos ativos ──
    with abas[0]:
        alunos = alunos_ativos(incluir_gestor=False)
        if alunos.empty:
            st.info("Nenhum aluno cadastrado.")
        else:
            editado = st.data_editor(
                alunos, use_container_width=True, hide_index=True,
                disabled=["id", "perfil"], key="editor_alunos",
            )
            col1, col2 = st.columns([2, 1])
            if col1.button("💾 Salvar alterações", type="primary"):
                try:
                    with st.spinner("Salvando…"):
                     with conectar() as conn:
                        for _, row in editado.iterrows():
                            conn.execute(
                                "UPDATE alunos SET nome = ?, email = ? WHERE id = ? AND perfil = 'Aluno'",
                                (limpar_texto(row["nome"]), normalizar_email(row["email"]), int(row["id"])),
                            )
                    limpar_cache()
                    _toast_sucesso("Dados dos alunos atualizados com sucesso.")
                    st.rerun()
                except (IntegrityError, UniqueViolation):
                    st.error("Nome ou e-mail duplicado.")

            st.markdown("---")
            render_html('<div class="section-title">🗑️ Excluir aluno</div>')
            render_html(
                '<div class="rule-warning">'
                '⚠️ A exclusão remove <strong>permanentemente</strong> o aluno e todo o seu histórico '
                '(execuções, status, horas, acertos). Esta ação <strong>não pode ser desfeita</strong>.'
                '</div>'
            )
            excluir_id = st.selectbox(
                "Selecione o aluno a excluir",
                alunos["id"].tolist(),
                format_func=lambda v: f"{alunos.loc[alunos['id']==v,'nome'].iloc[0]} ({alunos.loc[alunos['id']==v,'email'].iloc[0]})",
                key="sel_excluir_aluno",
            )
            nome_excluir = alunos.loc[alunos["id"] == excluir_id, "nome"].iloc[0]
            confirmar = st.checkbox(
                f"Confirmo que desejo excluir permanentemente o aluno **{nome_excluir}** e todos os seus dados.",
                key="chk_excluir_aluno",
            )
            if st.button("🗑️ Excluir aluno", type="primary", disabled=not confirmar):
                nome_removido = _excluir_aluno_cascade(excluir_id)
                _toast_sucesso(f"Aluno '{nome_removido}' e todos os seus dados foram removidos com sucesso.")
                st.rerun()

    # ── Aba 2: Novo aluno ──
    with abas[1]:
        with st.form("novo_aluno", clear_on_submit=True):
            col1, col2 = st.columns(2)
            nome  = col1.text_input("Nome")
            email = col2.text_input("E-mail (opcional)")
            salvar = st.form_submit_button("➕ Adicionar aluno", use_container_width=True)
        if salvar:
            nome  = limpar_texto(nome)
            email = normalizar_email(email) or email_local(nome)
            if not nome:
                st.error("Informe o nome do aluno.")
            else:
                try:
                    executar(
                        "INSERT INTO alunos (nome, email, senha, perfil, ativo, force_troca_senha) VALUES (?, ?, ?, 'Aluno', 1, 1)",
                        (nome, email, hash_senha("123")),
                    )
                    limpar_cache()
                    _toast_sucesso(f"Aluno '{nome}' cadastrado com sucesso. Senha inicial: 123")
                    st.rerun()
                except (IntegrityError, UniqueViolation):
                    st.error("Já existe um aluno com esse nome ou e-mail.")

    # ── Aba 3: Limpeza de usuários ──
    with abas[2]:
        render_html(
            '<div class="hero" style="margin-bottom:16px">'
            '<h1 style="font-size:1.2rem">🧹 Limpeza de usuários</h1>'
            '<p>Identifique e remova usuários duplicados, sem execuções ou indesejados. '
            'Use os filtros abaixo para selecionar quem remover, confirme e execute a limpeza.</p>'
            '</div>'
        )

        # Carrega TODOS os alunos (inclusive inativos) com contagem de execuções
        todos = consultar("""
            SELECT
                a.id,
                a.nome,
                a.email,
                a.perfil,
                a.ativo,
                COUNT(e.id) AS qtd_execucoes,
                MAX(e.atualizado_em) AS ultima_atividade
            FROM alunos a
            LEFT JOIN execucoes e ON e.aluno_id = a.id
            WHERE a.perfil = 'Aluno'
            GROUP BY a.id, a.nome, a.email, a.perfil, a.ativo
            ORDER BY a.ativo DESC, qtd_execucoes ASC, a.nome
        """)

        if todos.empty:
            st.info("Nenhum aluno cadastrado.")
        else:
            # Métricas rápidas
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total de alunos", len(todos))
            c2.metric("Ativos", int(todos["ativo"].sum()))
            c3.metric("Inativos", int((todos["ativo"] == 0).sum()))
            c4.metric("Sem nenhuma execução", int((todos["qtd_execucoes"] == 0).sum()))

            st.markdown("---")

            # Filtros de seleção para limpeza
            col_f1, col_f2 = st.columns(2)
            mostrar_sem_exec = col_f1.checkbox("Mostrar apenas alunos sem execuções", value=False)
            mostrar_inativos = col_f2.checkbox("Mostrar apenas alunos inativos", value=False)

            df_view = todos.copy()
            if mostrar_sem_exec:
                df_view = df_view[df_view["qtd_execucoes"] == 0]
            if mostrar_inativos:
                df_view = df_view[df_view["ativo"] == 0]

            if df_view.empty:
                st.info("Nenhum aluno encontrado com os filtros aplicados.")
            else:
                st.dataframe(
                    df_view.rename(columns={
                        "id": "ID", "nome": "Nome", "email": "E-mail",
                        "perfil": "Perfil", "ativo": "Ativo",
                        "qtd_execucoes": "Execuções", "ultima_atividade": "Última atividade",
                    }),
                    use_container_width=True, hide_index=True,
                )

                st.markdown("---")
                render_html('<div class="section-title">🗑️ Remover selecionados</div>')
                render_html(
                    '<div class="rule-warning">'
                    '⚠️ A remoção exclui permanentemente o aluno e todos os seus dados de execução. '
                    'Esta ação <strong>não pode ser desfeita</strong>. Revise a lista acima antes de confirmar.'
                    '</div>'
                )

                ids_disponiveis = df_view["id"].tolist()
                nomes_map = dict(zip(df_view["id"], df_view["nome"]))

                ids_remover = st.multiselect(
                    "Selecione os alunos a remover",
                    ids_disponiveis,
                    format_func=lambda v: f"{nomes_map.get(v,'?')} — {int(df_view.loc[df_view['id']==v,'qtd_execucoes'].iloc[0])} execuções",
                    key="ms_limpar_alunos",
                )

                if ids_remover:
                    nomes_selecionados = [nomes_map.get(i, str(i)) for i in ids_remover]
                    confirmar_limpeza = st.checkbox(
                        f"Confirmo a remoção de {len(ids_remover)} aluno(s): {', '.join(nomes_selecionados)}",
                        key="chk_limpar_alunos",
                    )
                    if st.button("🧹 Executar limpeza", type="primary", disabled=not confirmar_limpeza):
                        removidos = []
                        erros = []
                        for aid in ids_remover:
                            try:
                                nome_rem = _excluir_aluno_cascade(aid)
                                removidos.append(nome_rem)
                            except Exception as exc:
                                erros.append(f"{nomes_map.get(aid,'?')}: {exc}")
                        if removidos:
                            _toast_sucesso(f"{len(removidos)} aluno(s) removido(s) com sucesso: {', '.join(removidos)}.")
                        if erros:
                            for e in erros:
                                st.error(f"Erro ao remover: {e}")
                        st.rerun()

                st.markdown("---")
                render_html('<div class="section-title">⚡ Limpeza rápida</div>')
                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    st.caption("Remove todos os alunos **sem nenhuma execução** registrada.")
                    sem_exec_ids = todos[todos["qtd_execucoes"] == 0]["id"].tolist()
                    if sem_exec_ids:
                        conf_sem_exec = st.checkbox(
                            f"Confirmo remoção de {len(sem_exec_ids)} aluno(s) sem execuções",
                            key="chk_sem_exec",
                        )
                        if st.button("🗑️ Remover sem execuções", disabled=not conf_sem_exec, key="btn_sem_exec"):
                            for aid in sem_exec_ids:
                                _excluir_aluno_cascade(aid)
                            _toast_sucesso(f"{len(sem_exec_ids)} aluno(s) sem execuções removidos com sucesso.")
                            st.rerun()
                    else:
                        st.success("Nenhum aluno sem execuções.")

                with col_r2:
                    st.caption("Remove todos os alunos com status **inativo** (`ativo = 0`).")
                    inativos_ids = todos[todos["ativo"] == 0]["id"].tolist()
                    if inativos_ids:
                        conf_inativos = st.checkbox(
                            f"Confirmo remoção de {len(inativos_ids)} aluno(s) inativo(s)",
                            key="chk_inativos",
                        )
                        if st.button("🗑️ Remover inativos", disabled=not conf_inativos, key="btn_inativos"):
                            for aid in inativos_ids:
                                _excluir_aluno_cascade(aid)
                            _toast_sucesso(f"{len(inativos_ids)} aluno(s) inativo(s) removidos com sucesso.")
                            st.rerun()
                    else:
                        st.success("Nenhum aluno inativo.")


# ─────────────────────────────────────────────
# TELA: IMPORTAÇÕES
# ─────────────────────────────────────────────

def tela_importacao():
    st.header("Importações")

    abas = st.tabs([
        "📋 Planilha de referência",
        "📊 Ciclo Consolidado (multi-aluno)",
    ])

    # ── Aba 1: Planilha de referência ──
    with abas[0]:
        render_html("""
            <div class="insight-card info" style="margin-bottom:12px">
              <div class="insight-icon">🛡️</div>
              <div class="insight-body">
                <div class="insight-title">Importação segura — histórico preservado</div>
                <p class="insight-text">
                  A importação da planilha de referência <strong>nunca apaga execuções existentes</strong>.
                  Apenas insere novas disciplinas, aulas, assuntos e tarefas que ainda não existam,
                  ou atualiza campos de estrutura (trilha, tipo, exercícios previstos).
                  Status, datas, horas e acertos já registrados são sempre preservados.
                </p>
              </div>
            </div>
        """)
        st.caption(f"Planilha de referência padrão: `{PLANILHA_REFERENCIA}`")
        arquivo = st.file_uploader("Planilha referência (.xlsx)", type=["xlsx"], key="ref_upload")
        if arquivo is not None:
            destino = BASE_DIR / "planilha_referencia_importada.xlsx"
            destino.write_bytes(arquivo.getbuffer())
            if st.button("Importar referência enviada", type="primary", key="ref_envio"):
                with st.spinner("Importando planilha de referência…"):
                    ok = importar_planilha_referencia(destino)
                if ok:
                    _toast_sucesso("Planilha de referência importada com sucesso. Histórico preservado.")
                    st.rerun()
        elif st.button("Reimportar referência padrão", key="ref_padrao"):
            with st.spinner("Importando planilha padrão…"):
                ok = importar_planilha_referencia(PLANILHA_REFERENCIA)
            if ok:
                _toast_sucesso("Planilha de referência padrão importada com sucesso. Histórico preservado.")
                st.rerun()

        st.markdown("---")
        render_html(
            '<div class="insight-card info" style="margin-bottom:10px">'
            '<div class="insight-icon">🔗</div>'
            '<div class="insight-body">'
            '<div class="insight-title">Vincular tarefas pendentes aos alunos</div>'
            '<p class="insight-text">'
            'Se o dashboard não exibe tarefas novas da planilha, use este botão. '
            'Ele cria vínculos <strong>Não iniciada</strong> para tarefas que ainda não '
            'aparecem no acompanhamento de nenhum aluno. '
            '<strong>Não altera nenhum registro existente</strong> — histórico de execuções preservado.'
            '</p></div></div>'
        )
        if st.button("🔗 Vincular tarefas pendentes aos alunos", key="btn_vincular_pendentes"):
            with st.spinner("Verificando e vinculando tarefas pendentes…"):
                novos = vincular_tarefas_pendentes()
            if novos > 0:
                _toast_sucesso(f"✅ {novos} novo(s) vínculo(s) criado(s). O dashboard já reflete as novas tarefas.")
            else:
                st.info("Nenhuma tarefa pendente encontrada — todos os vínculos já existem.")

        st.markdown("---")
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Estudantes",  int(consultar("SELECT COUNT(*) qtd FROM alunos WHERE perfil='Aluno' AND ativo=1").iloc[0].qtd))
        m2.metric("Disciplinas", int(consultar("SELECT COUNT(*) qtd FROM disciplinas WHERE ativo=1").iloc[0].qtd))
        m3.metric("Aulas",       int(consultar("SELECT COUNT(*) qtd FROM aulas WHERE ativo=1").iloc[0].qtd))
        m4.metric("Tarefas",     int(consultar("SELECT COUNT(*) qtd FROM tarefas WHERE ativo=1").iloc[0].qtd))
        m5.metric("Execuções",   int(consultar("SELECT COUNT(*) qtd FROM execucoes").iloc[0].qtd))

    # ── Aba 2: Ciclo Consolidado ──
    with abas[1]:
        render_html("""
            <div class="insight-card info" style="margin-bottom:14px">
              <div class="insight-icon">ℹ️</div>
              <div class="insight-body">
                <div class="insight-title">Formato esperado: Ciclo Consolidado</div>
                <p class="insight-text">
                  Arquivo com aba <strong>CICLO_CONSOLIDADO</strong>, linha de cabeçalho na linha 3.<br>
                  Colunas: <em>BLOCO · DISCIPLINA · OBJETIVO · STATUS · DATA PROGRAMADA ·
                  CH/AULA/QUESTÕES/ACERTOS (Aluno A) · CH/AULA/QUESTÕES/ACERTOS (Aluno B) ·
                  CH TOTAL · TOTAL QUESTÕES · TOTAL ACERTOS · DESEMPENHO</em>.<br>
                  Os nomes dos alunos são lidos automaticamente da linha 2 do arquivo.
                </p>
              </div>
            </div>
        """)

        arquivo_cc = st.file_uploader(
            "Arquivo Ciclo Consolidado (.xlsx)",
            type=["xlsx"],
            key="cc_upload",
        )

        col_modo, col_btn = st.columns([2, 1])
        # "acumular" é o padrão: "substituir" apaga TODO o histórico dos dois
        # alunos do arquivo e antes vinha pré-selecionado, a um clique de distância.
        modo = col_modo.radio(
            "Modo de importação",
            ["acumular", "substituir"],
            format_func=lambda v: "➕ Acumular (não apaga execuções anteriores)" if v == "acumular" else "🔄 Substituir — APAGA o histórico dos alunos do arquivo",
            horizontal=True,
            key="cc_modo",
        )

        confirma_subst = True
        if modo == "substituir":
            try:
                resumo = consultar(
                    """SELECT COUNT(*) AS execucoes,
                              COALESCE(SUM(ch_efetiva), 0) AS horas,
                              COALESCE(SUM(qtd_questoes_feitas), 0) AS questoes
                       FROM execucoes
                       WHERE ch_efetiva > 0 OR qtd_questoes_feitas > 0
                          OR status <> 'NAO_INICIADA'"""
                ).iloc[0]
                n_sess = consultar("SELECT COUNT(*) AS n FROM sessoes_estudo").iloc[0]["n"]
                render_html(
                    '<div class="rule-warning" style="margin-bottom:10px">'
                    '⚠️ <strong>Operação destrutiva.</strong> O modo Substituir apaga as execuções '
                    'e as sessões de estudo dos alunos presentes no arquivo. '
                    f'Hoje o banco tem <strong>{int(resumo["execucoes"])}</strong> execução(ões) com estudo, '
                    f'<strong>{horas_para_hm(float(resumo["horas"]))}</strong>, '
                    f'<strong>{int(resumo["questoes"])}</strong> questões e '
                    f'<strong>{int(n_sess)}</strong> sessão(ões) registradas.'
                    '</div>'
                )
            except Exception:
                render_html(
                    '<div class="rule-warning" style="margin-bottom:10px">'
                    '⚠️ <strong>Operação destrutiva.</strong> O modo Substituir apaga execuções '
                    'e sessões de estudo dos alunos presentes no arquivo.</div>'
                )
            confirma_subst = st.checkbox(
                "Confirmo que quero apagar o histórico desses alunos antes de importar",
                value=False, key="cc_confirma_subst",
            )

        if arquivo_cc is not None:
            destino_cc = BASE_DIR / arquivo_cc.name
            destino_cc.write_bytes(arquivo_cc.getbuffer())

            # Pré-visualização: mostra nomes detectados
            try:
                import openpyxl as _opx
                _wb = _opx.load_workbook(str(destino_cc), read_only=True, data_only=True)
                if "CICLO_CONSOLIDADO" in _wb.sheetnames:
                    _linhas = list(_wb["CICLO_CONSOLIDADO"].iter_rows(values_only=True, max_row=3))
                    _nome_a = limpar_texto(_linhas[1][6]) if len(_linhas) > 1 and len(_linhas[1]) > 6 else "?"
                    _nome_b = limpar_texto(_linhas[1][10]) if len(_linhas) > 1 and len(_linhas[1]) > 10 else "?"
                    st.info(f"Alunos detectados: **{_nome_a}** e **{_nome_b}**")
            except Exception:
                pass

            if col_btn.button("▶ Importar", type="primary", use_container_width=True, key="cc_importar"):
                if modo == "substituir" and not confirma_subst:
                    render_html(
                        '<div class="rule-error">❌ Marque a confirmação para usar o modo Substituir, '
                        'ou troque para Acumular.</div>'
                    )
                    st.stop()
                with st.spinner("Importando Ciclo Consolidado… aguarde."):
                    resultado = importar_ciclo_consolidado(destino_cc, modo=modo)

                if resultado["erros"]:
                    for e in resultado["erros"]:
                        st.error(e)
                if resultado["avisos"]:
                    with st.expander(f"⚠️ {len(resultado['avisos'])} aviso(s)"):
                        for av in resultado["avisos"]:
                            st.warning(av)
                if resultado["ok"]:
                    _toast_sucesso(
                        f"✅ Importação concluída! "
                        f"**{resultado['registros']}** execuções gravadas."
                    )
                    st.rerun()


# ─────────────────────────────────────────────
# TELA: CONFIGURAÇÕES
# ─────────────────────────────────────────────

def tela_configuracoes():
    st.header("Configurações")
    usuario = aluno_logado()
    aba_senha, aba_usuarios, aba_backup = st.tabs(["Minha senha", "Usuários", "Backup"])

    with aba_senha:
        with st.form("senha"):
            atual     = st.text_input("Senha atual", type="password")
            nova      = st.text_input("Nova senha", type="password")
            confirmar = st.text_input("Confirmar nova senha", type="password")
            salvar    = st.form_submit_button("Alterar senha")
        if salvar:
            usuario_db = consultar("SELECT senha FROM alunos WHERE id = ?", (usuario["id"],))
            if usuario_db.empty or not verificar_senha(atual, usuario_db.iloc[0]["senha"]):
                st.error("Senha atual inválida.")
            elif not senha_valida(nova):
                st.error("Use uma senha com pelo menos 8 caracteres, contendo letras e números.")
            elif nova != confirmar:
                st.error("A confirmação não confere.")
            else:
                atualizar_senha_usuario(usuario["id"], nova, 0)
                _toast_sucesso("Senha alterada com sucesso.")

    with aba_usuarios:
        if usuario["perfil"] != "Gestor":
            st.info("Somente gestores podem administrar senhas e bloqueios.")
        else:
            usuarios = consultar("SELECT id, nome, email, perfil, ativo, force_troca_senha FROM alunos ORDER BY perfil, nome")
            st.dataframe(usuarios, use_container_width=True, hide_index=True)
            if not usuarios.empty:
                usuario_id = st.selectbox(
                    "Usuário", usuarios["id"].tolist(),
                    format_func=lambda v: f"{usuarios.loc[usuarios['id'] == v, 'nome'].iloc[0]} ({usuarios.loc[usuarios['id'] == v, 'perfil'].iloc[0]})",
                )
                col1, col2, col3 = st.columns(3)
                nova_senha = col1.text_input("Nova senha temporária", type="password")
                forcar     = col2.checkbox("Forçar troca no próximo login", value=True)
                if col3.button("Redefinir senha", type="primary"):
                    if not senha_valida(nova_senha):
                        st.error("A senha temporária deve ter pelo menos 8 caracteres, contendo letras e números.")
                    else:
                        atualizar_senha_usuario(usuario_id, nova_senha, 1 if forcar else 0)
                        _toast_sucesso("Senha redefinida com sucesso. O aluno deverá trocar no próximo acesso.")
                        st.rerun()
                ativo_atual = int(usuarios.loc[usuarios["id"] == usuario_id, "ativo"].iloc[0])
                col4, col5 = st.columns(2)
                if ativo_atual == 1 and col4.button("Bloquear usuário"):
                    executar("UPDATE alunos SET ativo = 0 WHERE id = ?", (int(usuario_id),))
                    _toast_sucesso("Usuário bloqueado com sucesso.")
                    st.rerun()
                if ativo_atual == 0 and col5.button("Reativar usuário"):
                    executar("UPDATE alunos SET ativo = 1, force_troca_senha = 1 WHERE id = ?", (int(usuario_id),))
                    _toast_sucesso("Usuário reativado. Será solicitado troca de senha no próximo acesso.")
                    st.rerun()

    with aba_backup:
        if usuario["perfil"] != "Gestor":
            st.info("Somente gestores podem exportar backups.")
        else:
            st.caption("Exportação lógica para backup e auditoria.")
            tabelas = {
                "alunos": consultar("SELECT id, nome, email, perfil, ativo, force_troca_senha FROM alunos"),
                "disciplinas": consultar("SELECT * FROM disciplinas"),
                "aulas": consultar("SELECT * FROM aulas"),
                "assuntos": consultar("SELECT * FROM assuntos"),
                "tarefas": consultar("SELECT * FROM tarefas"),
                "execucoes": consultar("SELECT * FROM execucoes"),
            }
            for nome_tab, dados in tabelas.items():
                st.download_button(
                    f"Exportar {nome_tab}.csv",
                    data=dados.to_csv(index=False).encode("utf-8"),
                    file_name=f"{nome_tab}_{date.today()}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )


# ─────────────────────────────────────────────
# INICIALIZAÇÃO E MAIN
# ─────────────────────────────────────────────

def inicializar():
    criar_tabelas()
    inserir_admin()
    qtd = consultar("SELECT COUNT(*) AS qtd FROM tarefas").iloc[0]["qtd"]
    if qtd == 0 and PLANILHA_REFERENCIA.exists():
        importar_planilha_referencia(PLANILHA_REFERENCIA, substituir=True)


def main():
    inicializar()

    if "usuario" not in st.session_state:
        tela_login()
        return

    usuario = aluno_logado()
    if int(usuario.get("force_troca_senha") or 0) == 1:
        tela_troca_obrigatoria()
        return

    # ── Exibe banner de sucesso persistente (sobrevive ao st.rerun) ──
    _render_banner_sucesso()

    st.sidebar.title(APP_NAME)
    st.sidebar.write(f"**Usuário:** {usuario['nome']}")
    st.sidebar.write(f"**Perfil:** {usuario['perfil']}")
    if st.sidebar.button("Sair", use_container_width=True):
        st.session_state.clear()
        st.rerun()

    if usuario["perfil"] == "Gestor":
        opcoes = ["Dashboard", "Registro rápido", "Tarefas", "Aulas e assuntos", "Disciplinas", "Alunos", "Importações", "Configurações"]
    else:
        opcoes = ["Dashboard", "Registro rápido", "Tarefas", "Aulas e assuntos", "Configurações"]

    opcao = st.sidebar.radio("Navegação", opcoes)
    paginas = {
        "Dashboard":       dashboard,
        "Registro rápido": tela_registro_rapido,
        "Tarefas":         tela_tarefas,
        "Aulas e assuntos":tela_aulas,
        "Disciplinas":     tela_disciplinas,
        "Alunos":          tela_alunos,
        "Importações":     tela_importacao,
        "Configurações":   tela_configuracoes,
    }
    paginas[opcao]()


if __name__ == "__main__":
    main()
